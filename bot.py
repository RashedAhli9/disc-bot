# ============================================================
# FLASK KEEPALIVE (STARTS IMMEDIATELY – KOYEB SAFE
# ===========================================================

from flask import Flask
import threading
import os

app = Flask(__name__)

@app.route("/")
def home():
    return "OK", 200

@app.route("/health")
def health():
    log_info("[HEALTH CHECK] ping received")
    return "healthy", 200


def run_flask():
    port = int(os.environ.get("PORT", 8000))
    app.run(host="0.0.0.0", port=port)

threading.Thread(target=run_flask, daemon=True).start()

# ============================================================
# DISCORD + SYSTEM IMPORTS
# ============================================================

import discord
from discord.ext import commands, tasks
from discord import app_commands
import json
import sqlite3
import pytz
from datetime import datetime, timedelta, date, time
from discord.ui import View, Select, Button, Modal, TextInput
import io
import zipfile
import asyncio
import anthropic
import matplotlib
matplotlib.use("Agg")  # headless backend — no display needed on a server
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import aiohttp
import openpyxl
import io

# ============================================================
# LOGGING SYSTEM
# ============================================================

import logging

logging.basicConfig(
    level=logging.INFO,
    format='[%(asctime)s] [%(levelname)s] %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# ============================================================
# GLOBAL CONFIG
# ============================================================

MY_TIMEZONE = "UTC"
channel_id = 1328658110897983549
update_channel_id = 1332676174995918859
OWNER_ID = 1084884048884797490
ROLE_ID = 1413532222396301322
BACKUP_CHANNEL_ID = 1444604637377204295
ABYSS_ROLE_ID = 1413532222396301322
EVENT_ANNOUNCEMENT_ROLE_ID = 1412464184746053693
ABYSS_CONFIG_FILE = "abyss_config.json"
DB = "/data/events.db"
DB_PROGRESS = "/data/season_progress.db"  # Separate database for seasonal data

# Ensure the data directory exists so sqlite3.connect() doesn't crash-loop
# if the persistent volume was momentarily unmounted or missing.
os.makedirs(os.path.dirname(DB), exist_ok=True)

# Anthropic API — powers !ask's natural-language understanding when set.
# Falls back to keyword-only routing if this env var isn't configured on Railway.
ANTHROPIC_API_KEY = os.getenv("ANTHROPIC_API_KEY")
_anthropic_client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY) if ANTHROPIC_API_KEY else None

# ============================================================
# CALL OF STATS ACCOUNT IDS & CONSTANTS
# ============================================================

REKZ_ACCOUNT_ID = "16322115"
FALLBACK_DAYS = 6  # Days to fallback when data is empty

# Discord ID -> Call of Stats Account ID mapping
# For members who may not have numeric roles in server
DISCORD_TO_ACCOUNT_ID = {
    1244330800019804180: "7979635",  # Havi
}

# ============================================================
# USERNAME LOOKUP (Map Discord usernames to Discord IDs)
# ============================================================
# Hardcoded mapping - add your members here
# Username: Discord ID
USERNAME_TO_DISCORD_ID = {
    "rekz": 1084884048884797490,
    "truvix": 797778630025019402,
    "azrael": 663715561951461376,
    "drakken": 401088806432014347,
    "gato": 937458459115413565,
    "truffles": 1285424051761713155,
    "havi": 1244330800019804180,
}


# ============================================================
# CACHE SETTINGS
# ============================================================

CACHE_EXPIRY_HOURS = 72  # 3 days - cache validity period

# ============================================================
# LOGGING SETUP (Simple alternative to print)
# ============================================================

import logging
logging.basicConfig(
    level=logging.INFO,
    format='[%(levelname)s] %(message)s'
)
logger = logging.getLogger(__name__)

def log_info(message):
    """Log info message"""
    logger.info(message)

def log_error(message):
    """Log error message"""
    logger.error(message)

def log_debug(message):
    """Log debug message"""
    logger.debug(message)
    
BACKUP_DIR = "backups"
MAX_BACKUPS = 10
os.makedirs(BACKUP_DIR, exist_ok=True)

# ============================================================
# REQUIRED HELPERS (FIXES ALL CRASHES)
# ============================================================

def has_admin(inter):
    if inter.user.id == OWNER_ID:
        return True
    return any(r.permissions.administrator for r in inter.user.roles)

def parse_datetime(input_str):
    input_str = input_str.strip()
    now = datetime.utcnow()

    # Case 1: full date + HH:MM
    try:
        return datetime.strptime(input_str, "%d-%m-%Y %H:%M")
    except Exception as e:
        pass

    # Case 2: full date + HHutc
    if "-" in input_str and "utc" in input_str:
        date_part, time_part = input_str.split()
        base_date = datetime.strptime(date_part, "%d-%m-%Y")
        hour = int(time_part.replace("utc", ""))
        return base_date.replace(
            hour=hour, minute=0, second=0, microsecond=0
        )

    # Relative parsing
    d = h = m = 0
    for part in input_str.lower().split():
        if part.endswith("d"):
            d = int(part[:-1])
        elif part.endswith("h"):
            h = int(part[:-1])
        elif part.endswith("m"):
            m = int(part[:-1])
        elif part.endswith("utc"):
            hour = int(part.replace("utc", ""))
            return now.replace(
                hour=hour, minute=0, second=0, microsecond=0
            )

    return now + timedelta(days=d, hours=h, minutes=m)


def day_name(i):
    return ["Mon","Tue","Wed","Thu","Fri","Sat","Sun"][i]

def pretty_days(days):
    return ", ".join(day_name(d) for d in sorted(days))

def pretty_hours(hours):
    return ", ".join(f"{h:02}:00" for h in sorted(hours))


async def dm_abyss_role(guild: discord.Guild, embed: discord.Embed):
    """
    Send a DM to all members with the Abyss reminder role.
    
    Args:
        guild: Discord guild/server
        embed: Discord embed to send
    
    Returns:
        None (logs success/errors)
    """
    role = guild.get_role(ABYSS_ROLE_ID)
    if not role:
        log_info(f"[ABYSS] Role not found in guild {guild.id}")
        return

    count = 0
    async for member in guild.fetch_members(limit=None):
        if role in member.roles:
            try:
                dm = await member.create_dm()
                await dm.send(embed=embed)
                count += 1
                # Rate limiting - wait 0.5 seconds between DMs to prevent hitting rate limits
                await asyncio.sleep(0.5)
            except discord.Forbidden:
                continue
            except Exception as e:
                log_info(f"[ABYSS] DM error: {e}")
    
    log_info(f"[ABYSS] Sent DM to {count} members with role {ABYSS_ROLE_ID}")



# ============================================================
# BACKUP SYSTEM
# ============================================================

def cleanup_old_backups():
    """
    Remove old backup files, keeping only the MAX_BACKUPS most recent.
    
    Returns:
        None
    """
    files = sorted(
        [f for f in os.listdir(BACKUP_DIR) if f.endswith(".zip")],
        reverse=True
    )
    for f in files[MAX_BACKUPS:]:
        try:
            os.remove(os.path.join(BACKUP_DIR, f))
            log_info(f"Deleted old backup: {f}")
        except Exception as e:
            log_error(f"Failed to delete backup {f}: {e}")

async def upload_backup(path):
    """
    Upload a backup file to the backup channel.
    
    Args:
        path: Path to the backup zip file
    
    Returns:
        None
    """
    try:
        ch = bot.get_channel(BACKUP_CHANNEL_ID)
        if ch:
            await ch.send("📦 **Backup created**", file=discord.File(path))
    except Exception as e:
        log_error(f"[UPLOAD BACKUP] Error: {e}")

def make_backup():
    ts = datetime.utcnow().strftime("%Y-%m-%d_%H-%M-%S")
    path = os.path.join(BACKUP_DIR, f"backup_{ts}.zip")

    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as z:
        if os.path.exists(DB):
            z.write(DB)
        if os.path.exists(ABYSS_CONFIG_FILE):
            z.write(ABYSS_CONFIG_FILE)

    cleanup_old_backups()

    if bot.loop.is_running():
        bot.loop.create_task(upload_backup(path))

def silent_backup():
    try:
        make_backup()
    except Exception as e:
        pass

def get_all_lords_from_guild(guild):
    """Get all numeric roles (account IDs) from the guild"""
    lords = []
    
    # Scan ALL roles in the server
    for role in guild.roles:
        if role.name.isdigit() and role.name != "@everyone":
            lords.append({
                "name": role.name,
                "account_id": role.name,
                "role": role
            })
    
    return lords

# ============================================================
# CALLOFSTATS CACHE SYSTEM
# ============================================================

_stats_cache = {}
CACHE_EXPIRY_HOURS = 72  # 3 days
CACHE_DURATION = 600  # 10 minutes - for in-memory cache of fetched stats

async def fetch_alliance_tag(account_id):
    """Fetch just the alliance tag from Call of Stats"""
    import re
    log_info(f"[ALLIANCE TAG] Fetching for {account_id}...")
    try:
        async with aiohttp.ClientSession() as session:
            url = f"https://www.callofstats.com/lord/{account_id}"
            async with session.get(url, timeout=aiohttp.ClientTimeout(total=30)) as resp:
                if resp.status == 200:
                    html = await resp.text()
                    
                    # Debug: find where alliance tag might be
                    if "higher-value" in html:
                        # Find the section with higher-value
                        idx = html.find("higher-value")
                        debug_section = html[max(0, idx-100):min(len(html), idx+300)]
                        log_info(f"[ALLIANCE TAG DEBUG] HTML around 'higher-value': {debug_section}")
                    
                    # Extract alliance tag from <h2 class="higher-value">[TAG]</h2>
                    match = re.search(r'<h2 class="higher-value">([^<]+)</h2>', html)
                    if match:
                        tag = match.group(1).strip()
                        log_info(f"[ALLIANCE TAG] Found for {account_id}: {tag}")
                        return tag
                    else:
                        log_info(f"[ALLIANCE TAG] No match found in HTML for {account_id}")
    except Exception as e:
        log_info(f"[ALLIANCE TAG] Error fetching for {account_id}: {e}")
    
    log_info(f"[ALLIANCE TAG] Returning empty string for {account_id}")
    return ""


async def fetch_highest_power(account_id):
    """
    Fetch the HIGHEST POWER from the normal profile page (no date range).
    Uses the authenticated global session (required - page needs login).
    Retries up to 3 times, re-authenticating if redirected to login page.
    Returns the highest power value as int, or None if not found.
    """
    import re

    url = f"https://callofstats.com/lord/{account_id}"
    patterns = [
        r'<span class="subtle">Highest Power</span>\s*<div class="value">\+?([0-9,]+)</div>',
        r'Highest Power</span>.*?<div class="value">\+?([0-9,]+)</div>',
        r'Highest Power.*?\+?([0-9,]+)',
    ]

    for attempt in range(3):
        try:
            session = await get_callofstats_session()
            if not session:
                log_info(f"[HIGHEST POWER] No session available, attempt {attempt+1}")
                await asyncio.sleep(2)
                continue

            async with session.get(url, allow_redirects=True) as response:
                if response.status != 200:
                    log_info(f"[HIGHEST POWER] HTTP {response.status} attempt {attempt+1} for {account_id}")
                    await asyncio.sleep(1)
                    continue

                html = await response.text()

                # Detect login redirect
                if "<title>Login" in html or "Sign in to Call of Stats" in html:
                    log_info(f"[HIGHEST POWER] Got login page on attempt {attempt+1}, forcing session refresh")
                    global _callofstats_session, _session_login_time
                    if _callofstats_session:
                        await _callofstats_session.close()
                        _callofstats_session = None
                        _session_login_time = None
                    await asyncio.sleep(2)
                    continue

                for i, pattern in enumerate(patterns):
                    match = re.search(pattern, html, re.DOTALL)
                    if match:
                        power_str = match.group(1).replace(",", "")
                        try:
                            highest_power = int(power_str)
                            log_info(f"[HIGHEST POWER] {account_id} = {highest_power:,} (pattern {i}, attempt {attempt+1})")
                            return highest_power
                        except Exception as e:
                            log_info(f"[HIGHEST POWER] Parse error: {e}")

                if attempt == 2:
                    idx = html.find("Highest Power")
                    if idx != -1:
                        log_info(f"[HIGHEST POWER DEBUG] HTML around 'Highest Power':\n{html[max(0,idx-100):idx+300]}")
                    else:
                        log_info(f"[HIGHEST POWER] Not found for {account_id}")
                        log_info(f"[HIGHEST POWER DEBUG] First 300 chars: {html[:300]}")
                else:
                    log_info(f"[HIGHEST POWER] No match attempt {attempt+1} for {account_id}, retrying...")
                    await asyncio.sleep(1)

        except asyncio.TimeoutError:
            log_info(f"[HIGHEST POWER] Timeout attempt {attempt+1} for {account_id}")
            await asyncio.sleep(1)
        except Exception as e:
            log_info(f"[HIGHEST POWER ERROR] attempt {attempt+1} for {account_id}: {e}")
            await asyncio.sleep(1)

    return None

async def fetch_highest_power_at_date(account_id, selected_date):
    """
    Fetch the HIGHEST POWER as it stood on a SPECIFIC date (selected_date, no range).
    Used for old/ended seasons instead of fetch_highest_power(), which always returns
    today's live value — for a past season we want the power at that season's end date.
    Returns the highest power value as int, or None if not found.
    """
    import re

    url = f"https://callofstats.com/lord/{account_id}?selected_date={selected_date}"
    patterns = [
        r'<span class="subtle">Highest Power</span>\s*<div class="value">\+?([0-9,]+)</div>',
        r'Highest Power</span>.*?<div class="value">\+?([0-9,]+)</div>',
        r'Highest Power.*?\+?([0-9,]+)',
    ]

    for attempt in range(3):
        try:
            session = await get_callofstats_session()
            if not session:
                log_info(f"[HIGHEST POWER AT DATE] No session available, attempt {attempt+1}")
                await asyncio.sleep(2)
                continue

            async with session.get(url, allow_redirects=True) as response:
                if response.status != 200:
                    log_info(f"[HIGHEST POWER AT DATE] HTTP {response.status} attempt {attempt+1} for {account_id} @ {selected_date}")
                    await asyncio.sleep(1)
                    continue

                html = await response.text()

                if "<title>Login" in html or "Sign in to Call of Stats" in html:
                    log_info(f"[HIGHEST POWER AT DATE] Got login page on attempt {attempt+1}, forcing session refresh")
                    global _callofstats_session, _session_login_time
                    if _callofstats_session:
                        await _callofstats_session.close()
                        _callofstats_session = None
                        _session_login_time = None
                    await asyncio.sleep(2)
                    continue

                for i, pattern in enumerate(patterns):
                    match = re.search(pattern, html, re.DOTALL)
                    if match:
                        power_str = match.group(1).replace(",", "")
                        try:
                            highest_power = int(power_str)
                            log_info(f"[HIGHEST POWER AT DATE] {account_id} @ {selected_date} = {highest_power:,} (pattern {i}, attempt {attempt+1})")
                            return highest_power
                        except Exception as e:
                            log_info(f"[HIGHEST POWER AT DATE] Parse error: {e}")

                if attempt == 2:
                    log_info(f"[HIGHEST POWER AT DATE] Not found for {account_id} @ {selected_date}")
                else:
                    log_info(f"[HIGHEST POWER AT DATE] No match attempt {attempt+1} for {account_id}, retrying...")
                    await asyncio.sleep(1)

        except asyncio.TimeoutError:
            log_info(f"[HIGHEST POWER AT DATE] Timeout attempt {attempt+1} for {account_id}")
            await asyncio.sleep(1)
        except Exception as e:
            log_info(f"[HIGHEST POWER AT DATE ERROR] attempt {attempt+1} for {account_id}: {e}")
            await asyncio.sleep(1)

    return None


async def fetch_advanced_stats_ranged(account_id, start_date, end_date):
    """
    Fetch the "Advanced War Stats" block (Infantry/Cavalry/Mage/Marksman/Other Merits,
    T4/T5 Units Rss Healed, T4/T5 Units Dead) for a date RANGE (start_date to end_date),
    same as the main season stats. Confirmed via direct browser test that this section
    IS present and correctly scoped on the ranged view (?start_date=X&end_date=Y).
    Returns a dict of the 7 fields (values as raw strings like "+16,161,386"), or all-None on failure.
    """
    import re

    url = f"https://callofstats.com/lord/{account_id}?start_date={start_date}&end_date={end_date}"
    field_labels = {
        "infantry_merits": "Infantry Merits",
        "cavalry_merits": "Cavalry Merits",
        "mage_merits": "Mage Merits",
        "marksman_merits": "Marksman Merits",
        "other_merits": "Other Merits",
        "t45_healed": "T4/T5 Units Rss Healed",
        "t45_dead": "T4/T5 Units Dead",
    }
    result = {k: None for k in field_labels}

    for attempt in range(3):
        try:
            session = await get_callofstats_session()
            if not session:
                log_info(f"[ADV RANGED] No session available, attempt {attempt+1}")
                await asyncio.sleep(2)
                continue

            async with session.get(url, allow_redirects=True) as response:
                if response.status != 200:
                    log_info(f"[ADV RANGED] HTTP {response.status} attempt {attempt+1} for {account_id} @ {start_date}->{end_date}")
                    await asyncio.sleep(1)
                    continue

                html = await response.text()

                if "<title>Login" in html or "Sign in to Call of Stats" in html:
                    log_info(f"[ADV RANGED] Got login page on attempt {attempt+1}, forcing session refresh")
                    global _callofstats_session, _session_login_time
                    if _callofstats_session:
                        await _callofstats_session.close()
                        _callofstats_session = None
                        _session_login_time = None
                    await asyncio.sleep(2)
                    continue

                for key, label in field_labels.items():
                    pattern = f'<span class="subtle">{re.escape(label)}</span>\\s*<div class="value">([^<]+)</div>'
                    match = re.search(pattern, html)
                    if match:
                        result[key] = match.group(1).strip()

                if any(v for v in result.values()):
                    log_info(f"[ADV RANGED] {account_id} @ {start_date}->{end_date}: t45_healed={result['t45_healed']}")
                    return result

                if attempt == 2:
                    log_info(f"[ADV RANGED] Not found for {account_id} @ {start_date}->{end_date}")
                else:
                    log_info(f"[ADV RANGED] No match attempt {attempt+1} for {account_id}, retrying...")
                    await asyncio.sleep(1)

        except asyncio.TimeoutError:
            log_info(f"[ADV RANGED] Timeout attempt {attempt+1} for {account_id}")
            await asyncio.sleep(1)
        except Exception as e:
            log_info(f"[ADV RANGED ERROR] attempt {attempt+1} for {account_id}: {e}")
            await asyncio.sleep(1)

    return result


async def fetch_achievement_stats(account_id, selected_date=None):
    """
    Fetch EXCHANGE COINS SPENT and MAX PETS achievement values from the profile page.
    If selected_date is given, fetches the value AS OF that date (for computing an increase
    over a range, same as Highest Power). Otherwise fetches the current/live value.
    Returns dict {"exchange_coins_spent": int|None, "max_pets": int|None}.
    """
    import re

    url = f"https://callofstats.com/lord/{account_id}"
    if selected_date:
        url += f"?selected_date={selected_date}"
    result = {"exchange_coins_spent": None, "max_pets": None}

    achievement_patterns = {
        "exchange_coins_spent": [
            r'<div class="achievement-name">Exchange Coins Spent</div>\s*<div class="achievement-labels">.*?<div class="achievement-values">\s*<span>\+?([0-9,]+)</span>',
            r'Exchange Coins Spent</div>.*?<span>\+?([0-9,]+)</span>',
        ],
        "max_pets": [
            r'<div class="achievement-name">Max Pets</div>\s*<div class="achievement-labels">.*?<div class="achievement-values">\s*<span>\+?([0-9,]+)</span>',
            r'Max Pets</div>.*?<span>\+?([0-9,]+)</span>',
        ],
    }

    for attempt in range(3):
        try:
            session = await get_callofstats_session()
            if not session:
                log_info(f"[ACHIEVEMENTS] No session available, attempt {attempt+1}")
                await asyncio.sleep(2)
                continue

            async with session.get(url, allow_redirects=True) as response:
                if response.status != 200:
                    log_info(f"[ACHIEVEMENTS] HTTP {response.status} attempt {attempt+1} for {account_id}")
                    await asyncio.sleep(1)
                    continue

                html = await response.text()

                if "<title>Login" in html or "Sign in to Call of Stats" in html:
                    log_info(f"[ACHIEVEMENTS] Got login page on attempt {attempt+1}, forcing session refresh")
                    global _callofstats_session, _session_login_time
                    if _callofstats_session:
                        await _callofstats_session.close()
                        _callofstats_session = None
                        _session_login_time = None
                    await asyncio.sleep(2)
                    continue

                for key, patterns in achievement_patterns.items():
                    for i, pattern in enumerate(patterns):
                        match = re.search(pattern, html, re.DOTALL)
                        if match:
                            val_str = match.group(1).replace(",", "")
                            try:
                                result[key] = int(val_str)
                                log_info(f"[ACHIEVEMENTS] {account_id} {key} = {result[key]:,} (pattern {i})")
                                break
                            except Exception as e:
                                log_info(f"[ACHIEVEMENTS] Parse error for {key}: {e}")

                # If we got at least one value, or this is the last attempt, return what we have
                if result["exchange_coins_spent"] is not None or result["max_pets"] is not None or attempt == 2:
                    if result["exchange_coins_spent"] is None and result["max_pets"] is None:
                        log_info(f"[ACHIEVEMENTS] Nothing found for {account_id}")
                    return result
                else:
                    log_info(f"[ACHIEVEMENTS] No match attempt {attempt+1} for {account_id}, retrying...")
                    await asyncio.sleep(1)

        except asyncio.TimeoutError:
            log_info(f"[ACHIEVEMENTS] Timeout attempt {attempt+1} for {account_id}")
            await asyncio.sleep(1)
        except Exception as e:
            log_info(f"[ACHIEVEMENTS ERROR] attempt {attempt+1} for {account_id}: {e}")
            await asyncio.sleep(1)

    return result

    return None


async def fetch_current_t_kills(account_id):
    """
    Fetch the CURRENT T5-T1 kill totals for a lord (no date range)
    Uses the authenticated Call of Stats session
    Returns dict: {"t5": 123456, "t4": 234567, ...} or empty dict if not found
    """
    import re
    
    try:
        # Use the authenticated session instead of creating a new one
        session = await get_callofstats_session()
        url = f"https://callofstats.com/lord/{account_id}"
        
        async with session.get(url, allow_redirects=True) as response:
            if response.status != 200:
                log_info(f"[CURRENT T-KILLS] Failed to fetch {url}: {response.status}")
                return {}
            
            html = await response.text()
            
            t_kills = {}
            
            # Look for T5 Kills pattern - very flexible
            for tier in ["T5", "T4", "T3", "T2", "T1"]:
                # Try multiple patterns
                patterns = [
                    f'{tier} Kills</span>.*?<div class="value">([0-9,]+)</div>',
                    f'<span class="subtle">{tier} Kills</span>.*?<div class="value">([0-9,]+)</div>',
                    f'{tier} Kills.*?([0-9,]+)',
                ]
                
                match = None
                for pattern in patterns:
                    match = re.search(pattern, html, re.DOTALL)
                    if match:
                        break
                
                if match:
                    kills_str = match.group(1).replace(",", "")
                    try:
                        t_kills[tier.lower()] = int(kills_str)
                        log_info(f"[CURRENT T-KILLS] Parsed {tier}: {kills_str}")
                    except Exception as e:
                        log_info(f"[CURRENT T-KILLS] Failed to parse {tier}: {e}")
                else:
                    log_info(f"[CURRENT T-KILLS] Pattern not found for {tier}")
                    # Debug
                    idx = html.find(f"{tier} Kills")
                    if idx != -1:
                        debug_html = html[max(0, idx-100):min(len(html), idx+300)]
                        log_info(f"[CURRENT T-KILLS DEBUG {tier}]:\n{debug_html}")
            
            if t_kills:
                log_info(f"[CURRENT T-KILLS] {account_id} = {t_kills}")
                return t_kills
            else:
                log_info(f"[CURRENT T-KILLS] Could not parse any T-kills for {account_id}")
                return {}
    except Exception as e:
        log_info(f"[CURRENT T-KILLS ERROR] {account_id}: {e}")
        import traceback
        traceback.print_exc()
        return {}


async def fetch_t_kills_at_date(account_id, selected_date):
    """
    Fetch T5-T1 kill totals for a lord as they stood on a SPECIFIC date (no range).
    Used for old/ended seasons instead of fetch_current_t_kills(), which always
    returns today's live totals.
    Returns dict: {"t5": 123456, ...} or empty dict if not found.
    """
    import re

    url = f"https://callofstats.com/lord/{account_id}?selected_date={selected_date}"

    try:
        session = await get_callofstats_session()
        if not session:
            return {}

        async with session.get(url, allow_redirects=True) as response:
            if response.status != 200:
                log_info(f"[T-KILLS AT DATE] Failed to fetch {url}: {response.status}")
                return {}

            html = await response.text()

            if "<title>Login" in html or "Sign in to Call of Stats" in html:
                log_info(f"[T-KILLS AT DATE] Got login page, forcing session refresh")
                global _callofstats_session, _session_login_time
                if _callofstats_session:
                    await _callofstats_session.close()
                    _callofstats_session = None
                    _session_login_time = None
                return {}

            t_kills = {}
            for tier in ["T5", "T4", "T3", "T2", "T1"]:
                patterns = [
                    f'{tier} Kills</span>.*?<div class="value">([0-9,]+)</div>',
                    f'<span class="subtle">{tier} Kills</span>.*?<div class="value">([0-9,]+)</div>',
                    f'{tier} Kills.*?([0-9,]+)',
                ]
                match = None
                for pattern in patterns:
                    match = re.search(pattern, html, re.DOTALL)
                    if match:
                        break
                if match:
                    kills_str = match.group(1).replace(",", "")
                    try:
                        t_kills[tier.lower()] = int(kills_str)
                    except Exception as e:
                        log_info(f"[T-KILLS AT DATE] Failed to parse {tier}: {e}")

            log_info(f"[T-KILLS AT DATE] {account_id} @ {selected_date} = {t_kills}")
            return t_kills
    except Exception as e:
        log_info(f"[T-KILLS AT DATE ERROR] {account_id}: {e}")
        return {}


async def fetch_latest_data_date(account_id):
    """
    Extract the latest data date from Call of Stats profile
    Reads from data-current-date attribute in linkacct-data div
    Returns date string in format "DD/MM/YYYY" or None if not found
    """
    import re
    from datetime import datetime
    
    try:
        session = await get_callofstats_session()
        url = f"https://callofstats.com/lord/{account_id}"
        
        async with session.get(url, allow_redirects=True) as response:
            if response.status != 200:
                log_info(f"[LATEST DATA DATE] Failed to fetch {url}: {response.status}")
                return None
            
            html = await response.text()
            
            # Extract from data-current-date attribute in linkacct-data div
            # Format in HTML: data-current-date="2026-03-25" (YYYY-MM-DD)
            pattern = r'data-current-date="(\d{4}-\d{2}-\d{2})"'
            match = re.search(pattern, html)
            
            if match:
                date_str = match.group(1)  # Returns YYYY-MM-DD format
                # Convert YYYY-MM-DD to DD/MM/YYYY for consistency
                dt = datetime.strptime(date_str, "%Y-%m-%d")
                formatted_date = dt.strftime("%d/%m/%Y")
                log_info(f"[LATEST DATA DATE] {account_id} = {formatted_date}")
                return formatted_date
            
            log_info(f"[LATEST DATA DATE] Could not find date for {account_id}")
            return None
    except Exception as e:
        log_info(f"[LATEST DATA DATE ERROR] {account_id}: {e}")
        return None


async def fetch_stats_with_fallback(account_id, start_date, end_date):
    """
    Fetch stats and automatically fallback to earlier dates if data is empty.
    Returns (stats, actual_end_date_used)
    Never falls back before start_date.
    """
    from datetime import timedelta
    
    end_dt = datetime.strptime(end_date, "%Y-%m-%d").date()
    start_dt = datetime.strptime(start_date, "%Y-%m-%d").date()
    
    last_stats = None

    # Try up to FALLBACK_DAYS back, but never below start_date
    for days_back in range(FALLBACK_DAYS):
        current_end_dt = end_dt - timedelta(days=days_back)
        
        # Don't go below season start
        if current_end_dt < start_dt:
            log_info(f"[FALLBACK] Reached season start, stopping")
            break
        
        current_end = current_end_dt.isoformat()
        log_info(f"[FALLBACK] Trying date: {current_end}")
        
        stats = await fetch_stats_for_account(account_id, start_date, current_end, skip_cache=True)
        
        if not stats:
            log_info(f"[FALLBACK] No stats returned for {current_end}, trying earlier")
            continue
        
        last_stats = stats
        
        # Check if any data exists (not all zeros/None)
        has_data = False
        for key in ["merits", "kills_gain", "deads_gain", "healed_gain",
                    "gold_gathered", "wood_gathered", "ore_gathered", "mana_gathered"]:
            val = stats.get(key)
            if val and val not in ("+0", "0", ""):
                has_data = True
                break
        
        if has_data:
            log_info(f"[FALLBACK] Found data for {current_end}")
            return stats, current_end
        else:
            log_info(f"[FALLBACK] All zeros for {current_end}, trying earlier")

    # Still nothing — the START date itself might be broken/deleted by COS, which shifting
    # the end date backward can't fix. Try shifting the START date forward a few days
    # against the original end_date before giving up entirely.
    for days_forward in range(1, 6):
        shifted_start = (start_dt + timedelta(days=days_forward)).isoformat()
        if datetime.strptime(shifted_start, "%Y-%m-%d").date() >= end_dt:
            break
        log_info(f"[FALLBACK] Trying shifted start_date={shifted_start} -> end_date={end_date}")
        stats = await fetch_stats_for_account(account_id, shifted_start, end_date, skip_cache=True)
        if not stats:
            continue
        has_data = any(
            stats.get(key) and stats.get(key) not in ("+0", "0", "")
            for key in ["merits", "kills_gain", "deads_gain", "healed_gain",
                        "gold_gathered", "wood_gathered", "ore_gathered", "mana_gathered"]
        )
        if has_data:
            log_info(f"[FALLBACK] Found data with shifted start_date={shifted_start}")
            return stats, end_date
        last_stats = stats

    # Return last fetched stats even if empty (last resort)
    return last_stats, end_date


def get_cached_stats(account_id, start_date, end_date):
    """Get stats from cache if valid (not expired)"""
    cache_key = f"{account_id}_{start_date}_{end_date}"
    if cache_key in _stats_cache:
        cached = _stats_cache[cache_key]
        age_hours = (datetime.utcnow() - cached["timestamp"]).total_seconds() / 3600
        if age_hours < CACHE_EXPIRY_HOURS:
            log_info(f"[CACHE HIT] {account_id} (age: {int(age_hours)}h)")
            return cached["stats"]
        else:
            del _stats_cache[cache_key]
    return None


def set_cached_stats(account_id, start_date, end_date, stats):
    """Store stats in cache with timestamp"""
    cache_key = f"{account_id}_{start_date}_{end_date}"
    _stats_cache[cache_key] = {"timestamp": datetime.utcnow(), "stats": stats}
    log_info(f"[CACHE SET] {account_id}")

# ============================================================
# SQLITE DATABASE
# ============================================================

def init_db():
    conn = sqlite3.connect(DB)
    c = conn.cursor()
    c.execute("""
        CREATE TABLE IF NOT EXISTS events (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            datetime TEXT NOT NULL,
            reminder INTEGER NOT NULL
        );
    """)
    c.execute("""
        CREATE TABLE IF NOT EXISTS seasons (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            season_name TEXT NOT NULL,
            start_date TEXT NOT NULL,
            created_at TEXT NOT NULL
        );
    """)
    c.execute("""
        CREATE TABLE IF NOT EXISTS lords (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            lord_name TEXT NOT NULL,
            account_id TEXT NOT NULL UNIQUE
        );
    """)
    c.execute("""
        CREATE TABLE IF NOT EXISTS callofstats_update (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            latest_data_date TEXT NOT NULL,
            last_checked TEXT NOT NULL,
            notified INTEGER DEFAULT 0
        );
    """)
    c.execute("""
        CREATE TABLE IF NOT EXISTS bot_status_config (
            id INTEGER PRIMARY KEY,
            mode TEXT NOT NULL DEFAULT 'default',
            custom_text TEXT,
            updated_at TEXT
        );
    """)
    c.execute("""
        CREATE TABLE IF NOT EXISTS queued_tasks (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            tool_name TEXT NOT NULL,
            tool_input TEXT NOT NULL,
            description TEXT NOT NULL,
            scheduled_time TEXT NOT NULL,
            status TEXT NOT NULL DEFAULT 'pending',
            created_by TEXT,
            channel_id TEXT NOT NULL,
            created_at TEXT NOT NULL,
            result TEXT
        );
    """)
    c.execute("""
        CREATE TABLE IF NOT EXISTS server_config (
            id INTEGER PRIMARY KEY,
            server_num INTEGER NOT NULL,
            set_at TEXT NOT NULL
        );
    """)
    c.execute("""
        CREATE TABLE IF NOT EXISTS server_lord_stats (
            server_num INTEGER NOT NULL,
            account_id TEXT NOT NULL,
            lord_name TEXT,
            current_power INTEGER DEFAULT 0,
            highest_power INTEGER DEFAULT 0,
            deaths INTEGER DEFAULT 0,
            total_merits INTEGER DEFAULT 0,
            gathering INTEGER DEFAULT 0,
            infantry_merits INTEGER DEFAULT 0,
            cavalry_merits INTEGER DEFAULT 0,
            marksman_merits INTEGER DEFAULT 0,
            mage_merits INTEGER DEFAULT 0,
            other_merits INTEGER DEFAULT 0,
            healing INTEGER DEFAULT 0,
            t4_deaths INTEGER DEFAULT 0,
            t5_deaths INTEGER DEFAULT 0,
            t4_severely_wounded INTEGER DEFAULT 0,
            t5_severely_wounded INTEGER DEFAULT 0,
            enemy_merits INTEGER DEFAULT 0,
            t4_healed INTEGER DEFAULT 0,
            t5_healed INTEGER DEFAULT 0,
            start_date TEXT,
            end_date TEXT,
            uploaded_at TEXT,
            PRIMARY KEY (server_num, account_id)
        );
    """)
    c.execute("""
        CREATE TABLE IF NOT EXISTS kvk_matchups (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nickname TEXT NOT NULL,
            created_at TEXT NOT NULL,
            num_zones INTEGER NOT NULL,
            zones_json TEXT NOT NULL,
            team1_zones TEXT NOT NULL,
            team2_zones TEXT NOT NULL,
            team1_json TEXT NOT NULL,
            team2_json TEXT NOT NULL
        );
    """)
    migrate_db(conn)
    conn.commit()
    conn.close()

def init_db_progress():
    """Initialize separate database for season progress tracking"""
    conn = sqlite3.connect(DB_PROGRESS)
    c = conn.cursor()
    c.execute("""
        CREATE TABLE IF NOT EXISTS season_progress (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            season_id INTEGER NOT NULL,
            account_id TEXT NOT NULL,
            data_date TEXT NOT NULL,
            lord_name TEXT NOT NULL,
            power_gain TEXT,
            merits TEXT,
            kills_gain TEXT,
            deads_gain TEXT,
            healed_gain TEXT,
            t5_gain TEXT,
            t4_gain TEXT,
            t3_gain TEXT,
            t2_gain TEXT,
            t1_gain TEXT,
            gold_spent TEXT,
            wood_spent TEXT,
            ore_spent TEXT,
            mana_spent TEXT,
            gold_gathered TEXT,
            wood_gathered TEXT,
            ore_gathered TEXT,
            mana_gathered TEXT,
            infantry_merits TEXT,
            cavalry_merits TEXT,
            mage_merits TEXT,
            marksman_merits TEXT,
            other_merits TEXT,
            t45_healed TEXT,
            t45_dead TEXT,
            created_at TEXT NOT NULL,
            UNIQUE(season_id, account_id, data_date)
        );
    """)
    conn.commit()
    conn.close()

def migrate_db(conn):
    """Handle database schema migrations for backwards compatibility"""
    try:
        c = conn.cursor()
        c.execute("PRAGMA table_info(seasons)")
        existing = {row[1] for row in c.fetchall()}
        if "end_date" not in existing:
            c.execute("ALTER TABLE seasons ADD COLUMN end_date TEXT")
            log_info(f"[DB MIGRATE] Added column: seasons.end_date")
            conn.commit()

        # server_lord_stats — COS's server-page Excel export added new columns
        c.execute("PRAGMA table_info(server_lord_stats)")
        existing_sls = {row[1] for row in c.fetchall()}
        new_sls_columns = [
            "t4_deaths", "t5_deaths", "t4_severely_wounded", "t5_severely_wounded",
            "enemy_merits", "t4_healed", "t5_healed",
        ]
        for col in new_sls_columns:
            if col not in existing_sls:
                c.execute(f"ALTER TABLE server_lord_stats ADD COLUMN {col} INTEGER DEFAULT 0")
                log_info(f"[DB MIGRATE] Added column: server_lord_stats.{col}")
        conn.commit()
    except Exception as e:
        log_error(f"[DB MIGRATION] Error: {e}")

def migrate_db_progress():
    """Add new columns to season_progress if they don't exist yet (safe to run every boot)"""
    new_columns = [
        ("infantry_merits", "TEXT"),
        ("cavalry_merits", "TEXT"),
        ("mage_merits", "TEXT"),
        ("marksman_merits", "TEXT"),
        ("other_merits", "TEXT"),
        ("t45_healed", "TEXT"),
        ("t45_dead", "TEXT"),
        ("highest_power", "TEXT"),
        ("exchange_coins_spent", "TEXT"),
        ("max_pets", "TEXT"),
        # Prepped ahead of a COS update that's expected to expose these — see the game's own
        # export settings panel. Columns exist now so no further migration is needed once
        # parsing for them gets wired up; until then they'll just stay NULL/unused.
        ("t4_deaths", "TEXT"),
        ("t5_deaths", "TEXT"),
        ("t4_severely_wounded", "TEXT"),
        ("t5_severely_wounded", "TEXT"),
        ("enemy_merits", "TEXT"),
        ("t4_healed_split", "TEXT"),
        ("t5_healed_split", "TEXT"),
        ("alliance_donations", "TEXT"),
        ("build_time", "TEXT"),
        ("destruction_time", "TEXT"),
        ("resource_assistance", "TEXT"),
        ("behemoth_raid_wins", "TEXT"),
        ("alliance_help", "TEXT"),
    ]
    try:
        conn = sqlite3.connect(DB_PROGRESS)
        c = conn.cursor()
        c.execute("PRAGMA table_info(season_progress)")
        existing = {row[1] for row in c.fetchall()}
        for col_name, col_type in new_columns:
            if col_name not in existing:
                c.execute(f"ALTER TABLE season_progress ADD COLUMN {col_name} {col_type}")
                log_info(f"[DB MIGRATE] Added column: {col_name}")
        conn.commit()
        conn.close()
    except Exception as e:
        log_error(f"[DB MIGRATE PROGRESS] Error: {e}")

def db_add_event(name, dt, reminder):
    conn = sqlite3.connect(DB)
    c = conn.cursor()
    c.execute(
        "INSERT INTO events (name, datetime, reminder) VALUES (?, ?, ?)",
        (name, dt, reminder)
    )
    log_info(f"[DB ADD EVENT] writing to: {DB}")
    conn.commit()
    conn.close()
    silent_backup()

def db_get_events():
    conn = sqlite3.connect(DB)
    c = conn.cursor()
    c.execute(
        "SELECT id, name, datetime, reminder FROM events ORDER BY datetime ASC"
    )
    rows = c.fetchall()
    conn.close()
    return rows

def db_update_event(event_id, name=None, dt=None, reminder=None):
    conn = sqlite3.connect(DB)
    c = conn.cursor()

    if name is not None:
        c.execute("UPDATE events SET name=? WHERE id=?", (name, event_id))
    if dt is not None:
        c.execute("UPDATE events SET datetime=? WHERE id=?", (dt, event_id))
    if reminder is not None:
        c.execute("UPDATE events SET reminder=? WHERE id=?", (reminder, event_id))

    conn.commit()
    conn.close()
    silent_backup()

def db_delete_event(event_id):
    conn = sqlite3.connect(DB)
    c = conn.cursor()
    c.execute("DELETE FROM events WHERE id=?", (event_id,))
    conn.commit()
    conn.close()
    silent_backup()

init_db()
init_db_progress()

def db_save_kvk_matchup(nickname, num_zones, zones_data, team1_zones, team2_zones, team1_totals, team2_totals):
    import json
    conn = sqlite3.connect(DB)
    c = conn.cursor()
    c.execute(
        "INSERT INTO kvk_matchups (nickname, created_at, num_zones, zones_json, team1_zones, team2_zones, team1_json, team2_json) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
        (
            nickname,
            datetime.utcnow().strftime("%Y-%m-%d %H:%M"),
            num_zones,
            json.dumps(zones_data),
            json.dumps(team1_zones),
            json.dumps(team2_zones),
            json.dumps(team1_totals),
            json.dumps(team2_totals),
        )
    )
    conn.commit()
    conn.close()
    log_info(f"[KVK] Saved matchup: {nickname}")

def db_get_kvk_matchups():
    conn = sqlite3.connect(DB)
    c = conn.cursor()
    c.execute("SELECT id, nickname, created_at, num_zones, team1_zones, team2_zones, team1_json, team2_json FROM kvk_matchups ORDER BY id DESC")
    rows = c.fetchall()
    conn.close()
    return rows

def db_get_kvk_matchup(matchup_id):
    import json
    conn = sqlite3.connect(DB)
    c = conn.cursor()
    c.execute("SELECT id, nickname, created_at, num_zones, zones_json, team1_zones, team2_zones, team1_json, team2_json FROM kvk_matchups WHERE id = ?", (matchup_id,))
    row = c.fetchone()
    conn.close()
    if not row:
        return None
    return {
        "id": row[0], "nickname": row[1], "created_at": row[2],
        "num_zones": row[3],
        "zones_data": json.loads(row[4]),
        "team1_zones": json.loads(row[5]),
        "team2_zones": json.loads(row[6]),
        "team1_totals": json.loads(row[7]),
        "team2_totals": json.loads(row[8]),
    }

def db_delete_kvk_matchup(matchup_id):
    conn = sqlite3.connect(DB)
    c = conn.cursor()
    c.execute("DELETE FROM kvk_matchups WHERE id = ?", (matchup_id,))
    conn.commit()
    conn.close()


def db_get_server_pick():
    conn = sqlite3.connect(DB)
    c = conn.cursor()
    c.execute("SELECT server_num FROM server_config ORDER BY id DESC LIMIT 1")
    row = c.fetchone()
    conn.close()
    return row[0] if row else None

def db_set_server_pick(server_num):
    conn = sqlite3.connect(DB)
    c = conn.cursor()
    c.execute("DELETE FROM server_config")
    c.execute("INSERT INTO server_config (server_num, set_at) VALUES (?, ?)",
              (server_num, date.today().isoformat()))
    conn.commit()
    conn.close()

def db_replace_server_lord_stats(server_num, rows, start_date, end_date):
    """Delete all existing rows for this server, then insert the new upload."""
    conn = sqlite3.connect(DB)
    c = conn.cursor()
    c.execute("DELETE FROM server_lord_stats WHERE server_num = ?", (server_num,))
    now = datetime.utcnow().isoformat()
    for r in rows:
        c.execute("""
            INSERT INTO server_lord_stats (
                server_num, account_id, lord_name, current_power, highest_power,
                deaths, total_merits, gathering, infantry_merits, cavalry_merits,
                marksman_merits, mage_merits, other_merits, healing,
                t4_deaths, t5_deaths, t4_severely_wounded, t5_severely_wounded,
                enemy_merits, t4_healed, t5_healed,
                start_date, end_date, uploaded_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            server_num, r["account_id"], r["lord_name"], r["current_power"], r["highest_power"],
            r["deaths"], r["total_merits"], r["gathering"], r["infantry_merits"], r["cavalry_merits"],
            r["marksman_merits"], r["mage_merits"], r["other_merits"], r["healing"],
            r.get("t4_deaths", 0), r.get("t5_deaths", 0), r.get("t4_severely_wounded", 0),
            r.get("t5_severely_wounded", 0), r.get("enemy_merits", 0), r.get("t4_healed", 0),
            r.get("t5_healed", 0),
            start_date, end_date, now
        ))
    conn.commit()
    conn.close()

def db_get_server_lord_stats(server_num):
    conn = sqlite3.connect(DB)
    c = conn.cursor()
    c.execute("""
        SELECT account_id, lord_name, current_power, highest_power, deaths, total_merits,
               gathering, infantry_merits, cavalry_merits, marksman_merits, mage_merits,
               other_merits, healing, t4_deaths, t5_deaths, t4_severely_wounded,
               t5_severely_wounded, enemy_merits, t4_healed, t5_healed, start_date, end_date
        FROM server_lord_stats WHERE server_num = ?
    """, (server_num,))
    rows = c.fetchall()
    conn.close()
    cols = ["account_id","lord_name","current_power","highest_power","deaths","total_merits",
            "gathering","infantry_merits","cavalry_merits","marksman_merits","mage_merits",
            "other_merits","healing","t4_deaths","t5_deaths","t4_severely_wounded",
            "t5_severely_wounded","enemy_merits","t4_healed","t5_healed","start_date","end_date"]
    return [dict(zip(cols, r)) for r in rows]
migrate_db_progress()

# ============================================================
# SEASON TRACKER DATABASE FUNCTIONS
# ============================================================

def db_queue_task(tool_name, tool_input, description, scheduled_time_iso, created_by, channel_id):
    conn = sqlite3.connect(DB)
    c = conn.cursor()
    c.execute(
        "INSERT INTO queued_tasks (tool_name, tool_input, description, scheduled_time, status, created_by, channel_id, created_at) "
        "VALUES (?, ?, ?, ?, 'pending', ?, ?, ?)",
        (tool_name, json.dumps(tool_input), description, scheduled_time_iso, str(created_by), str(channel_id), datetime.utcnow().isoformat())
    )
    task_id = c.lastrowid
    conn.commit()
    conn.close()
    return task_id

def db_get_queued_tasks(status="pending"):
    conn = sqlite3.connect(DB)
    c = conn.cursor()
    if status:
        c.execute("SELECT id, tool_name, tool_input, description, scheduled_time, status, created_by, channel_id, result FROM queued_tasks WHERE status=? ORDER BY scheduled_time ASC", (status,))
    else:
        c.execute("SELECT id, tool_name, tool_input, description, scheduled_time, status, created_by, channel_id, result FROM queued_tasks ORDER BY scheduled_time ASC")
    rows = c.fetchall()
    conn.close()
    return rows

def db_get_due_tasks():
    now = datetime.utcnow().isoformat()
    conn = sqlite3.connect(DB)
    c = conn.cursor()
    c.execute("SELECT id, tool_name, tool_input, description, scheduled_time, channel_id FROM queued_tasks WHERE status='pending' AND scheduled_time<=?", (now,))
    rows = c.fetchall()
    conn.close()
    return rows

def db_get_queued_task(task_id):
    conn = sqlite3.connect(DB)
    c = conn.cursor()
    c.execute("SELECT id, tool_name, tool_input, description, scheduled_time, status, created_by, channel_id, result FROM queued_tasks WHERE id=?", (task_id,))
    row = c.fetchone()
    conn.close()
    return row

def db_set_task_status(task_id, status, result=None):
    conn = sqlite3.connect(DB)
    c = conn.cursor()
    c.execute("UPDATE queued_tasks SET status=?, result=? WHERE id=?", (status, result, task_id))
    conn.commit()
    conn.close()

def db_update_task_time(task_id, new_scheduled_time_iso):
    conn = sqlite3.connect(DB)
    c = conn.cursor()
    c.execute("UPDATE queued_tasks SET scheduled_time=? WHERE id=? AND status='pending'", (new_scheduled_time_iso, task_id))
    changed = c.rowcount
    conn.commit()
    conn.close()
    return changed > 0

def db_add_season(season_name, start_date):
    conn = sqlite3.connect(DB)
    c = conn.cursor()
    created_at = datetime.utcnow().isoformat()
    c.execute(
        "INSERT INTO seasons (season_name, start_date, created_at) VALUES (?, ?, ?)",
        (season_name, start_date, created_at)
    )
    conn.commit()
    conn.close()
    silent_backup()

def count_season_data_dates(season_id, account_id):
    """Count unique dates with data for account in season"""
    try:
        conn = sqlite3.connect(DB_PROGRESS)
        c = conn.cursor()
        c.execute(
            "SELECT COUNT(DISTINCT data_date) FROM season_progress WHERE season_id = ? AND account_id = ?",
            (season_id, account_id)
        )
        count = c.fetchone()[0]
        conn.close()
        return count
    except Exception as e:
        log_info(f"[COUNT DATA DATES] Error: {e}")
        return 0



    conn = sqlite3.connect(DB)
    c = conn.cursor()
    c.execute("SELECT id, season_name, start_date, created_at FROM seasons ORDER BY created_at DESC LIMIT 1")
    row = c.fetchone()
    conn.close()
    return row

def db_get_current_season():
    conn = sqlite3.connect(DB)
    c = conn.cursor()
    c.execute("SELECT id, season_name, start_date, created_at FROM seasons ORDER BY created_at DESC LIMIT 1")
    row = c.fetchone()
    conn.close()
    return row

def db_get_season_by_name(season_name):
    """Get season by name (case-insensitive)"""
    conn = sqlite3.connect(DB)
    c = conn.cursor()
    c.execute("SELECT id, season_name, start_date, created_at FROM seasons WHERE LOWER(season_name) = LOWER(?) LIMIT 1", (season_name,))
    row = c.fetchone()
    conn.close()
    return row

def db_get_season_by_id(season_id):
    """Get season by its numeric ID"""
    conn = sqlite3.connect(DB)
    c = conn.cursor()
    c.execute("SELECT id, season_name, start_date, created_at FROM seasons WHERE id = ? LIMIT 1", (season_id,))
    row = c.fetchone()
    conn.close()
    return row

def resolve_season_input(season_input):
    """
    Resolve a user-provided season reference to a season tuple.
    Accepts either a numeric season ID (e.g. '1') or a season name (e.g. 'sos1').
    Returns the season tuple (id, season_name, start_date, created_at) or None.
    """
    if season_input is None:
        return None
    season_input = str(season_input).strip()
    if season_input.isdigit():
        season = db_get_season_by_id(int(season_input))
        if season:
            return season
        # Fall through to name lookup in case a season is literally named "1"
    return db_get_season_by_name(season_input)

def db_get_season_end_date(season_id):
    """Get the end_date for a season, or None if still active/ongoing."""
    conn = sqlite3.connect(DB)
    c = conn.cursor()
    c.execute("SELECT end_date FROM seasons WHERE id = ?", (season_id,))
    row = c.fetchone()
    conn.close()
    return row[0] if row and row[0] else None

def resolve_leaderboard_end_date(season_id):
    """
    For the CURRENT active season, query up to today. For an OLD/ended season, cap at
    that season's actual end_date instead — querying an ended season up to today bleeds
    the range into whatever season(s) came after it and produces wrong/zero results.
    Used by every !top*/!rss leaderboard command.
    """
    current_season = db_get_current_season()
    is_current = current_season and current_season[0] == season_id
    if is_current:
        return date.today().isoformat()
    return db_get_season_end_date(season_id) or date.today().isoformat()

def db_set_season_end_date(season_id, end_date):
    conn = sqlite3.connect(DB)
    c = conn.cursor()
    c.execute("UPDATE seasons SET end_date = ? WHERE id = ?", (end_date, season_id))
    conn.commit()
    conn.close()

def db_end_previous_active_season(new_season_start_date):
    """
    Mark the current most-recent season as ended — end_date is set to the day
    before the new season's start date — if it doesn't already have an end_date.
    Called right before a new season is created.
    """
    current = db_get_current_season()
    if not current:
        return
    season_id = current[0]
    if db_get_season_end_date(season_id) is None:
        start_dt = datetime.strptime(new_season_start_date, "%Y-%m-%d")
        end_dt = start_dt - timedelta(days=1)
        db_set_season_end_date(season_id, end_dt.strftime("%Y-%m-%d"))

def db_update_season(season_id, season_name=None, start_date=None, end_date=None):
    """Update one or more fields of a season. Pass None to leave a field unchanged."""
    conn = sqlite3.connect(DB)
    c = conn.cursor()
    if season_name is not None:
        c.execute("UPDATE seasons SET season_name = ? WHERE id = ?", (season_name, season_id))
    if start_date is not None:
        c.execute("UPDATE seasons SET start_date = ? WHERE id = ?", (start_date, season_id))
    if end_date is not None:
        c.execute("UPDATE seasons SET end_date = ? WHERE id = ?", (end_date, season_id))
    conn.commit()
    conn.close()

def db_get_all_seasons():
    """Get all seasons ordered by creation date"""
    conn = sqlite3.connect(DB)
    c = conn.cursor()
    c.execute("SELECT id, season_name, start_date, created_at FROM seasons ORDER BY created_at ASC")
    rows = c.fetchall()
    conn.close()
    return rows

def db_add_lord(lord_name, account_id):
    conn = sqlite3.connect(DB)
    c = conn.cursor()
    c.execute(
        "INSERT INTO lords (lord_name, account_id) VALUES (?, ?)",
        (lord_name, account_id)
    )
    conn.commit()
    conn.close()
    silent_backup()

def db_get_all_lords():
    conn = sqlite3.connect(DB)
    c = conn.cursor()
    c.execute("SELECT id, lord_name, account_id FROM lords ORDER BY lord_name ASC")
    rows = c.fetchall()
    conn.close()
    return rows

# ============================================================
# CALL OF STATS UPDATE TRACKING
# ============================================================

def db_get_last_known_data_date():
    """Get the last known Call of Stats data date"""
    conn = sqlite3.connect(DB)
    c = conn.cursor()
    c.execute("SELECT latest_data_date FROM callofstats_update ORDER BY id DESC LIMIT 1")
    row = c.fetchone()
    conn.close()
    return row[0] if row else None

def db_get_bot_status():
    """Get current bot status config: (mode, custom_text). mode is 'default' or 'custom'."""
    conn = sqlite3.connect(DB)
    c = conn.cursor()
    c.execute("SELECT mode, custom_text FROM bot_status_config ORDER BY id DESC LIMIT 1")
    row = c.fetchone()
    conn.close()
    if row:
        return row[0], row[1]
    return "default", None

def db_set_bot_status(mode, custom_text=None):
    """Set bot status config. mode is 'default' or 'custom'."""
    conn = sqlite3.connect(DB)
    c = conn.cursor()
    c.execute("DELETE FROM bot_status_config")
    c.execute(
        "INSERT INTO bot_status_config (mode, custom_text, updated_at) VALUES (?, ?, ?)",
        (mode, custom_text, datetime.utcnow().isoformat())
    )
    conn.commit()
    conn.close()

def db_update_data_date(new_date):
    """Update the latest data date and reset notified flag"""
    conn = sqlite3.connect(DB)
    c = conn.cursor()
    now = datetime.utcnow().isoformat()
    
    # Check if there's an existing record
    c.execute("SELECT id FROM callofstats_update LIMIT 1")
    if c.fetchone():
        c.execute("UPDATE callofstats_update SET latest_data_date=?, last_checked=?, notified=0", (new_date, now))
    else:
        c.execute("INSERT INTO callofstats_update (latest_data_date, last_checked, notified) VALUES (?, ?, 0)", (new_date, now))
    
    conn.commit()
    conn.close()

def db_mark_update_notified():
    """Mark that we've sent the notification"""
    conn = sqlite3.connect(DB)
    c = conn.cursor()
    c.execute("UPDATE callofstats_update SET notified=1 WHERE id=(SELECT id FROM callofstats_update ORDER BY id DESC LIMIT 1)")
    conn.commit()
    conn.close()

def db_is_update_notified():
    """Check if we've already notified about this update"""
    conn = sqlite3.connect(DB)
    c = conn.cursor()
    c.execute("SELECT notified FROM callofstats_update ORDER BY id DESC LIMIT 1")
    row = c.fetchone()
    conn.close()
    return row[0] if row else 0

def db_snapshot_exists(season_id, account_id, data_date):
    """Check if a snapshot already exists for a given date"""
    try:
        conn = sqlite3.connect(DB_PROGRESS)
        c = conn.cursor()
        c.execute(
            "SELECT id FROM season_progress WHERE season_id=? AND account_id=? AND data_date=?",
            (season_id, account_id, data_date)
        )
        row = c.fetchone()
        conn.close()
        return row is not None
    except Exception as e:
        log_error(f"[DB CHECK] Error checking snapshot: {e}")
        return False

def is_stats_empty(stats):
    """Check if stats are empty/all zeros"""
    if not stats:
        return True
    
    # Check if key stat fields are empty or zero
    key_fields = ["power_gain", "merits", "kills_gain", "lord_name"]
    
    for field in key_fields:
        value = stats.get(field, "")
        if value and value != "+0" and value != "0" and value != "Unknown":
            return False
    
    return True

def db_save_season_progress(season_id, account_id, lord_name, stats, data_date=None):
    """
    Save a member's progress for a specific date in a season.
    MERGES with any existing row for this (season_id, account_id, data_date) instead of
    blindly overwriting — a field only gets updated if the new value is real (not None
    and not "+0"/"-0"). This protects previously-archived good data (e.g. advanced war
    stats) from being silently wiped out by a later fetch that happens to come back
    worse or incomplete (COS delay, partial parse, temporary glitch, etc).
    """
    try:
        if not data_date:
            data_date = date.today().isoformat()

        field_names = [
            "power_gain", "merits", "kills_gain", "deads_gain", "healed_gain",
            "t5_gain", "t4_gain", "t3_gain", "t2_gain", "t1_gain",
            "gold_spent", "wood_spent", "ore_spent", "mana_spent",
            "gold_gathered", "wood_gathered", "ore_gathered", "mana_gathered",
            "infantry_merits", "cavalry_merits", "mage_merits", "marksman_merits", "other_merits",
            "t45_healed", "t45_dead", "highest_power", "exchange_coins_spent", "max_pets",
            # Prepped for a pending COS update — unused until parsing is wired up for them
            "t4_deaths", "t5_deaths", "t4_severely_wounded", "t5_severely_wounded",
            "enemy_merits", "t4_healed_split", "t5_healed_split", "alliance_donations",
            "build_time", "destruction_time", "resource_assistance", "behemoth_raid_wins",
            "alliance_help",
        ]

        def _is_real(v):
            if v is None:
                return False
            s = str(v).strip()
            return s not in ("", "+0", "-0", "0")

        conn = sqlite3.connect(DB_PROGRESS)
        try:
            c = conn.cursor()
            now = datetime.utcnow().isoformat()

            # Read existing row (if any) to merge against
            c.execute(
                f"SELECT {', '.join(field_names)} FROM season_progress WHERE season_id=? AND account_id=? AND data_date=?",
                (season_id, account_id, data_date)
            )
            existing_row = c.fetchone()
            existing = dict(zip(field_names, existing_row)) if existing_row else {}

            merged = {}
            for f in field_names:
                new_val = stats.get(f)
                merged[f] = new_val if _is_real(new_val) else existing.get(f, new_val)

            c.execute(f"""
                INSERT OR REPLACE INTO season_progress 
                (season_id, account_id, data_date, lord_name, {', '.join(field_names)}, created_at)
                VALUES (?, ?, ?, ?, {', '.join('?' for _ in field_names)}, ?)
            """, (
                season_id, account_id, data_date, lord_name,
                *[merged[f] for f in field_names],
                now
            ))
            conn.commit()
            log_info(f"[DB SAVE] {lord_name} ({account_id}) for {data_date}")
            return True
        finally:
            conn.close()
    except Exception as e:
        log_error(f"[DB SAVE PROGRESS] Error: {e}")
        return False

def db_save_advanced_stats(season_id, account_id, data_date, adv_stats, lord_name=None):
    """
    Merge advanced war stats fields into an existing (or new) season_progress row,
    WITHOUT touching any other fields (power_gain, merits, kills_gain, etc). Safe to call
    even when we only have the advanced stats and not the full core stats for that date.
    This permanently archives data locally so it survives COS deleting old dates later.
    A field only gets updated if the new value is real (not None/"+0") — never lets a
    worse/incomplete fetch overwrite a previously-good archived value.
    """
    adv_fields = ["infantry_merits", "cavalry_merits", "mage_merits", "marksman_merits",
                  "other_merits", "t45_healed", "t45_dead"]

    def _is_real(v):
        if v is None:
            return False
        s = str(v).strip()
        return s not in ("", "+0", "-0", "0")

    try:
        conn = sqlite3.connect(DB_PROGRESS)
        c = conn.cursor()
        c.execute(f"SELECT {', '.join(adv_fields)} FROM season_progress WHERE season_id=? AND account_id=? AND data_date=?",
                   (season_id, account_id, data_date))
        existing_row = c.fetchone()
        exists = existing_row is not None
        existing = dict(zip(adv_fields, existing_row)) if existing_row else {}
        now = datetime.utcnow().isoformat()

        merged = {f: (adv_stats.get(f) if _is_real(adv_stats.get(f)) else existing.get(f)) for f in adv_fields}

        if exists:
            c.execute(f"""
                UPDATE season_progress SET
                    {', '.join(f'{f}=?' for f in adv_fields)}
                WHERE season_id=? AND account_id=? AND data_date=?
            """, (*[merged[f] for f in adv_fields], season_id, account_id, data_date))
        else:
            c.execute(f"""
                INSERT INTO season_progress (season_id, account_id, data_date, lord_name,
                    {', '.join(adv_fields)}, created_at)
                VALUES (?, ?, ?, ?, {', '.join('?' for _ in adv_fields)}, ?)
            """, (
                season_id, account_id, data_date, lord_name or str(account_id),
                *[merged[f] for f in adv_fields], now
            ))
        conn.commit()
        conn.close()
        log_info(f"[DB SAVE ADV] {account_id} for {data_date} archived")
        return True
    except Exception as e:
        log_error(f"[DB SAVE ADV] Error: {e}")
        return False

def db_get_latest_advanced_snapshot(season_id, account_id):
    """
    Find the most recent date (any date, not limited to a fixed window) within this
    season that has real (nonzero) advanced war stats data archived for this account.
    Used as a last-resort fallback when the usual today/yesterday/day-before window
    comes up empty (e.g. COS fell further behind than expected, or deleted a date).
    Returns a dict with the 7 advanced fields + data_date, or None.
    """
    try:
        conn = sqlite3.connect(DB_PROGRESS)
        c = conn.cursor()
        c.execute("""
            SELECT data_date, infantry_merits, cavalry_merits, mage_merits, marksman_merits,
                   other_merits, t45_healed, t45_dead
            FROM season_progress
            WHERE season_id=? AND account_id=?
              AND (
                (infantry_merits IS NOT NULL AND infantry_merits != '+0') OR
                (cavalry_merits IS NOT NULL AND cavalry_merits != '+0') OR
                (mage_merits IS NOT NULL AND mage_merits != '+0') OR
                (marksman_merits IS NOT NULL AND marksman_merits != '+0') OR
                (other_merits IS NOT NULL AND other_merits != '+0') OR
                (t45_healed IS NOT NULL AND t45_healed != '+0') OR
                (t45_dead IS NOT NULL AND t45_dead != '+0')
              )
            ORDER BY data_date DESC LIMIT 1
        """, (season_id, account_id))
        row = c.fetchone()
        conn.close()
        if not row:
            return None
        return {
            "data_date": row[0],
            "infantry_merits": row[1], "cavalry_merits": row[2], "mage_merits": row[3],
            "marksman_merits": row[4], "other_merits": row[5],
            "t45_healed": row[6], "t45_dead": row[7],
        }
    except Exception as e:
        log_error(f"[DB GET LATEST ADV] Error: {e}")
        return None

def db_save_extra_stats(season_id, account_id, data_date, highest_power=None, exchange_coins_spent=None, max_pets=None, lord_name=None):
    """
    Merge-save highest_power / exchange_coins_spent / max_pets for a date, without
    touching any other columns. Same protection as db_save_advanced_stats — a field
    only gets updated if the new value is real, never overwritten with worse data.
    Permanently archives these so they survive COS deleting old dates.
    """
    updates = {}
    if highest_power is not None:
        updates["highest_power"] = str(highest_power)
    if exchange_coins_spent is not None:
        updates["exchange_coins_spent"] = str(exchange_coins_spent)
    if max_pets is not None:
        updates["max_pets"] = str(max_pets)
    if not updates:
        return False

    def _is_real(v):
        if v is None:
            return False
        s = str(v).strip()
        return s not in ("", "+0", "-0", "0")

    try:
        conn = sqlite3.connect(DB_PROGRESS)
        c = conn.cursor()
        fields = list(updates.keys())
        c.execute(f"SELECT {', '.join(fields)} FROM season_progress WHERE season_id=? AND account_id=? AND data_date=?",
                   (season_id, account_id, data_date))
        existing_row = c.fetchone()
        exists = existing_row is not None
        existing = dict(zip(fields, existing_row)) if existing_row else {}
        now = datetime.utcnow().isoformat()

        merged = {f: (updates[f] if _is_real(updates[f]) else existing.get(f)) for f in fields}

        if exists:
            c.execute(f"UPDATE season_progress SET {', '.join(f'{f}=?' for f in fields)} WHERE season_id=? AND account_id=? AND data_date=?",
                       (*[merged[f] for f in fields], season_id, account_id, data_date))
        else:
            c.execute(f"""
                INSERT INTO season_progress (season_id, account_id, data_date, lord_name, {', '.join(fields)}, created_at)
                VALUES (?, ?, ?, ?, {', '.join('?' for _ in fields)}, ?)
            """, (season_id, account_id, data_date, lord_name or str(account_id), *[merged[f] for f in fields], now))
        conn.commit()
        conn.close()
        log_info(f"[DB SAVE EXTRA] {account_id} for {data_date} archived: {list(updates.keys())}")
        return True
    except Exception as e:
        log_error(f"[DB SAVE EXTRA] Error: {e}")
        return False

def db_get_latest_field_value(season_id, account_id, field_name):
    """
    Find the most recent date (and value) where this ONE specific field is populated
    and nonzero, independent of whether other advanced fields are present on that same
    row. A row can have some fields filled and others still None (e.g. one field failed
    to parse on a given day while the rest succeeded) — this looks past that gap for
    just the field that's actually needed.
    Returns the raw value string (e.g. "+16,161,386"), or None.
    """
    if field_name not in ("infantry_merits", "cavalry_merits", "mage_merits", "marksman_merits",
                           "other_merits", "t45_healed", "t45_dead",
                           "highest_power", "exchange_coins_spent", "max_pets"):
        return None
    try:
        conn = sqlite3.connect(DB_PROGRESS)
        c = conn.cursor()
        c.execute(f"""
            SELECT {field_name} FROM season_progress
            WHERE season_id=? AND account_id=? AND {field_name} IS NOT NULL
              AND {field_name} != '+0' AND {field_name} != '0'
            ORDER BY data_date DESC LIMIT 1
        """, (season_id, account_id))
        row = c.fetchone()
        conn.close()
        return row[0] if row else None
    except Exception as e:
        log_error(f"[DB GET FIELD] Error: {e}")
        return None

def db_get_season_progress(season_id, account_id, data_date=None):
    """Get a member's progress for a specific date in a season (defaults to today)"""
    try:
        if not data_date:
            data_date = date.today().isoformat()
        
        conn = sqlite3.connect(DB_PROGRESS)
        try:
            c = conn.cursor()
            c.execute("""
                SELECT power_gain, merits, kills_gain, deads_gain, healed_gain,
                       t5_gain, t4_gain, t3_gain, t2_gain, t1_gain,
                       gold_spent, wood_spent, ore_spent, mana_spent,
                       gold_gathered, wood_gathered, ore_gathered, mana_gathered, lord_name, data_date,
                       infantry_merits, cavalry_merits, mage_merits, marksman_merits, other_merits,
                       t45_healed, t45_dead
                FROM season_progress
                WHERE season_id=? AND account_id=? AND data_date=?
            """, (season_id, account_id, data_date))
            row = c.fetchone()
            
            if not row:
                log_info(f"[DB QUERY] No data: season={season_id}, account={account_id}, date={data_date}")
                return None
            
            # Convert to stats dict format
            stats = {
                "power_gain": row[0],
                "merits": row[1],
                "kills_gain": row[2],
                "deads_gain": row[3],
                "healed_gain": row[4],
                "t5_gain": row[5],
                "t4_gain": row[6],
                "t3_gain": row[7],
                "t2_gain": row[8],
                "t1_gain": row[9],
                "gold_spent": row[10],
                "wood_spent": row[11],
                "ore_spent": row[12],
                "mana_spent": row[13],
                "gold_gathered": row[14],
                "wood_gathered": row[15],
                "ore_gathered": row[16],
                "mana_gathered": row[17],
                "lord_name": row[18],
                "data_date": row[19],
                "infantry_merits": row[20],
                "cavalry_merits": row[21],
                "mage_merits": row[22],
                "marksman_merits": row[23],
                "other_merits": row[24],
                "t45_healed": row[25],
                "t45_dead": row[26]
            }
            log_info(f"[DB HIT] {row[18]} ({account_id}) for {data_date}")
            return stats
        finally:
            conn.close()
    except Exception as e:
        log_error(f"[DB GET PROGRESS] Error: {e}")
        return None

def db_get_latest_season_progress(season_id, account_id):
    """Get the latest (most recent date) progress for a member in a season"""
    try:
        conn = sqlite3.connect(DB_PROGRESS)
        try:
            c = conn.cursor()
            c.execute("""
                SELECT power_gain, merits, kills_gain, deads_gain, healed_gain,
                       t5_gain, t4_gain, t3_gain, t2_gain, t1_gain,
                       gold_spent, wood_spent, ore_spent, mana_spent,
                       gold_gathered, wood_gathered, ore_gathered, mana_gathered, lord_name, data_date,
                       infantry_merits, cavalry_merits, mage_merits, marksman_merits, other_merits,
                       t45_healed, t45_dead, highest_power
                FROM season_progress
                WHERE season_id=? AND account_id=?
                ORDER BY data_date DESC
                LIMIT 1
            """, (season_id, account_id))
            row = c.fetchone()
            
            if not row:
                return None
            
            # Convert to stats dict format
            stats = {
                "power_gain": row[0],
                "merits": row[1],
                "kills_gain": row[2],
                "deads_gain": row[3],
                "healed_gain": row[4],
                "t5_gain": row[5],
                "t4_gain": row[6],
                "t3_gain": row[7],
                "t2_gain": row[8],
                "t1_gain": row[9],
                "gold_spent": row[10],
                "wood_spent": row[11],
                "ore_spent": row[12],
                "mana_spent": row[13],
                "gold_gathered": row[14],
                "wood_gathered": row[15],
                "ore_gathered": row[16],
                "mana_gathered": row[17],
                "lord_name": row[18],
                "data_date": row[19],
                "infantry_merits": row[20],
                "cavalry_merits": row[21],
                "mage_merits": row[22],
                "marksman_merits": row[23],
                "other_merits": row[24],
                "t45_healed": row[25],
                "t45_dead": row[26],
                "highest_power": row[27]
            }
            return stats
        finally:
            conn.close()
    except Exception as e:
        log_error(f"[DB GET LATEST PROGRESS] Error: {e}")
        return None

def db_get_lord(account_id):
    try:
        conn = sqlite3.connect(DB)
        try:
            c = conn.cursor()
            c.execute("SELECT id, lord_name, account_id FROM lords WHERE account_id=?", (account_id,))
            row = c.fetchone()
            return row
        finally:
            conn.close()
    except Exception as e:
        log_error(f"[DB GET LORD] Error: {e}")
        return None

# ============================================================
# CALLOFSTATS LOGIN & STATS FETCHING
# ============================================================

# Global session cache
_callofstats_session = None
_session_login_time = None

async def get_callofstats_session():
    """Get or create cached authenticated session"""
    global _callofstats_session, _session_login_time
    
    username = os.getenv("CALLOFSTATS_USERNAME")
    password = os.getenv("CALLOFSTATS_PASSWORD")
    
    if not username or not password:
        log_info("[CALLOFSTATS] Missing credentials in env variables")
        return None
    
    # Reuse session if it exists and is less than 30 minutes old
    if _callofstats_session and _session_login_time:
        age = (datetime.utcnow() - _session_login_time).total_seconds()
        if age < 1800:  # 30 minutes
            log_info(f"[CALLOFSTATS] Reusing cached session (age: {int(age)}s)")
            return _callofstats_session
        else:
            log_info("[CALLOFSTATS] Session expired, creating new one")
            await _callofstats_session.close()
            _callofstats_session = None
    
    # Create new session and login
    try:
        # Create session with timeout
        timeout = aiohttp.ClientTimeout(total=30, connect=10, sock_read=10)
        session = aiohttp.ClientSession(timeout=timeout)
        
        log_info("[CALLOFSTATS] Logging in...")
        async with session.post(
            "https://callofstats.com/login",
            data={"username": username, "password": password},
            allow_redirects=True
        ) as resp:
            if resp.status != 200:
                log_info(f"[CALLOFSTATS] Login failed: {resp.status}")
                await session.close()
                return None
            log_info("[CALLOFSTATS] Login successful (session cached)")
        
        _callofstats_session = session
        _session_login_time = datetime.utcnow()
        return session
    except asyncio.TimeoutError:
        log_info("[CALLOFSTATS] Login timed out (30 seconds)")
        return None
    except Exception as e:
        log_info(f"[CALLOFSTATS] Login error: {e}")
        return None

async def fetch_stats(start_date, end_date):
    """Fetch player stats from callofstats"""
    account_id = os.getenv("CALLOFSTATS_ACCOUNT_ID")
    if not account_id:
        log_info("[CALLOFSTATS] Missing CALLOFSTATS_ACCOUNT_ID")
        return None
    
    return await fetch_stats_for_account(account_id, start_date, end_date)

async def fetch_stats_for_account(account_id, start_date, end_date, skip_cache=False):
    """Fetch player stats from callofstats for a specific account (with caching)"""
    global _stats_cache
    
    if not account_id:
        log_info("[CALLOFSTATS] Missing account_id")
        return None
    
    # Check cache first (unless skip_cache=True)
    cache_key = f"{account_id}_{start_date}_{end_date}"
    if not skip_cache and cache_key in _stats_cache:
        cached = _stats_cache[cache_key]
        age = (datetime.utcnow() - cached["timestamp"]).total_seconds()
        if age < CACHE_DURATION:
            log_info(f"[CACHE HIT] Account {account_id} (age: {int(age)}s)")
            return cached["stats"]
        else:
            log_info(f"[CACHE EXPIRED] Account {account_id}, fetching fresh data")
            del _stats_cache[cache_key]
    elif skip_cache:
        log_info(f"[SKIP CACHE] Fetching fresh data for {account_id}")
    
    try:
        # Get cached session (reuses login)
        session = await get_callofstats_session()
        if not session:
            return None
        
        # Format dates properly
        try:
            start_dt = datetime.strptime(start_date, "%Y-%m-%d")
            end_dt = datetime.strptime(end_date, "%Y-%m-%d")
            start_date_formatted = start_dt.strftime("%Y-%m-%d")
            end_date_formatted = end_dt.strftime("%Y-%m-%d")
        except Exception as e:
            start_date_formatted = start_date
            end_date_formatted = end_date
        
        url = f"https://callofstats.com/lord/{account_id}?start_date={start_date_formatted}&end_date={end_date_formatted}"
        log_info(f"[CALLOFSTATS] Fetching: {url}")
        
        async with session.get(url, allow_redirects=True) as resp:
            if resp.status == 200:
                html = await resp.text()
                log_info(f"[CALLOFSTATS] Fetch successful ({len(html)} bytes)")
                stats = parse_stats(html)
                
                # Only cache if result has real data (don't cache all-zero responses)
                if stats and not skip_cache:
                    has_real_data = any(
                        stats.get(k) and stats.get(k) not in ("+0", "0", "")
                        for k in ["merits", "kills_gain", "healed_gain", "mana_gathered"]
                    )
                    if has_real_data:
                        _stats_cache[cache_key] = {"timestamp": datetime.utcnow(), "stats": stats}
                
                return stats
            else:
                log_info(f"[CALLOFSTATS] Fetch failed: {resp.status}")
                return None
    except asyncio.TimeoutError:
        log_info("[CALLOFSTATS] Request timed out (30 seconds)")
        return None
    except Exception as e:
        log_info(f"[CALLOFSTATS] Fetch error: {e}")
        return None

def parse_stats(html):
    """Parse stats from HTML using proper HTML parsing"""
    import re
    
    stats = {
        "lord_name": None,
        "alliance_tag": None,
        "power": None,
        "power_gain": None,
        "merits": None,
        "merits_pct": None,
        "kills_gain": None,
        "deads_gain": None,
        "healed_gain": None,
        "t5_gain": None,
        "t4_gain": None,
        "t3_gain": None,
        "t2_gain": None,
        "t1_gain": None,
        "gold_spent": None,
        "wood_spent": None,
        "ore_spent": None,
        "mana_spent": None,
        "rss_spent_total": None,
        "gold_gathered": None,
        "wood_gathered": None,
        "ore_gathered": None,
        "mana_gathered": None,
        "rss_gathered_total": None,
        "infantry_merits": None,
        "cavalry_merits": None,
        "mage_merits": None,
        "marksman_merits": None,
        "other_merits": None,
        "t45_healed": None,
        "t45_dead": None
    }
    
    try:
        def find_stat_value(label_name):
            """Find value for a stat by its label using HTML structure"""
            # Look for: <span class="subtle">LABEL</span> ... <div class="value">VALUE</div>
            pattern = f'<span class="subtle">{re.escape(label_name)}</span>\\s*<div class="value">([^<]+)</div>'
            match = re.search(pattern, html)
            if match:
                val = match.group(1).strip()
                log_debug(f"[PARSE DEBUG] Found {label_name}: {val}")
                return val
            else:
                log_debug(f"[PARSE DEBUG] NOT FOUND: {label_name}")
            return None
        
        # Extract lord name - look for <h1 class="higher-value">NAME</h1>
        name_match = re.search(r'<h1 class="higher-value">([^<]+)</h1>', html)
        if name_match:
            stats["lord_name"] = name_match.group(1).strip()
        
        # Extract alliance tag - look for <h2 class="higher-value">[TAG]</h2>
        tag_match = re.search(r'<h2 class="higher-value">([^<]+)</h2>', html)
        if tag_match:
            stats["alliance_tag"] = tag_match.group(1).strip()
        
        # Power stats
        stats["power_gain"] = find_stat_value("Highest Power")
        
        # Merits
        stats["merits"] = find_stat_value("Merits")
        stats["merits_pct"] = find_stat_value("Merit to Power Ratio")
        
        # War stats
        stats["kills_gain"] = find_stat_value("Units Killed")
        stats["deads_gain"] = find_stat_value("Units Dead")
        stats["healed_gain"] = find_stat_value("Units Healed")
        
        # Tiered Kills
        stats["t5_gain"] = find_stat_value("T5 Kills")
        stats["t4_gain"] = find_stat_value("T4 Kills")
        stats["t3_gain"] = find_stat_value("T3 Kills")
        stats["t2_gain"] = find_stat_value("T2 Kills")
        stats["t1_gain"] = find_stat_value("T1 Kills")
        
        # Resources Gathered
        stats["gold_gathered"] = find_stat_value("Gold Gathered")
        stats["wood_gathered"] = find_stat_value("Wood Gathered")
        stats["ore_gathered"] = find_stat_value("Ore Gathered")
        stats["mana_gathered"] = find_stat_value("Mana Gathered")
        
        # Resources Spent
        stats["gold_spent"] = find_stat_value("Gold Spent")
        stats["wood_spent"] = find_stat_value("Wood Spent")
        stats["ore_spent"] = find_stat_value("Ore Spent")
        stats["mana_spent"] = find_stat_value("Mana Spent")
        
        # Advanced War Stats
        stats["infantry_merits"] = find_stat_value("Infantry Merits")
        stats["cavalry_merits"] = find_stat_value("Cavalry Merits")
        stats["mage_merits"] = find_stat_value("Mage Merits")
        stats["marksman_merits"] = find_stat_value("Marksman Merits")
        stats["other_merits"] = find_stat_value("Other Merits")
        stats["t45_healed"] = find_stat_value("T4/T5 Units Rss Healed")
        stats["t45_dead"] = find_stat_value("T4/T5 Units Dead")
        
        found_count = len([v for v in stats.values() if v])
        log_info(f"[PARSE] Successfully parsed stats: {found_count} fields found")
        log_info(f"[PARSE] Lord name: {stats['lord_name']}")
        log_info(f"[PARSE] Alliance tag: {stats['alliance_tag']}")
        log_info(f"[PARSE] RSS Spent - Gold: {stats['gold_spent']}, Wood: {stats['wood_spent']}, Ore: {stats['ore_spent']}, Mana: {stats['mana_spent']}")
        log_info(f"[PARSE] T-Kills - T5: {stats['t5_gain']}, T4: {stats['t4_gain']}, T3: {stats['t3_gain']}, T2: {stats['t2_gain']}, T1: {stats['t1_gain']}")
        return stats
    except Exception as e:
        log_info(f"[PARSE STATS] Error: {e}")
        return None

# ============================================================
# DEFAULT ABYSS CONFIG
# ============================================================

DEFAULT_ABYSS = {
    "days": [1, 4, 6],
    "hours": [0, 4, 8, 12, 16, 20],
    "reminder_hours": [0, 4, 8, 12, 16, 20],
    "round2": True
}

def ensure_file(path, default):
    if not os.path.exists(path):
        with open(path, "w") as f:
            json.dump(default, f, indent=2)

def load_json(path, default):
    ensure_file(path, default)
    with open(path, "r") as f:
        return json.load(f)

def save_json(path, data):
    with open(path, "w") as f:
        json.dump(data, f, indent=2)

cfg = load_json(ABYSS_CONFIG_FILE, DEFAULT_ABYSS)
for k in DEFAULT_ABYSS:
    if k not in cfg:
        cfg[k] = DEFAULT_ABYSS[k]
save_json(ABYSS_CONFIG_FILE, cfg)

ABYSS_DAYS = cfg["days"]
ABYSS_HOURS = cfg["hours"]
REMINDER_HOURS = cfg["reminder_hours"]
REMINDER_MINS = cfg.get("reminder_mins", 15)
ROUND2_ENABLED = cfg["round2"]

# ============================================================
# DISCORD BOT SETUP
# ============================================================

intents = discord.Intents.default()
intents.members = True
intents.message_content = True

bot = commands.Bot(command_prefix="!", intents=intents, help_command=None)
bot.active_edit = None

# ============================================================
# BACKUP SLASH COMMANDS (OWNER ONLY)
# ============================================================

@bot.tree.command(name="backup", description="List recent backups")
async def backup(inter):
    if inter.user.id != OWNER_ID:
        return await inter.response.send_message("❌ Owner only.", ephemeral=True)

    files = sorted(
        [f for f in os.listdir(BACKUP_DIR) if f.endswith(".zip")],
        reverse=True
    )

    if not files:
        return await inter.response.send_message("📦 No backups found.", ephemeral=True)

    msg = "**📦 Recent Backups:**\n" + "\n".join(
        f"- `{f}`" for f in files[:MAX_BACKUPS]
    )
    await inter.response.send_message(msg, ephemeral=True)

@bot.tree.command(name="setstatus", description="Set the bot's Discord status (owner only)")
async def setstatus(inter, text: str):
    if inter.user.id != OWNER_ID:
        return await inter.response.send_message("❌ Owner only.", ephemeral=True)

    if text.strip().lower() == "default":
        db_set_bot_status("default")
        await update_bot_presence()
        latest_date = db_get_last_known_data_date()
        shown = f"Data: {latest_date}" if latest_date else "Abyss events ⚔️"
        return await inter.response.send_message(
            f"✅ Status reset to default — now showing: **{shown}**", ephemeral=True
        )

    db_set_bot_status("custom", text)
    await update_bot_presence()
    await inter.response.send_message(f"✅ Status set to: **{text}**", ephemeral=True)


@bot.tree.command(name="say", description="Make the bot say something (owner only)")
async def say(inter, text: str):
    if inter.user.id != OWNER_ID:
        return await inter.response.send_message("❌ Owner only.", ephemeral=True)

    await inter.response.send_message("✅ Sent.", ephemeral=True)
    await inter.channel.send(text)


@bot.tree.command(name="forcebackup", description="Create a backup now")
async def forcebackup(inter):
    if inter.user.id != OWNER_ID:
        return await inter.response.send_message("❌ Owner only.", ephemeral=True)

    make_backup()
    await inter.response.send_message("✅ Backup created.", ephemeral=True)

@bot.tree.command(name="restorebackup", description="Restore a backup")
async def restorebackup(inter):
    if inter.user.id != OWNER_ID:
        return await inter.response.send_message("❌ Owner only.", ephemeral=True)

    files = sorted(
        [f for f in os.listdir(BACKUP_DIR) if f.endswith(".zip")],
        reverse=True
    )

    if not files:
        return await inter.response.send_message("❌ No backups found.", ephemeral=True)

    await inter.response.defer(ephemeral=True)  # ✅ CRITICAL

    select = Select(
        placeholder="Select a backup to restore",
        options=[discord.SelectOption(label=f, value=f) for f in files]
    )

    async def cb(i):
        await i.response.defer(ephemeral=True)

        path = os.path.join(BACKUP_DIR, select.values[0])

        with zipfile.ZipFile(path, "r") as z:
            z.extractall(".")

        await i.followup.send(
            f"♻️ Restored `{select.values[0]}`. Restarting bot...",
            ephemeral=True
        )

        await asyncio.sleep(2)
       # auto-restart on Koyeb

    select.callback = cb
    view = View(timeout=60)
    view.add_item(select)

    await inter.followup.send("Choose a backup to restore:", view=view, ephemeral=True)

# ============================================================
# ON READY
# ============================================================


async def force_refresh_all_stats():
    """
    Force-fetch and cache all members' stats for current season.
    Called when Call of Stats update is detected.
    Saves progress to database for future !progress queries on past seasons.
    """
    try:
        season = db_get_current_season()
        if not season:
            log_info("[CACHE REFRESH] No active season")
            return
        
        season_id, season_name, start_date, created_at = season
        today = date.today().isoformat()
        
        guild = bot.get_guild(bot.guilds[0].id) if bot.guilds else None
        if not guild:
            log_info("[CACHE REFRESH] No guild found")
            return
        
        lords = get_all_lords_from_guild(guild)
        accounts_to_refresh = []
        
        # Add lords from guild roles
        for lord in lords:
            accounts_to_refresh.append(lord["account_id"])
        
        # Add mapped accounts (like Havi)
        for discord_id, account_id in DISCORD_TO_ACCOUNT_ID.items():
            accounts_to_refresh.append(account_id)
        
        log_info(f"[CACHE REFRESH] Starting bulk refresh for {len(accounts_to_refresh)} members")
        
        # Fetch and cache stats for each member
        count = 0
        for account_id in accounts_to_refresh:
            try:
                # Fetch today's stats (will fallback if needed)
                stats_today, actual_date_today = await fetch_stats_with_fallback(account_id, start_date, today)
                
                if stats_today:
                    # Cache today's stats (in-memory)
                    set_cached_stats(account_id, start_date, today, stats_today)
                    log_info(f"[FORCEFETCH] Cached today {account_id} for {today}")
                    
                    # SAVE to database with actual date (handles missed dates like 24/03)
                    db_save_season_progress(season_id, account_id, stats_today.get("lord_name", account_id), stats_today, actual_date_today)
                    log_info(f"[FORCEFETCH] Saved today {account_id} for {actual_date_today}")

                    # Also fetch and save the "original page" data (Highest Power, Achievements)
                    # that the ranged/date-comparison fetch above doesn't include — this used to
                    # only get archived opportunistically whenever !progress happened to run for
                    # a given account, leaving gaps for anyone nobody checked that day.
                    try:
                        hp_val, achievements = await asyncio.gather(
                            fetch_highest_power(account_id),
                            fetch_achievement_stats(account_id, actual_date_today),
                        )
                        db_save_extra_stats(
                            season_id, account_id, actual_date_today,
                            highest_power=hp_val,
                            exchange_coins_spent=achievements.get("exchange_coins_spent"),
                            max_pets=achievements.get("max_pets"),
                            lord_name=stats_today.get("lord_name", account_id),
                        )
                        log_info(f"[FORCEFETCH] Saved extra stats (power/achievements) for {account_id} on {actual_date_today}")
                    except Exception as e:
                        log_info(f"[FORCEFETCH] Extra stats fetch failed for {account_id}: {e}")
                    
                    # Also cache the day before for comparisons
                    day_before = (datetime.strptime(actual_date_today, "%Y-%m-%d").date() - timedelta(days=1)).isoformat()
                    stats_yesterday, actual_date_yesterday = await fetch_stats_with_fallback(account_id, start_date, day_before)
                    
                    if stats_yesterday:
                        set_cached_stats(account_id, start_date, day_before, stats_yesterday)
                        log_info(f"[FORCEFETCH] Cached yesterday {account_id} for {day_before} (actual: {actual_date_yesterday})")
                        
                        # Also save yesterday to database
                        db_save_season_progress(season_id, account_id, stats_yesterday.get("lord_name", account_id), stats_yesterday, actual_date_yesterday)
                        log_info(f"[FORCEFETCH] Saved yesterday {account_id} for {actual_date_yesterday}")
                    else:
                        log_info(f"[FORCEFETCH] ⚠️ No yesterday data for {account_id} (tried {day_before})")
                    
                    count += 1
            except Exception as e:
                log_error(f"[CACHE REFRESH] Error for {account_id}: {e}")
                continue
        
        log_info(f"[CACHE REFRESH] Complete! Cached & saved {count}/{len(accounts_to_refresh)} members to database")
    except Exception as e:
        log_error(f"[CACHE REFRESH ERROR] {e}")


@tasks.loop(minutes=1)
async def check_callofstats_update():
    """
    Check every 5 minutes if new Call of Stats data is available
    Only notify if the date changed AND actual data exists for new date
    """
    try:
        # Always check Rekz's profile for latest data
        account_id = REKZ_ACCOUNT_ID
        
        # Fetch latest data date from website
        latest_date = await fetch_latest_data_date(account_id)
        if not latest_date:
            log_info(f"[CALLOFSTATS UPDATE] Could not fetch latest date")
            return
        
        # Get last known date
        last_known = db_get_last_known_data_date()
        
        # If date changed, verify actual data exists before notifying
        if last_known != latest_date:
            log_info(f"[CALLOFSTATS UPDATE] Date changed {last_known} -> {latest_date}, verifying data exists...")
            
            # Try to fetch stats for the new date to verify it exists
            # Convert DD/MM/YYYY to YYYY-MM-DD for the query
            from datetime import datetime
            try:
                date_obj = datetime.strptime(latest_date, "%d/%m/%Y")
                new_date_iso = date_obj.strftime("%Y-%m-%d")
                
                # Fetch season start date
                season = db_get_current_season()
                if not season:
                    log_info(f"[CALLOFSTATS UPDATE] No active season")
                    return
                season_id, season_name, start_date, created_at = season
                
                # Use a short, always-recent window for this verification query instead of
                # the season's actual start_date — COS has been deleting old dates, and if
                # the season's start_date gets deleted, a range anchored there returns empty/
                # broken data, silently blocking notifications until someone manually edits
                # the season's start date forward. A short recent window sidesteps that
                # entirely since it never touches old dates that might be gone.
                verify_start = (date_obj - timedelta(days=3)).strftime("%Y-%m-%d")
                test_stats = await fetch_stats_for_account(account_id, verify_start, new_date_iso, skip_cache=True)
                
                # Verify we got valid data (not empty/unknown)
                if not test_stats or test_stats.get("lord_name") == "Unknown":
                    log_info(f"[CALLOFSTATS UPDATE] No actual data for {new_date_iso} yet (stats missing), skipping notification")
                    return
                
                # STRICT verification: check MULTIPLE stats to confirm NEW data is ready
                # Not just one stat - need at least 2+ stats with real values
                real_stat_count = 0
                for stat_key in ["power_gain", "merits", "kills_gain", "deads_gain", "healed_gain"]:
                    stat_val = test_stats.get(stat_key, "0")
                    try:
                        num_val = int(str(stat_val).replace("+", "").replace(",", "") or 0)
                        if num_val > 0:
                            real_stat_count += 1
                            log_info(f"[CALLOFSTATS UPDATE] Verified stat: {stat_key}={num_val}")
                    except:
                        pass
                
                # Require at least 1 stat with a real value (not just an "Unknown" placeholder).
                # Used to require 2+, but a legitimate new-data day can have some stats at
                # genuinely zero (e.g. no kills that day) while others are real — requiring
                # 2+ was rejecting valid updates and blocking notifications for extended periods.
                if real_stat_count < 1:
                    log_info(f"[CALLOFSTATS UPDATE] ⚠️ Only {real_stat_count} stat(s) with data for {new_date_iso} - data not ready yet, skipping notification")
                    return
                
                log_info(f"[CALLOFSTATS UPDATE] ✅ Verified {real_stat_count} stats for {new_date_iso} - data is real!")
                
            except Exception as e:
                log_info(f"[CALLOFSTATS UPDATE] Error verifying data: {e}")
                return
            
            # Update database
            db_update_data_date(latest_date)

            # Refresh bot status/presence to reflect new date (only affects "default" mode)
            await update_bot_presence()

            # IMMEDIATELY refresh cache with new data
            log_info(f"[CALLOFSTATS UPDATE] Triggering cache refresh...")
            await force_refresh_all_stats()
            
            # Send notification
            try:
                guild = bot.get_guild(bot.guilds[0].id) if bot.guilds else None
                if guild:
                    update_channel = guild.get_channel(BACKUP_CHANNEL_ID)
                    if update_channel:
                        embed = discord.Embed(
                            title="🔄 Call of Stats Update",
                            description=f"<@{OWNER_ID}> New data for **{latest_date}** cached and ready!",
                            color=0x00FF00
                        )
                        embed.set_footer(text="Cache refreshed ✅")
                        await update_channel.send(embed=embed)
                        
                        # Also mark as notified
                        db_mark_update_notified()
                        log_info(f"[CALLOFSTATS UPDATE] Notification sent to channel {BACKUP_CHANNEL_ID}")
            except Exception as e:
                log_info(f"[CALLOFSTATS UPDATE CHANNEL ERROR] {e}")
    except Exception as e:
        log_info(f"[CALLOFSTATS UPDATE ERROR] {e}")
        import traceback
        traceback.print_exc()

@tasks.loop(minutes=5)
async def self_ping():
    try:
        url = os.getenv("KOYEB_PUBLIC_URL")
        if not url:
            return

        async with aiohttp.ClientSession() as session:
            async with session.get(url) as resp:
                pass
    except Exception as e:
        log_info("[Self Ping Error]", e)

def preload_cache_from_db():
    """Load all recent season data from database into cache on bot startup"""
    try:
        season = db_get_current_season()
        if not season:
            log_info("[PRELOAD] No active season")
            return
        
        season_id, season_name, start_date, created_at = season
        
        conn = sqlite3.connect(DB_PROGRESS)
        c = conn.cursor()
        
        # Get all unique account_ids and dates in the current season
        c.execute("""
            SELECT DISTINCT account_id, data_date
            FROM season_progress
            WHERE season_id = ?
            ORDER BY data_date DESC
            LIMIT 200
        """, (season_id,))
        
        rows = c.fetchall()
        conn.close()
        
        if not rows:
            log_info("[PRELOAD] No data found in database")
            return
        
        # Load each into cache
        count = 0
        for account_id, data_date in rows:
            stats = db_get_season_progress(season_id, account_id, data_date)
            if stats:
                set_cached_stats(account_id, start_date, data_date, stats)
                count += 1
        
        log_info(f"[PRELOAD] Loaded {count} entries into cache from database")
    except Exception as e:
        log_error(f"[PRELOAD ERROR] {e}")

async def update_bot_presence():
    """Set the bot's Discord status based on bot_status_config: default (latest COS data date) or custom text."""
    mode, custom_text = db_get_bot_status()

    if mode == "custom" and custom_text:
        text = custom_text
    else:
        latest_date = db_get_last_known_data_date()
        text = f"Data: {latest_date}" if latest_date else "Abyss events ⚔️"

    try:
        activity = discord.Activity(type=discord.ActivityType.watching, name=text)
        await bot.change_presence(activity=activity)
    except Exception as e:
        log_info(f"[PRESENCE] Failed to update: {e}")


@bot.event
async def on_ready():
    print(f"Logged in as {bot.user}")

    # Set bot status/activity
    await update_bot_presence()

    # Sync commands once
    try:
        synced = await bot.tree.sync()
        print(f"✅ Synced {len(synced)} commands")
    except Exception as e:
        print(f"❌ Command sync failed: {e}")
    
    # Pre-load cache from database on startup (so commands work immediately)
    preload_cache_from_db()

    # ✅ DELETE OLD DATA (ONE-TIME CLEANUP ON RESTART)
    try:
        season = db_get_current_season()
        if season:
            season_id, season_name, start_date, created_at = season
            conn = sqlite3.connect(DB_PROGRESS)
            c = conn.cursor()
            c.execute("DELETE FROM season_progress WHERE data_date < ?", (start_date,))
            deleted = c.rowcount
            conn.commit()
            conn.close()
            if deleted > 0:
                log_info(f"🧹 [STARTUP] Deleted {deleted} old snapshots before {start_date}")
                print(f"✅ Cleaned up {deleted} old data entries")
    except Exception as e:
        log_error(f"[STARTUP CLEANUP] Error: {e}")

    # ✅ START SELF-PING (CRITICAL)
    if not self_ping.is_running():
        self_ping.start()

    # ✅ CHECK CALLOFSTATS UPDATES
    if not check_callofstats_update.is_running():
        check_callofstats_update.start()

    # ✅ QUEUED TASK RUNNER (!ask scheduled actions)
    if not check_queued_tasks.is_running():
        check_queued_tasks.start()

    # ✅ SAFE LOOP STARTS
    if not abyss_reminder_loop.is_running():
        abyss_reminder_loop.start()

    if not custom_event_loop.is_running():
        custom_event_loop.start()

    ch = bot.get_channel(update_channel_id)
    if ch:
        await ch.send("🤖 Bot restarted successfully.")


# ============================================================
# HELP COMMAND
# ============================================================

def build_help_embed(is_owner: bool):
    embed = discord.Embed(
        title="📖 Bot Commands",
        description="Here's everything you can do — commands are grouped by category below.",
        color=0x3498db
    )

    # ============ MEMBER COMMANDS ============
    embed.add_field(
        name="🟢 ​ M E M B E R ​ C O M M A N D S",
        value="\u200b",
        inline=False
    )

    embed.add_field(
        name="📊 Progress & Stats",
        value=(
            "`!progress [user] [season]` — Full season stats (works for past seasons too, e.g. `!progress rekz 1`)\n"
            "`!chart <user> <stat> [season]` — Line chart of a stat's growth over time\n"
            "`!groupchart <stat> [season]` — One chart with every tracked member's growth\n"
            "`!q [user]` — Quick one-liner stats\n"
            "`!compare lord1 lord2` — Compare two players\n"
            "`!gains [season] [user]` — View gains\n"
            "`/gain start_date end_date [user]` — Gains with date autocomplete\n"
            "`!active` — Active vs inactive members"
        ),
        inline=False
    )

    embed.add_field(
        name="🧠 Ask (Natural Language)",
        value=(
            "`!ask <question>` — leaderboards, progress, comparisons, seasons, events, pace, math & general help\n"
            "e.g. `!ask top 10 mana leaderboard`\n"
            "e.g. `!ask who is slacking this season`\n"
            "e.g. `!ask at my pace, how long until I hit 1B mana gathered`\n"
            "e.g. `!ask compare rekz and truvix`\n"
            "e.g. `!ask when does the season end`\n"
            "e.g. `!ask what's the next event`\n"
            "e.g. `!ask add events : August\\n* Aug 28 (Fri): — KvK Start\\n...` (admin only)\n"
            "e.g. `!ask at 0 UTC create a new season called \"Nvr vs Yss\"` (admin only, queues it)\n"
            "e.g. `!ask what tasks are queued` / `!ask cancel task 3`\n"
            "e.g. `!ask what's 15% of 480` or `!ask help me solve 2x+5=17`\n"
            "*Remembers recent conversation per channel for follow-ups*"
        ),
        inline=False
    )

    embed.add_field(
        name="🏆 Leaderboards",
        value=(
            "`!topmana [season]` — Top mana gathered\n"
            "`!topinf / !topcav / !topmage / !toparcher [season]` — Merit breakdowns (total + daily gain)\n"
            "`!topheal [season]` — Top T4/T5 RSS healed\n"
            "`!topspent [season]` — Top estimated mana spent (T4/T5 Healed × 72, assuming T5)\n"
            "`!topdeaths [season]` — Most deaths\n"
            "`!topmerits [season]` — Highest merits\n"
            "`!rss [season]` — Top resource spenders\n\n"
            "*All support an optional `[season]` — e.g. `!topmerits sos1`*"
        ),
        inline=False
    )

    embed.add_field(
        name="🌐 Server Lookup",
        value=(
            "`!servertop[N]` — Top N servers by power (e.g. `!servertop25`)\n"
            "`!servercheck[#]` — Find a server's rank (e.g. `!servercheck698`)"
        ),
        inline=False
    )

    embed.add_field(
        name="🖥️ Server Leaderboards",
        value=(
            "`!stopmerits / !stopdeaths [server] [top]` — Core stats\n"
            "`!stopheal [top]` — Top healing\n"
            "`!stopmana [top]` — Top estimated mana spent (Healing × 72)\n"
            "`!stopinf / !stopcav / !stopmage / !stoparcher / !stopother [server] [top]` — Merit breakdowns\n"
            "`!stoppower [server] [top]` — Current power\n"
            "`!stophighest [top]` — Historical highest power\n\n"
            "*Data comes from the latest Excel upload — see admin section below*"
        ),
        inline=False
    )

    embed.add_field(
        name="⚔️ KvK Matchup",
        value=(
            "`!kvkmatchup` — Interactive zone/team comparison\n"
            "`!matchups` — List all saved matchups\n"
            "`!matchup [id]` — View a saved matchup's details"
        ),
        inline=False
    )

    embed.add_field(
        name="📚 Historical Data",
        value=(
            "`!seasonhistory` — All seasons with dates\n"
            "`!datahistory` — Oldest/newest saved dates"
        ),
        inline=False
    )

    embed.add_field(
        name="👤 Roles & Events",
        value=(
            "`/addrole` — Get Abyss reminders\n"
            "`/removerole` — Stop Abyss reminders\n"
            "`/weeklyevent` — Show current weekly events\n"
            "`/kvkevent` — Show custom events"
        ),
        inline=False
    )

    # ============ ADMIN / OWNER COMMANDS ============
    if is_owner:
        embed.add_field(
            name="\u200b",
            value="\u200b",
            inline=False
        )
        embed.add_field(
            name="🔒 ​ A D M I N ​ /  ​ O W N E R ​ C O M M A N D S",
            value="\u200b",
            inline=False
        )

        embed.add_field(
            name="🖥️ Server Data",
            value="`!serverupdate` — Upload Excel file with server stats (**admin only**)\n`!delmatchup [id]` — Delete a saved KvK matchup",
            inline=False
        )

        embed.add_field(
            name="⚙️ Season & Data Management",
            value=(
                "`/newseason` — Start a new season (auto-ends the previous one)\n"
                "`!editseason` / `/editseason` — Edit a season's name, start date, or end date\n"
                "`/deleteseason` / `/removeseason` — Delete a season and its progress data\n"
                "`!forcefetch` — Fetch latest stats immediately\n"
                "`!loadhistory` — Load season data from season start\n"
                "`!loadhistory all` — Load all Call of Stats data (auto-detects oldest date)\n"
                "`!datahistory` — Show oldest/newest data range\n"
                "`!cleandata` — Delete empty/zero-data snapshots\n\n"
                "*Seasons have an auto ID (see `!seasonhistory`) — use it in place of a season name, e.g. `!progress rekz 1`*"
            ),
            inline=False
        )

        embed.add_field(
            name="📅 Event Management",
            value="`/addevent` — Add custom event\n`/editevent` — Edit event\n`/removeevent` — Delete event\n`/abyssconfig` — Configure Abyss settings",
            inline=False
        )

        embed.add_field(
            name="🛠️ System",
            value="`/testdm` — Test DM system\n`/backup` — List database backups\n`/forcebackup` — Create backup now\n`/setstatus text` — Set bot status (`default` = latest data date)\n`/say text` — Make the bot say something",
            inline=False
        )

    embed.set_footer(text="Use / to see all slash commands  •  Use ! for text commands")
    return embed


@bot.tree.command(name="help", description="Show all available commands")
async def help_cmd(inter):
    embed = build_help_embed(inter.user.id == OWNER_ID)
    await inter.response.send_message(embed=embed, ephemeral=True)


# Autocomplete for dates
async def date_autocomplete(
    interaction: discord.Interaction,
    current: str,
) -> list[app_commands.Choice[str]]:
    """Autocomplete dates from database"""
    try:
        conn = sqlite3.connect(DB_PROGRESS)
        c = conn.cursor()
        c.execute("SELECT DISTINCT data_date FROM season_progress ORDER BY data_date DESC")
        all_dates = [row[0] for row in c.fetchall()]
        conn.close()
        
        # Filter dates that start with current input
        filtered = [d for d in all_dates if d.startswith(current.lower())] if current else all_dates[:25]
        
        # Return top 25 matches
        return [
            app_commands.Choice(name=date, value=date)
            for date in filtered[:25]
        ]
    except Exception as e:
        log_error(f"[AUTOCOMPLETE] Error: {e}")
        return []


# ============================================================
# SEASON TRACKING COMMANDS
# ============================================================

class NewSeasonModal(Modal, title="📅 New Season"):
    season_name = TextInput(
        label="Season Name",
        placeholder="e.g., S053, Season 5, Spring 2026"
    )
    start_date = TextInput(
        label="Start Date (YYYY-MM-DD)",
        placeholder="2025-12-01"
    )

    async def on_submit(self, interaction: discord.Interaction):
        await interaction.response.defer(ephemeral=True)

        try:
            # Validate date format
            start_dt = datetime.strptime(self.start_date.value.strip(), "%Y-%m-%d")

            # Mark the previous active season as ended — end_date = day before this new season starts
            db_end_previous_active_season(self.start_date.value.strip())

            # Store season
            db_add_season(self.season_name.value.strip(), self.start_date.value.strip())
            
            await interaction.followup.send(
                f"✅ **Season Started!**\n"
                f"**Name:** {self.season_name.value}\n"
                f"**Start Date:** {self.start_date.value}\n\n"
                f"Use `!progress` to check your season stats!",
                ephemeral=True
            )
        except ValueError:
            await interaction.followup.send(
                "❌ Invalid date format. Use YYYY-MM-DD (e.g., 2025-12-01)",
                ephemeral=True
            )
        except Exception as e:
            log_info(f"[NEW SEASON ERROR] {e}")
            await interaction.followup.send(
                "❌ Failed to create season.",
                ephemeral=True
            )

@bot.tree.command(name="newseason", description="Start a new season and track progress")
async def newseason(inter: discord.Interaction):
    if inter.user.id != OWNER_ID:
        return await inter.response.send_message("❌ Owner only.", ephemeral=True)

    await inter.response.send_modal(NewSeasonModal())


# ============================================================
# EDIT SEASON COMMAND
# ============================================================

@bot.tree.command(name="editseason", description="Edit an existing season's name or start date")
@app_commands.default_permissions(administrator=True)
async def editseason(inter: discord.Interaction):
    if not inter.user.guild_permissions.administrator and inter.user.id != OWNER_ID:
        return await inter.response.send_message("❌ Admin only.", ephemeral=True)

    seasons = db_get_all_seasons()
    if not seasons:
        return await inter.response.send_message("❌ No seasons found.", ephemeral=True)

    options = [
        discord.SelectOption(
            label=f"#{s[0]} — {s[1]} (starts {s[2]})",
            value=str(s[0])
        )
        for s in seasons
    ]

    select = Select(placeholder="Select a season to edit", options=options)

    async def select_cb(i: discord.Interaction):
        season_id = int(select.values[0])
        season = next((s for s in seasons if s[0] == season_id), None)
        if not season:
            return await i.response.send_message("❌ Season not found.", ephemeral=True)

        class EditSeasonModal(Modal, title="✏️ Edit Season"):
            season_name = TextInput(
                label="Season Name",
                default=season[1]
            )
            start_date = TextInput(
                label="Start Date (YYYY-MM-DD)",
                default=season[2]
            )
            end_date = TextInput(
                label="End Date (YYYY-MM-DD, blank = ongoing)",
                default=db_get_season_end_date(season[0]) or "",
                required=False
            )

            async def on_submit(self, modal_inter: discord.Interaction):
                await modal_inter.response.defer(ephemeral=True)
                try:
                    datetime.strptime(self.start_date.value.strip(), "%Y-%m-%d")
                    end_date_val = self.end_date.value.strip()
                    if end_date_val:
                        datetime.strptime(end_date_val, "%Y-%m-%d")

                    conn = sqlite3.connect(DB)
                    c = conn.cursor()
                    c.execute(
                        "UPDATE seasons SET season_name=?, start_date=?, end_date=? WHERE id=?",
                        (self.season_name.value.strip(), self.start_date.value.strip(), end_date_val or None, season_id)
                    )
                    conn.commit()
                    conn.close()
                    silent_backup()

                    end_display = end_date_val if end_date_val else "Ongoing"
                    await modal_inter.followup.send(
                        f"✅ Season updated!\n**Name:** {self.season_name.value}\n**Start Date:** {self.start_date.value}\n**End Date:** {end_display}",
                        ephemeral=True
                    )
                except ValueError:
                    await modal_inter.followup.send("❌ Invalid date format. Use YYYY-MM-DD", ephemeral=True)
                except Exception as e:
                    await modal_inter.followup.send(f"❌ Error: {e}", ephemeral=True)

        await i.response.send_modal(EditSeasonModal())

    select.callback = select_cb
    view = View()
    view.add_item(select)
    await inter.response.send_message("Select a season to edit:", view=view, ephemeral=True)


@bot.command(name="editseason")
async def editseason_text(ctx):
    """[ADMIN ONLY] Edit a season's name, start date, or end date. Shows a season picker."""
    is_admin = ctx.author.id == OWNER_ID or (
        ctx.guild and ctx.author.guild_permissions.administrator
    )
    if not is_admin:
        return await ctx.send("❌ Admin only.")

    seasons = db_get_all_seasons()
    if not seasons:
        return await ctx.send("❌ No seasons found.")

    options = [
        discord.SelectOption(
            label=f"#{s[0]} — {s[1]} (starts {s[2]})",
            value=str(s[0])
        )
        for s in seasons
    ]

    select = Select(placeholder="Select a season to edit", options=options)

    async def select_cb(i: discord.Interaction):
        season_id = int(select.values[0])
        season = next((s for s in seasons if s[0] == season_id), None)
        if not season:
            return await i.response.send_message("❌ Season not found.", ephemeral=True)

        class EditSeasonModal(Modal, title="✏️ Edit Season"):
            season_name = TextInput(
                label="Season Name",
                default=season[1]
            )
            start_date = TextInput(
                label="Start Date (YYYY-MM-DD)",
                default=season[2]
            )
            end_date = TextInput(
                label="End Date (YYYY-MM-DD, blank = ongoing)",
                default=db_get_season_end_date(season[0]) or "",
                required=False
            )

            async def on_submit(self, modal_inter: discord.Interaction):
                await modal_inter.response.defer(ephemeral=True)
                try:
                    datetime.strptime(self.start_date.value.strip(), "%Y-%m-%d")
                    end_date_val = self.end_date.value.strip()
                    if end_date_val:
                        datetime.strptime(end_date_val, "%Y-%m-%d")

                    conn = sqlite3.connect(DB)
                    c = conn.cursor()
                    c.execute(
                        "UPDATE seasons SET season_name=?, start_date=?, end_date=? WHERE id=?",
                        (self.season_name.value.strip(), self.start_date.value.strip(), end_date_val or None, season_id)
                    )
                    conn.commit()
                    conn.close()
                    silent_backup()

                    end_display = end_date_val if end_date_val else "Ongoing"
                    await modal_inter.followup.send(
                        f"✅ Season updated!\n**Name:** {self.season_name.value}\n**Start Date:** {self.start_date.value}\n**End Date:** {end_display}",
                        ephemeral=True
                    )
                except ValueError:
                    await modal_inter.followup.send("❌ Invalid date format. Use YYYY-MM-DD", ephemeral=True)
                except Exception as e:
                    await modal_inter.followup.send(f"❌ Error: {e}", ephemeral=True)

        await i.response.send_modal(EditSeasonModal())

    select.callback = select_cb
    view = View()
    view.add_item(select)
    await ctx.send("Select a season to edit:", view=view)


# ============================================================
# DELETE SEASON COMMAND
# ============================================================

@bot.tree.command(name="deleteseason", description="Delete a season and all its progress data")
@app_commands.default_permissions(administrator=True)
async def deleteseason(inter: discord.Interaction):
    if not inter.user.guild_permissions.administrator and inter.user.id != OWNER_ID:
        return await inter.response.send_message("❌ Admin only.", ephemeral=True)

    seasons = db_get_all_seasons()
    if not seasons:
        return await inter.response.send_message("❌ No seasons found.", ephemeral=True)

    options = [
        discord.SelectOption(
            label=f"#{s[0]} — {s[1]} (starts {s[2]})",
            value=str(s[0])
        )
        for s in seasons
    ]

    select = Select(placeholder="Select a season to DELETE", options=options)

    async def select_cb(i: discord.Interaction):
        season_id = int(select.values[0])
        season = next((s for s in seasons if s[0] == season_id), None)
        if not season:
            return await i.response.send_message("❌ Season not found.", ephemeral=True)

        # Confirm button
        confirm_btn = Button(label=f"⚠️ Confirm Delete '{season[1]}'", style=discord.ButtonStyle.danger)

        async def confirm_cb(confirm_inter: discord.Interaction):
            try:
                # Delete season progress data
                conn_p = sqlite3.connect(DB_PROGRESS)
                c_p = conn_p.cursor()
                c_p.execute("DELETE FROM season_progress WHERE season_id=?", (season_id,))
                deleted_rows = c_p.rowcount
                conn_p.commit()
                conn_p.close()

                # Delete season record
                conn = sqlite3.connect(DB)
                c = conn.cursor()
                c.execute("DELETE FROM seasons WHERE id=?", (season_id,))
                conn.commit()
                conn.close()
                silent_backup()

                await confirm_inter.response.send_message(
                    f"🗑️ Deleted season **{season[1]}** and {deleted_rows} progress snapshots.",
                    ephemeral=True
                )
            except Exception as e:
                await confirm_inter.response.send_message(f"❌ Error: {e}", ephemeral=True)

        confirm_btn.callback = confirm_cb
        confirm_view = View()
        confirm_view.add_item(confirm_btn)

        await i.response.send_message(
            f"⚠️ Are you sure you want to delete **{season[1]}** (starts {season[2]})?\nThis will also delete all progress data for this season.",
            view=confirm_view,
            ephemeral=True
        )

    select.callback = select_cb
    view = View()
    view.add_item(select)
    await inter.response.send_message("Select a season to delete:", view=view, ephemeral=True)


@bot.tree.command(name="removeseason", description="Delete a season and all its progress data (alias of /deleteseason)")
@app_commands.default_permissions(administrator=True)
async def removeseason(inter: discord.Interaction):
    if not inter.user.guild_permissions.administrator and inter.user.id != OWNER_ID:
        return await inter.response.send_message("❌ Admin only.", ephemeral=True)

    seasons = db_get_all_seasons()
    if not seasons:
        return await inter.response.send_message("❌ No seasons found.", ephemeral=True)

    options = [
        discord.SelectOption(
            label=f"#{s[0]} — {s[1]} (starts {s[2]})",
            value=str(s[0])
        )
        for s in seasons
    ]

    select = Select(placeholder="Select a season to REMOVE", options=options)

    async def select_cb(i: discord.Interaction):
        season_id = int(select.values[0])
        season = next((s for s in seasons if s[0] == season_id), None)
        if not season:
            return await i.response.send_message("❌ Season not found.", ephemeral=True)

        confirm_btn = Button(label=f"⚠️ Confirm Remove '{season[1]}'", style=discord.ButtonStyle.danger)

        async def confirm_cb(confirm_inter: discord.Interaction):
            try:
                conn_p = sqlite3.connect(DB_PROGRESS)
                c_p = conn_p.cursor()
                c_p.execute("DELETE FROM season_progress WHERE season_id=?", (season_id,))
                deleted_rows = c_p.rowcount
                conn_p.commit()
                conn_p.close()

                conn = sqlite3.connect(DB)
                c = conn.cursor()
                c.execute("DELETE FROM seasons WHERE id=?", (season_id,))
                conn.commit()
                conn.close()
                silent_backup()

                await confirm_inter.response.send_message(
                    f"🗑️ Removed season **{season[1]}** and {deleted_rows} progress snapshots.",
                    ephemeral=True
                )
            except Exception as e:
                await confirm_inter.response.send_message(f"❌ Error: {e}", ephemeral=True)

        confirm_btn.callback = confirm_cb
        confirm_view = View()
        confirm_view.add_item(confirm_btn)

        await i.response.send_message(
            f"⚠️ Are you sure you want to remove **{season[1]}** (starts {season[2]})?\nThis will also delete all progress data for this season.",
            view=confirm_view,
            ephemeral=True
        )

    select.callback = select_cb
    view = View()
    view.add_item(select)
    await inter.response.send_message("Select a season to remove:", view=view, ephemeral=True)


@bot.command(name="progress")
async def progress(ctx, user_input: str = None, season_input: str = None):
    """
    Check season progress. 
    Usage: 
      !progress               (your progress, current season)
      !progress truvix        (truvix's progress, current season)
      !progress truvix sos1   (truvix's progress in sos1 season)
      !progress 16322115 sos2 (account 16322115's progress in sos2)
      !progress rekz 1        (rekz's progress in season ID 1)
    """
    # Determine which season to use
    if season_input:
        # Look up specific season by ID or name
        try:
            season = resolve_season_input(season_input)
            if not season:
                return await ctx.send(f"❌ Season '{season_input}' not found. Use `!seasonhistory` to see all seasons.")
        except Exception as e:
            return await ctx.send(f"❌ Error looking up season: {e}")
    else:
        # Use current season
        season = db_get_current_season()
        
        if not season:
            return await ctx.send("❌ No season active. Use `/newseason` to start one.")
    
    account_id = None
    
    # If no input provided, find from user's numeric role
    if not user_input:
        for role in ctx.author.roles:
            if role.name.isdigit():
                account_id = role.name
                break
        
        if not account_id:
            return await ctx.send("❌ You don't have a numeric role with your account ID.\nAsk the owner to give you a role with your account ID number (e.g., role name: `16322115`).\n\nOr use: `!progress truvix` or `!progress 16322115`")
    
    # If input is numeric, use as account ID
    elif user_input.isdigit():
        account_id = user_input
    
    # If input is text, check username lookup
    else:
        username_lower = user_input.lower()
        found_discord_id = None
        
        # Check if username exists in mapping
        for username, discord_id in USERNAME_TO_DISCORD_ID.items():
            if username.lower() == username_lower:
                found_discord_id = discord_id
                break
        
        if not found_discord_id:
            return await ctx.send(f"❌ Username '{user_input}' not found. Available usernames: {', '.join(USERNAME_TO_DISCORD_ID.keys())}")
        
        # Get the member from guild and find their numeric role
        try:
            member = ctx.guild.get_member(found_discord_id)
            if not member:
                return await ctx.send(f"❌ User '{user_input}' is not in this server.")
            
            for role in member.roles:
                if role.name.isdigit():
                    account_id = role.name
                    break
            
            if not account_id:
                return await ctx.send(f"❌ User '{user_input}' doesn't have a numeric role with their account ID.")
        except Exception as e:
            log_info(f"[PROGRESS USERNAME ERROR] {e}")
            return await ctx.send(f"❌ Error looking up user '{user_input}'")
    
    season_id, season_name, start_date, created_at = season
    
    msg = await ctx.send(f"📊 Fetching stats for account {account_id} in season {season_name}...")
    
    # For the CURRENT active season, always query up to today.
    # For an OLD/ended season, the query must be capped at that season's actual end_date —
    # otherwise the range bleeds into whatever season(s) came after it and COS returns
    # garbage (wrong totals, negative merits, wrong timespan display, etc).
    current_season_check = db_get_current_season()
    is_current_season = current_season_check and current_season_check[0] == season_id

    if is_current_season:
        today = date.today().isoformat()
    else:
        stored_end = db_get_season_end_date(season_id)
        if stored_end:
            today = stored_end
        else:
            # Legacy season with no stored end_date — fall back to last known data date
            conn = sqlite3.connect(DB_PROGRESS)
            c = conn.cursor()
            c.execute("SELECT MAX(data_date) FROM season_progress WHERE season_id = ?", (season_id,))
            row = c.fetchone()
            conn.close()
            today = row[0] if row and row[0] else date.today().isoformat()
        log_info(f"[PROGRESS] Old season #{season_id} — capping query end date to {today}")
    
    try:
        # FIRST: Try database for today
        stats = db_get_season_progress(season_id, account_id, today)
        
        # If not today, try yesterday
        if not stats:
            yesterday = (date.today() - timedelta(days=1)).isoformat()
            stats = db_get_season_progress(season_id, account_id, yesterday)
        
        # FALLBACK: Check cache
        if not stats:
            stats = get_cached_stats(account_id, start_date, today)
        
        # FALLBACK 2: Fetch from API (for unlisted people)
        if not stats:
            log_info(f"[PROGRESS] {account_id} not in DB, fetching from API")
            stats, end_date_used = await fetch_stats_with_fallback(account_id, start_date, today)
        else:
            end_date_used = stats.get("data_date", today)
        
        if not stats:
            return await msg.edit(content="❌ Failed to fetch stats. Call of Stats may not have released data yet.")

        # --- Fix missing/negative/zero stats by shifting the range's start date forward ---
        # COS's own range-computed delta (season_start -> today) can occasionally be wrong
        # for a specific account (e.g. after a server/alliance transfer, or when the stored
        # season start_date is off by even one day). When that happens, we re-run the SAME
        # query with the start date shifted forward one day at a time until COS itself
        # returns a valid number for that field.
        def _parse_stat_num(s):
            if not s:
                return 0
            try:
                return int(str(s).replace("+", "").replace(",", "").strip() or 0)
            except:
                return 0

        raw_power_gain = _parse_stat_num(stats.get("power_gain"))
        raw_merits = _parse_stat_num(stats.get("merits"))

        # Figure out which fields actually need fixing before doing any extra fetching
        needs_fix = {}
        if raw_power_gain <= 0:
            needs_fix["power_gain"] = "num"   # accept when value > 0
        if raw_merits <= 0:
            needs_fix["merits"] = "num"
        for field in ["kills_gain", "deads_gain", "healed_gain",
                      "gold_gathered", "wood_gathered", "ore_gathered", "mana_gathered",
                      "infantry_merits", "cavalry_merits", "mage_merits", "marksman_merits",
                      "other_merits", "t45_healed", "t45_dead"]:
            if stats.get(field) is None:
                needs_fix[field] = "missing"  # accept as soon as it's populated at all

        if needs_fix:
            # One shared day-by-day search — a single fetch per candidate day resolves
            # ALL still-missing fields at once, instead of each field re-fetching the
            # same URL independently (which was the main reason !progress got slow).
            try:
                cursor_date = datetime.strptime(start_date, "%Y-%m-%d") + timedelta(days=1)
                end_dt = datetime.strptime(end_date_used, "%Y-%m-%d")
            except Exception:
                cursor_date = None
                end_dt = None

            days_tried = 0
            max_days = 7
            while cursor_date and cursor_date < end_dt and needs_fix and days_tried < max_days:
                candidate_start = cursor_date.strftime("%Y-%m-%d")
                try:
                    snap, _ = await fetch_stats_with_fallback(account_id, candidate_start, end_date_used)
                except Exception:
                    snap = None

                if snap:
                    resolved_now = []
                    for field, mode in needs_fix.items():
                        val_raw = snap.get(field)
                        if mode == "num":
                            val = _parse_stat_num(val_raw)
                            if val > 0:
                                stats[field] = f"+{val:,}"
                                log_info(f"[PROGRESS] Corrected {field} for {account_id} using start_date={candidate_start}: {val}")
                                resolved_now.append(field)
                        else:  # "missing"
                            if val_raw is not None:
                                stats[field] = val_raw
                                log_info(f"[PROGRESS] Recovered missing {field} for {account_id} using start_date={candidate_start}")
                                resolved_now.append(field)
                    for f in resolved_now:
                        del needs_fix[f]

                cursor_date += timedelta(days=1)
                days_tried += 1

            for field in needs_fix:
                log_info(f"[PROGRESS] Could not resolve {field} for {account_id}, leaving as-is")

        # Advanced war stats (Infantry/Cavalry/Mage/Marksman/Other Merits, T4/T5 Healed,
        # T4/T5 Dead) are parsed by the SAME parse_stats() function as the core stats above,
        # from the SAME page — they're already in `stats` at this point (recovered by the
        # backtracking search above if they were initially missing). No separate fetch needed.
        stats_adv_today = stats
        adv_yesterday = end_date_used
        adv_is_up_to_date = is_current_season and end_date_used == today

        # For the "+X today" gain figures, grab yesterday's totals using the same core fetch
        stats_adv_prev = None
        try:
            prev_date = (datetime.strptime(end_date_used, "%Y-%m-%d") - timedelta(days=1)).strftime("%Y-%m-%d")
            stats_adv_prev, _ = await fetch_stats_with_fallback(account_id, start_date, prev_date)
        except Exception as e:
            log_info(f"[ADV STATS] Prev-day fetch for gain calc failed: {e}")

        def _adv_int(s):
            if not s:
                return 0
            try:
                return int(str(s).replace(",", "").replace("+", "").replace("-", "").strip())
            except:
                return 0

        def _adv_gain(field):
            t = _adv_int(stats_adv_today.get(field) if stats_adv_today else None)
            p = _adv_int(stats_adv_prev.get(field) if stats_adv_prev else None)
            return t - p if t and t > p else None

        def _adv_total(field):
            val = _adv_int(stats_adv_today.get(field) if stats_adv_today else None)
            if val:
                return val
            # This field might be missing from the resolved row even though other fields
            # on it are fine — look past it for the most recent date that has it.
            fallback = db_get_latest_field_value(season_id, account_id, field)
            return _adv_int(fallback) or None

        # "Only 1 day of data" should reflect the actual reported date RANGE (start_date to
        # end_date_used) spanning less than a day — e.g. an account added today. The raw DB
        # row count isn't a reliable signal for this: a missed daily save for one day doesn't
        # mean the account has only been tracked for 1 day, and using row count caused false
        # positives for long-tracked accounts that just had an incidental save gap somewhere.
        try:
            elapsed_days = (datetime.strptime(end_date_used, "%Y-%m-%d") - datetime.strptime(start_date, "%Y-%m-%d")).days
        except Exception:
            elapsed_days = 1
        is_single_day = elapsed_days < 1
        
        if is_single_day:
            log_info(f"[PROGRESS] Only 1 day in reported range for {account_id} in season {season_id} ({start_date} -> {end_date_used})")
        
        # Get highest power — for the CURRENT season use the live/today value,
        # but for an OLD/ended season use the power as it stood at that season's end date.
        # Check our own archive first (protects against COS deleting old dates), only
        # live-fetch if we don't have it yet, and save whatever we fetch permanently.
        current_season = db_get_current_season()
        is_current_season = current_season and current_season[0] == season_id
        season_end_date = db_get_season_end_date(season_id) or end_date_used
        power_date_key = end_date_used if is_current_season else season_end_date

        def _has_real_value(v):
            return v is not None and str(v).strip() not in ("", "0", "+0", "-0")

        async def _resolve_highest_power():
            db_row = db_get_season_progress(season_id, account_id, power_date_key)
            cached = db_row.get("highest_power") if db_row else None
            if _has_real_value(cached):
                return int(cached)
            val = await (fetch_highest_power(account_id) if is_current_season
                         else fetch_highest_power_at_date(account_id, season_end_date))
            return val

        async def _resolve_achievements():
            db_row = db_get_season_progress(season_id, account_id, end_date_used)
            cached_coins = db_row.get("exchange_coins_spent") if db_row else None
            if _has_real_value(cached_coins):
                cached_pets = db_row.get("max_pets") if db_row else None
                return {"exchange_coins_spent": int(cached_coins),
                        "max_pets": int(cached_pets) if _has_real_value(cached_pets) else None}
            return await fetch_achievement_stats(account_id, end_date_used)

        # These four are all independent of each other — run them concurrently instead
        # of one after another, since that was a meaningful chunk of !progress's runtime.
        highest_power, alliance_tag, current_t_kills, achievement_stats = await asyncio.gather(
            _resolve_highest_power(),
            fetch_alliance_tag(account_id),
            fetch_current_t_kills(account_id),
            _resolve_achievements(),
        )

        if not is_current_season:
            log_info(f"[PROGRESS] Old season #{season_id} — using highest power @ {season_end_date}: {highest_power}")

        if _has_real_value(highest_power):
            db_save_extra_stats(season_id, account_id, power_date_key, highest_power=highest_power)

        if achievement_stats["exchange_coins_spent"] is None and achievement_stats["max_pets"] is None:
            # COS may not have a page for a date it hasn't scanned yet (e.g. today) — try yesterday
            fallback_date = (datetime.strptime(end_date_used, "%Y-%m-%d") - timedelta(days=1)).strftime("%Y-%m-%d")
            log_info(f"[ACHIEVEMENTS] No data for {end_date_used}, trying {fallback_date}")
            achievement_stats = await fetch_achievement_stats(account_id, fallback_date)
        exchange_coins_spent = achievement_stats["exchange_coins_spent"]
        max_pets = achievement_stats["max_pets"]

        if exchange_coins_spent is not None or max_pets is not None:
            db_save_extra_stats(season_id, account_id, end_date_used,
                                 exchange_coins_spent=exchange_coins_spent, max_pets=max_pets)

        # Also fetch coins spent AT the season's start date, to show the increase like Power does
        exchange_coins_gain = None
        if exchange_coins_spent is not None:
            db_row_start = db_get_season_progress(season_id, account_id, start_date)
            cached_baseline = db_row_start.get("exchange_coins_spent") if db_row_start else None
            if _has_real_value(cached_baseline):
                baseline_coins = int(cached_baseline)
            else:
                baseline_achievement_stats = await fetch_achievement_stats(account_id, start_date)
                baseline_coins = baseline_achievement_stats["exchange_coins_spent"]
                if baseline_coins is not None:
                    db_save_extra_stats(season_id, account_id, start_date, exchange_coins_spent=baseline_coins)
            if baseline_coins is not None and exchange_coins_spent >= baseline_coins:
                exchange_coins_gain = exchange_coins_spent - baseline_coins
        
        # Calculate merit to power ratio using highest power and merits
        if stats.get("merits") and highest_power:
            try:
                merits_str = stats["merits"].replace("+", "").replace(",", "")
                merits_val = int(merits_str) if merits_str.isdigit() else 0
                
                # Calculate ratio: (Merits / Highest Power) × 100
                if highest_power > 0:
                    ratio = (merits_val / highest_power) * 100
                    stats["merits_pct"] = f"{ratio:.1f}%"
                else:
                    stats["merits_pct"] = "0%"
            except Exception as e:
                stats["merits_pct"] = "0%"
        else:
            stats["merits_pct"] = "0%"
        
        # Debug: log what we parsed
        log_info(f"[PROGRESS] Parsed stats: {stats}")
        log_info(f"[PROGRESS] Highest power: {highest_power}")
        log_info(f"[PROGRESS] Merit ratio: {stats.get('merits_pct')}")
        log_info(f"[PROGRESS] Current T-kills: {current_t_kills}")
        
        # Get rankings for all stats
        power_rank = await get_rankings_for_stat(ctx, "power_gain", start_date, end_date_used)
        merits_rank = await get_rankings_for_stat(ctx, "merits", start_date, end_date_used)
        kills_rank = await get_rankings_for_stat(ctx, "kills_gain", start_date, end_date_used)
        deads_rank = await get_rankings_for_stat(ctx, "deads_gain", start_date, end_date_used)
        healed_rank = await get_rankings_for_stat(ctx, "healed_gain", start_date, end_date_used)
        infantry_rank = await get_rankings_for_stat(ctx, "infantry_merits", start_date, end_date_used)
        cavalry_rank = await get_rankings_for_stat(ctx, "cavalry_merits", start_date, end_date_used)
        mage_rank = await get_rankings_for_stat(ctx, "mage_merits", start_date, end_date_used)
        marksman_rank = await get_rankings_for_stat(ctx, "marksman_merits", start_date, end_date_used)
        other_rank = await get_rankings_for_stat(ctx, "other_merits", start_date, end_date_used)
        
        # Calculate power_gain
        power_gain = 0
        if stats["power_gain"]:
            power_gain_str = stats["power_gain"].replace("+", "")
            try:
                power_gain = int(power_gain_str.replace(",", ""))
            except Exception as e:
                power_gain = 0
        
        # Build ranking strings
        power_rank_str = f" (#{power_rank[account_id][0]})" if account_id in power_rank else ""
        merits_rank_str = f" (#{merits_rank[account_id][0]})" if account_id in merits_rank else ""
        kills_rank_str = f" (#{kills_rank[account_id][0]})" if account_id in kills_rank else ""
        deads_rank_str = f" (#{deads_rank[account_id][0]})" if account_id in deads_rank else ""
        healed_rank_str = f" (#{healed_rank[account_id][0]})" if account_id in healed_rank else ""
        infantry_rank_str = f" (#{infantry_rank[account_id][0]})" if account_id in infantry_rank else ""
        cavalry_rank_str = f" (#{cavalry_rank[account_id][0]})" if account_id in cavalry_rank else ""
        mage_rank_str = f" (#{mage_rank[account_id][0]})" if account_id in mage_rank else ""
        marksman_rank_str = f" (#{marksman_rank[account_id][0]})" if account_id in marksman_rank else ""
        other_rank_str = f" (#{other_rank[account_id][0]})" if account_id in other_rank else ""
        
        
        # Calculate totals for RSS
        total_spent = 0
        total_gathered = 0
        
        for key in ["gold_spent", "wood_spent", "ore_spent", "mana_spent"]:
            val_str = (stats.get(key) or "+0").replace(",", "").replace("+", "")
            try:
                total_spent += abs(int(val_str)) if val_str.lstrip("-").isdigit() else 0
            except Exception as e:
                pass
        
        for key in ["gold_gathered", "wood_gathered", "ore_gathered", "mana_gathered"]:
            val_str = (stats.get(key) or "+0").replace(",", "").replace("+", "")
            try:
                total_gathered += abs(int(val_str)) if val_str.lstrip("-").isdigit() else 0
            except Exception as e:
                pass
        
        # Build Discord embed - matches the reference "card" layout
        lord_name = stats.get("lord_name", "Unknown")

        embed = discord.Embed(
            title=f"📈 Progress Report for {alliance_tag} {lord_name}",
            color=0x2ecc71
        )
        embed.set_author(name=f"Season: {season_name}")

        if is_single_day:
            embed.description = "⚠️ Only 1 day of data available - showing absolute values"

        # Power + Merits + Merit Ratio combined into one compact field
        power_line = ""
        if highest_power or power_gain:
            if highest_power and power_gain:
                power_line = f"🟢 **Power:** {highest_power:,} (+{power_gain:,}){power_rank_str}"
            elif highest_power:
                power_line = f"🟢 **Power:** {highest_power:,}{power_rank_str}"
            else:
                power_line = f"🟢 **Power:** +{power_gain:,}{power_rank_str}"

        merits_line = ""
        if stats.get("merits"):
            pct = f" ({stats['merits_pct']})" if stats.get("merits_pct") else ""
            merits_line = f"🔴 **Merits:** {stats['merits']}{pct}{merits_rank_str}"

        if power_line or merits_line:
            embed.add_field(name="\u200b", value="\n".join(l for l in [power_line, merits_line] if l), inline=False)

        # Kills / Deads / T4/T5 Units Rss Healed combined into one line
        t45_healed_val = stats_adv_today.get("t45_healed") if stats_adv_today else None
        if not t45_healed_val:
            # The resolved row might be missing this one specific field even though other
            # fields on it are fine — look past it for the most recent date that has it.
            t45_healed_val = db_get_latest_field_value(season_id, account_id, "t45_healed")
            if t45_healed_val:
                log_info(f"[PROGRESS] t45_healed fallback found in DB: {t45_healed_val}")

        if not t45_healed_val:
            # DB genuinely has no good copy anywhere — make one final LIVE attempt using the
            # same shift-start-date search that already works for the core stats, since
            # fetch_stats_with_fallback/parse_stats is what actually parses this field
            # correctly (not the old dedicated fetch_advanced_stats_ranged helper).
            try:
                cursor_date = datetime.strptime(start_date, "%Y-%m-%d") + timedelta(days=1)
                end_dt = datetime.strptime(end_date_used, "%Y-%m-%d")
                days_tried = 0
                while cursor_date < end_dt and days_tried < 7 and not t45_healed_val:
                    candidate_start = cursor_date.strftime("%Y-%m-%d")
                    live_retry, _ = await fetch_stats_with_fallback(account_id, candidate_start, end_date_used)
                    if live_retry and live_retry.get("t45_healed") and live_retry["t45_healed"] not in ("+0", "-0", "0"):
                        t45_healed_val = live_retry["t45_healed"]
                        log_info(f"[PROGRESS] t45_healed live retry found using start_date={candidate_start}: {t45_healed_val}")
                        db_save_advanced_stats(season_id, account_id, end_date_used, live_retry, stats.get("lord_name"))
                    cursor_date += timedelta(days=1)
                    days_tried += 1
            except Exception as e:
                log_info(f"[PROGRESS] t45_healed live retry failed: {e}")

        combat_parts = []
        if stats.get("kills_gain"):
            combat_parts.append(f"⚔️ {stats['kills_gain']}{kills_rank_str}")
        if stats.get("deads_gain"):
            combat_parts.append(f"💀 {stats['deads_gain']}{deads_rank_str}")
        if t45_healed_val:
            combat_parts.append(f"❤️ {t45_healed_val}{healed_rank_str}")
        if combat_parts:
            embed.add_field(name="Kills / Deads / T4/T5 Units Rss Healed", value="  ·  ".join(combat_parts), inline=False)

        # Mana Spent — estimated the same way as !stopmana used to, assuming it's all T5.
        # Guild-side !progress only has the COMBINED T4/T5 Healed figure (unlike the server
        # Excel export used by !stopmana, which has the real T4/T5 split) — so this can't be
        # computed exactly here, only estimated using the real T5 rate as a conservative guess.
        mana_source = _parse_stat_num(t45_healed_val)
        if mana_source:
            mana_spent_est = mana_source * MANA_PER_T5_HEAL
            embed.add_field(
                name="💧 Mana Spent (est.)",
                value=f"+{mana_spent_est:,} _(assuming its T5)_",
                inline=False
            )

        # RSS Gathered — single compact line
        rss_parts = []
        if stats.get("gold_gathered"):
            rss_parts.append(f"🪙{stats['gold_gathered'].lstrip('+')}")
        if stats.get("wood_gathered"):
            rss_parts.append(f"🪵{stats['wood_gathered'].lstrip('+')}")
        if stats.get("ore_gathered"):
            rss_parts.append(f"⛏️{stats['ore_gathered'].lstrip('+')}")
        if stats.get("mana_gathered"):
            rss_parts.append(f"💧{stats['mana_gathered'].lstrip('+')}")
        if rss_parts:
            embed.add_field(
                name="🧑‍🌾 RSS Gathered",
                value=f"{'  '.join(rss_parts)}  ·  **Total: {total_gathered:,}**",
                inline=False
            )

        # Troop Merits (Server Rank) — two-column style
        def _fmt(n):
            return f"{n:,}" if n else None

        troop_lines = []
        infantry_total = _adv_total("infantry_merits")
        cavalry_total = _adv_total("cavalry_merits")
        marksman_total = _adv_total("marksman_merits")
        mage_total = _adv_total("mage_merits")
        other_total = _adv_total("other_merits")

        if infantry_total:
            troop_lines.append(f"⚔️ Infantry: {_fmt(infantry_total)}{infantry_rank_str}")
        if cavalry_total:
            troop_lines.append(f"🐴 Cavalry: {_fmt(cavalry_total)}{cavalry_rank_str}")
        if marksman_total:
            troop_lines.append(f"🏹 Archer: {_fmt(marksman_total)}{marksman_rank_str}")
        if mage_total:
            troop_lines.append(f"🔮 Magic: {_fmt(mage_total)}{mage_rank_str}")

        other_lines = []
        if other_total:
            other_lines.append(f"🌀 Other: {_fmt(other_total)}{other_rank_str}")
        t45_dead_total = _adv_total("t45_dead")
        if t45_dead_total:
            other_lines.append(f"💀 T4/T5 Dead: {_fmt(t45_dead_total)}")

        if troop_lines:
            embed.add_field(name="Troop Merits (Server Rank)", value="\n".join(troop_lines), inline=True)
        if other_lines:
            embed.add_field(name="Other (Server Rank)", value="\n".join(other_lines), inline=True)
        if not troop_lines and not other_lines and is_current_season:
            embed.add_field(name="Troop Merits (Server Rank)", value="_(not yet available from COS)_", inline=False)

        # Achievements — single compact line
        if exchange_coins_spent is not None and exchange_coins_gain is not None:
            coins_str = f"{exchange_coins_spent:,} (+{exchange_coins_gain:,})"
        elif exchange_coins_spent is not None:
            coins_str = f"{exchange_coins_spent:,}"
        else:
            coins_str = "n/a"
        pets_str = f"{max_pets:,}" if max_pets is not None else "n/a"
        embed.add_field(
            name="🏆 Achievements",
            value=f"🪙 Coins Spent: {coins_str}  ·  🐾 Max Pets: {pets_str}",
            inline=False
        )

        # Footer: timespan + advanced-stats freshness note + tip — all as small footer text, not a field
        footer_parts = [f"📅 {start_date} → {end_date_used}"]
        if troop_lines or other_lines:
            if is_current_season and not adv_is_up_to_date:
                footer_parts.append(f"Merit breakdown as of {adv_yesterday}")
            elif not is_current_season:
                footer_parts.append(f"Merit breakdown final as of {adv_yesterday}")
        footer_parts.append(f"Tip: !progress {account_id} sos2 (or a season ID) for another season")
        embed.set_footer(text="  •  ".join(footer_parts))

        await msg.edit(content=None, embed=embed)
        
    except Exception as e:
        log_info(f"[PROGRESS ERROR] {e}")
        await msg.edit(content=f"❌ Error: {str(e)}")




@bot.tree.command(name="addlord", description="[DEPRECATED] Use numeric roles instead")
async def addlord_deprecated(inter: discord.Interaction):
    """Deprecated - use numeric roles instead"""
    await inter.response.send_message(
        "❌ `/addlord` is deprecated!\n\n"
        "Instead, create a Discord role with your account ID as the name (e.g., `16322115`)\n"
        "The bot will auto-detect all numeric roles!\n\n"
        "Use `/forcefetch` to fetch everyone's stats.",
        ephemeral=True
    )




@bot.command(name="forcefetch")
async def forcefetch(ctx):
    """Force fetch stats for all members with numeric roles (Owner only)"""
    if ctx.author.id != OWNER_ID:
        return await ctx.send("❌ Owner only.")
    
    season = db_get_current_season()
    if not season:
        return await ctx.send("❌ No season active. Use `/newseason` to start one.")
    
    season_id, season_name, start_date, created_at = season
    today = date.today().isoformat()
    
    # Get all lords from server members
    lords = get_all_lords_from_guild(ctx.guild)
    if not lords:
        return await ctx.send("❌ No members with numeric roles found.")
    
    await ctx.send(f"⏳ Fetching stats for {len(lords)} members from season {season_name}...")
    
    fetched = 0
    failed = 0
    
    for lord in lords:
        try:
            account_id = lord["account_id"]
            name = lord["name"]
            
            stats = await fetch_stats_for_account(account_id, start_date, today)
            if stats:
                fetched += 1
                log_info(f"[FORCEFETCH] ✅ {name} ({account_id})")
            else:
                failed += 1
                log_info(f"[FORCEFETCH] ❌ {name} ({account_id})")
        except Exception as e:
            failed += 1
            log_info(f"[FORCEFETCH] ERROR {name}: {e}")
    
    embed = discord.Embed(
        title="📊 Force Fetch Complete",
        description=f"Season: {season_name}",
        color=0x2ecc71
    )
    embed.add_field(name="✅ Fetched", value=str(fetched), inline=True)
    embed.add_field(name="❌ Failed", value=str(failed), inline=True)
    embed.add_field(name="💾 Cache", value=f"All data cached for 3 days", inline=False)
    embed.set_footer(text=f"Fetched: {start_date} → {today}")
    
    await ctx.send(embed=embed)

    # Also check whether the bot's status/presence date needs updating, since forcefetch
    # is often run specifically because someone noticed new data is out but the status
    # is still showing the old date.
    await check_callofstats_update()


@bot.command(name="loadhistory")
async def loadhistory(ctx, mode: str = None):
    """
    [OWNER ONLY]
    Load historical data - DAILY SNAPSHOTS.
    
    Usage:
      !loadhistory        (loads from season start to today)
      !loadhistory all    (finds oldest Call of Stats data and loads EVERYTHING to today!)
    """
    if ctx.author.id != OWNER_ID:
        return await ctx.send("❌ Owner only.")
    
    season = db_get_current_season()
    if not season:
        return await ctx.send("❌ No season active.")
    
    season_id, season_name, season_start_date, created_at = season
    today = date.today()
    
    # Get all members to fetch for
    lords = get_all_lords_from_guild(ctx.guild)
    for discord_id, account_id in DISCORD_TO_ACCOUNT_ID.items():
        if not any(l["account_id"] == account_id for l in lords):
            lords.append({"account_id": account_id, "name": f"Account {account_id}"})
    
    # Determine start date
    if mode and mode.lower() == "all":
        msg = await ctx.send(f"🔍 Scanning Call of Stats starting from Dec 1, 2025...\nThis may take 2-3 minutes...")
        
        # Start from December 1, 2025
        start = date(2025, 12, 1)
        date_range_text = f"{start.isoformat()} → {today.isoformat()} (ALL available data!)"
        load_mode = "all"
        
    else:
        start = datetime.strptime(season_start_date, "%Y-%m-%d").date()
        date_range_text = f"{start.isoformat()} → {today.isoformat()}"
        load_mode = "season"
        msg = await ctx.send(f"⏳ Loading historical data...\n📅 {date_range_text}")
    
    total_days = (today - start).days + 1
    
    await msg.edit(content=f"⏳ Loading historical data...\n📅 {date_range_text}\n👥 {len(lords)} members × {total_days} days = {total_days * len(lords)} snapshots\nThis may take 5-30 minutes depending on data size...")
    
    # Fetch data for each day - fetch EACH DAY INDIVIDUALLY
    current_date = start
    saved_count = 0
    skipped_count = 0
    day_num = 0
    failed_count = 0
    
    while current_date <= today:
        day_num += 1
        date_str = current_date.isoformat()
        
        for lord in lords:
            try:
                account_id = lord["account_id"]
                name = lord.get("name", account_id)
                
                # CHECK: Does this snapshot already exist?
                if db_snapshot_exists(season_id, account_id, date_str):
                    # Already saved, skip it!
                    skipped_count += 1
                    continue
                
                # Fetch for THIS DAY ONLY (date_str to date_str)
                stats, actual_date = await fetch_stats_with_fallback(account_id, date_str, date_str)
                
                # CHECK: Is the data empty/all zeros? If so, DON'T save it
                if is_stats_empty(stats):
                    # Empty data, skip but don't delete anything
                    log_info(f"[LOADHISTORY] Skipping empty data for {account_id} on {date_str}")
                    skipped_count += 1
                    continue
                
                if stats:
                    # Save to database with the correct date
                    db_save_season_progress(season_id, account_id, stats.get("lord_name", name), stats, actual_date)
                    saved_count += 1
                else:
                    failed_count += 1
                    
            except Exception as e:
                log_error(f"[LOADHISTORY] Error for {account_id} on {date_str}: {e}")
                failed_count += 1
                continue
        
        # Update progress message every 5 days or on last day
        if day_num % 5 == 0 or current_date == today:
            progress = (day_num / total_days) * 100
            try:
                await msg.edit(content=f"⏳ Loading historical data...\n📅 {date_range_text}\nDay {day_num}/{total_days} ({progress:.0f}%)\n✅ Saved {saved_count} | ⏭️ Skipped {skipped_count}")
            except:
                pass
        
        current_date += timedelta(days=1)
    
    # Create result embed
    embed = discord.Embed(
        title="📚 Historical Data Load Complete",
        description=f"Season: {season_name}",
        color=0x2ecc71
    )
    embed.add_field(name="📅 Date Range", value=date_range_text, inline=False)
    embed.add_field(name="🎯 Load Mode", value="ALL Available Data" if load_mode == "all" else "Season Data", inline=True)
    embed.add_field(name="👥 Members Tracked", value=str(len(lords)), inline=True)
    embed.add_field(name="📊 Snapshots Saved", value=str(saved_count), inline=True)
    embed.add_field(name="⏭️ Snapshots Skipped", value=str(skipped_count), inline=True)
    embed.add_field(name="📈 Total Days Covered", value=str(total_days), inline=True)
    embed.add_field(name="✅ Status", value="Complete database created!\n\n🎯 You can now use `!gains` with dates from this entire range!", inline=False)
    embed.set_footer(text="!gains is now fully powered with all available historical data")
    
    await ctx.send(embed=embed)
    log_info(f"[LOADHISTORY] Complete! Mode={load_mode}, Saved {saved_count}, Skipped {skipped_count} from {start.isoformat()} to {today.isoformat()}")


@bot.command(name="seasonhistory")
async def seasonhistory(ctx):
    """
    Show all seasons with their ID, start date and end date (or ongoing if active).
    """
    try:
        conn = sqlite3.connect(DB)
        c = conn.cursor()
        c.execute("SELECT id, season_name, start_date, created_at, end_date FROM seasons ORDER BY created_at ASC")
        seasons = c.fetchall()
        conn.close()
        
        if not seasons:
            return await ctx.send("❌ No seasons found in database.")
        
        current_season = db_get_current_season()
        current_season_id = current_season[0] if current_season else None
        today = date.today().isoformat()
        
        # Build output
        output = "```📅 SEASON HISTORY\n\n"
        
        for season_id, season_name, start_date, created_at, end_date_val in seasons:
            is_current = (season_id == current_season_id)
            
            if end_date_val:
                # Explicit end_date stored (set automatically when a new season starts, or edited manually)
                end_display = end_date_val
                status = "✅ ENDED"
            elif is_current:
                # Current season, no end_date yet: show start date → today
                end_display = today
                status = "🔴 ACTIVE"
            else:
                # Legacy season with no end_date recorded: fall back to last day with data
                conn = sqlite3.connect(DB_PROGRESS)
                c = conn.cursor()
                c.execute(
                    "SELECT MAX(data_date) FROM season_progress WHERE season_id = ?",
                    (season_id,)
                )
                row = c.fetchone()
                conn.close()
                end_display = row[0] if row and row[0] else "Unknown"
                status = "✅ ENDED"
            
            output += f"{status} #{season_id} — {season_name}\n"
            output += f"   📍 {start_date} → {end_display}\n\n"
        
        output += "```"
        await ctx.send(output)
        
    except Exception as e:
        log_error(f"[SEASONHISTORY ERROR] {e}")
        await ctx.send(f"❌ Error: {str(e)}")




@bot.command(name="datahistory")
async def datahistory(ctx):
    """
    [OWNER ONLY]
    Show the oldest and newest data saved in database.
    Useful to know the date range for !gains with specific dates.
    """
    if ctx.author.id != OWNER_ID:
        return await ctx.send("❌ Owner only.")
    
    try:
        conn = sqlite3.connect(DB_PROGRESS)
        c = conn.cursor()
        
        # Get oldest and newest dates
        c.execute("SELECT MIN(data_date), MAX(data_date), COUNT(*) FROM season_progress")
        row = c.fetchone()
        conn.close()
        
        oldest_date, newest_date, total_snapshots = row if row else (None, None, 0)
        
        if not oldest_date:
            return await ctx.send("❌ No data found in database.")
        
        # Create result embed
        embed = discord.Embed(
            title="📊 Database History",
            description="Complete data range saved",
            color=0x3498db
        )
        embed.add_field(name="📅 Oldest Data", value=f"`{oldest_date}`", inline=True)
        embed.add_field(name="📅 Newest Data", value=f"`{newest_date}`", inline=True)
        embed.add_field(name="📦 Total Snapshots", value=str(total_snapshots), inline=True)
        embed.add_field(
            name="📝 Usage Example",
            value=f"`!gains {oldest_date} {newest_date}`\nor\n`!gains {oldest_date} {newest_date} rekz`",
            inline=False
        )
        embed.set_footer(text="You can use any date within this range with !gains command")
        
        await ctx.send(embed=embed)
        log_info(f"[DATAHISTORY] {oldest_date} → {newest_date} ({total_snapshots} snapshots)")
        
    except Exception as e:
        log_error(f"[DATAHISTORY ERROR] {e}")
        await ctx.send(f"❌ Error: {str(e)}")


class CleanupModal(discord.ui.Modal, title="🗑️ Delete Data"):
    """Modal for cleanup inputs"""
    date_input = discord.ui.TextInput(
        label="Date (YYYY-MM-DD) - skip if using 'current'",
        placeholder="e.g., 2026-03-13 (optional)",
        min_length=0,
        max_length=10,
        required=False
    )
    mode_input = discord.ui.TextInput(
        label="Mode: before, after, or current",
        placeholder="Type: before OR after OR current",
        min_length=5,
        max_length=7
    )
    
    async def on_submit(self, interaction: discord.Interaction):
        await interaction.response.defer()
        
        try:
            date_str = self.date_input.value.strip() if self.date_input.value else None
            mode = self.mode_input.value.strip().lower()
            
            if mode not in ["before", "after", "current"]:
                return await interaction.followup.send("❌ Mode must be 'before', 'after', or 'current'")
            
            # If mode is "current", get season start date
            if mode == "current":
                season = db_get_current_season()
                if not season:
                    return await interaction.followup.send("❌ No active season found")
                date_str = season[2]  # season_start_date
                mode_display = "before current season start"
                delete_mode = "before"
            else:
                if not date_str:
                    return await interaction.followup.send("❌ Date is required for 'before' or 'after' mode")
                
                try:
                    datetime.strptime(date_str, "%Y-%m-%d")
                except ValueError:
                    return await interaction.followup.send("❌ Invalid date format. Use YYYY-MM-DD")
                
                mode_display = mode
                delete_mode = mode
            
            conn = sqlite3.connect(DB_PROGRESS)
            c = conn.cursor()
            
            if delete_mode == "before":
                c.execute("DELETE FROM season_progress WHERE data_date < ?", (date_str,))
            else:
                c.execute("DELETE FROM season_progress WHERE data_date > ?", (date_str,))
            
            deleted_count = c.rowcount
            conn.commit()
            conn.close()
            
            embed = discord.Embed(
                title="🗑️ Data Cleanup Complete",
                description=f"Deleted data {mode_display}",
                color=0x2ecc71
            )
            embed.add_field(name="📅 Cutoff Date", value=date_str, inline=True)
            embed.add_field(name="🎯 Mode", value=mode.capitalize(), inline=True)
            embed.add_field(name="🗑️ Rows Deleted", value=str(deleted_count), inline=True)
            
            await interaction.followup.send(embed=embed)
            log_info(f"[CLEANUP] Deleted {deleted_count} snapshots {mode_display}")
            
        except Exception as e:
            log_error(f"[CLEANUP ERROR] {e}")
            await interaction.followup.send(f"❌ Error: {str(e)}")


@bot.tree.command(name="cleandata", description="Delete data before or after a specific date")
@app_commands.default_permissions(administrator=True)
async def cleandata_slash(interaction: discord.Interaction):
    """Delete data - Admin only"""
    if not interaction.user.guild_permissions.administrator:
        return await interaction.response.send_message("❌ Admin only.", ephemeral=True)
    
    modal = CleanupModal()
    await interaction.response.send_modal(modal)


@bot.command(name="cleandata")
async def cleandata_text(ctx):
    """[ADMIN ONLY] Use /cleandata slash command instead"""
    if not ctx.author.guild_permissions.administrator:
        return await ctx.send("❌ Admin only.")
    await ctx.send("Use the slash command `/cleandata` instead!")


class GainsDateSelector(discord.ui.View):
    """Interactive selector for gains date range, with a season dropdown to switch seasons"""
    
    def __init__(self, available_dates, account_id, season_id, start_date, ctx, all_seasons=None):
        super().__init__(timeout=120)
        self.available_dates = available_dates
        self.account_id = account_id
        self.season_id = season_id
        self.start_date = start_date
        self.ctx = ctx
        self.all_seasons = all_seasons or []
        
        self.selected_start = None
        self.selected_end = None
        
        self._build_components()

    def _build_components(self):
        """(Re)build the season select, date selects, and button based on current state."""
        self.clear_items()

        # Season dropdown — switching this rebuilds the date dropdowns for that season
        if self.all_seasons:
            season_options = [
                discord.SelectOption(
                    label=f"#{s[0]} — {s[1]} (starts {s[2]})",
                    value=str(s[0]),
                    default=(s[0] == self.season_id)
                )
                for s in self.all_seasons
            ]
            season_select = Select(placeholder="📁 Pick a season", options=season_options)
            season_select.callback = self.on_season_select
            self.add_item(season_select)

        available_dates = self.available_dates
        total_days = len(available_dates)
        
        if total_days <= 25:
            start_dates = available_dates
            end_dates = available_dates
            start_label = "📅 Pick Start Date"
            end_label = "📅 Pick End Date"
        elif total_days <= 50:
            start_dates = available_dates[:25]
            end_dates = available_dates[-25:]
            start_label = f"📅 Pick Start Date (oldest {len(start_dates)})"
            end_label = f"📅 Pick End Date (latest {len(end_dates)})"
        else:
            start_dates = available_dates[:25]
            end_dates = available_dates[-25:]
            start_label = "📅 Pick Start Date (oldest 25)"
            end_label = "📅 Pick End Date (latest 25)"
        
        start_select = Select(
            placeholder=start_label,
            min_values=1,
            max_values=1,
            options=[discord.SelectOption(label=d, value=d) for d in start_dates] if start_dates else [discord.SelectOption(label="No data", value="none")]
        )
        start_select.callback = self.on_start_select
        self.add_item(start_select)
        
        end_select = Select(
            placeholder=end_label,
            min_values=1,
            max_values=1,
            options=[discord.SelectOption(label=d, value=d) for d in end_dates] if end_dates else [discord.SelectOption(label="No data", value="none")]
        )
        end_select.callback = self.on_end_select
        self.add_item(end_select)

        show_button = Button(label="📊 Show Gains", style=discord.ButtonStyle.green)
        show_button.callback = self.show_gains
        self.add_item(show_button)

    async def on_season_select(self, interaction: discord.Interaction):
        new_season_id = int(interaction.data["values"][0])
        season = next((s for s in self.all_seasons if s[0] == new_season_id), None)
        if not season:
            return await interaction.response.send_message("❌ Season not found.", ephemeral=True)

        self.season_id = new_season_id
        self.start_date = season[2]
        self.selected_start = None
        self.selected_end = None

        conn = sqlite3.connect(DB_PROGRESS)
        c = conn.cursor()
        c.execute(
            "SELECT DISTINCT data_date FROM season_progress WHERE season_id=? AND account_id=? ORDER BY data_date ASC",
            (self.season_id, self.account_id)
        )
        self.available_dates = [row[0] for row in c.fetchall()]
        conn.close()

        self._build_components()

        if not self.available_dates:
            await interaction.response.edit_message(
                content=f"⚠️ No saved data found for **{season[1]}**. Pick another season or run `!loadhistory`.",
                view=self
            )
        else:
            await interaction.response.edit_message(
                content=f"📁 Season switched to **{season[1]}** — {len(self.available_dates)} days of data available.",
                view=self
            )
    
    async def on_start_select(self, interaction: discord.Interaction):
        self.selected_start = interaction.data["values"][0]
        await interaction.response.defer()
    
    async def on_end_select(self, interaction: discord.Interaction):
        self.selected_end = interaction.data["values"][0]
        await interaction.response.defer()
    
    async def show_gains(self, interaction: discord.Interaction, button: discord.ui.Button = None):
        if not self.selected_start or not self.selected_end:
            return await interaction.response.send_message("❌ Please select both dates", ephemeral=True)
        
        if self.selected_start > self.selected_end:
            return await interaction.response.send_message("❌ Start date must be before end date", ephemeral=True)
        
        await interaction.response.defer()
        
        # Get stats for both dates
        stats_start = db_get_season_progress(self.season_id, self.account_id, self.selected_start)
        stats_end = db_get_season_progress(self.season_id, self.account_id, self.selected_end)
        
        if not stats_start:
            stats_start = get_cached_stats(self.account_id, self.start_date, self.selected_start)
        if not stats_end:
            stats_end = get_cached_stats(self.account_id, self.start_date, self.selected_end)
        
        if not stats_start or not stats_end:
            return await interaction.followup.send("❌ Missing data for selected dates")
        
        # Helper to parse stats
        def parse_stat(s):
            if not s:
                return 0
            return int(str(s).replace("+", "").replace(",", "") or 0)
        
        # Calculate gains
        power_gain = parse_stat(stats_end.get("power_gain", "0")) - parse_stat(stats_start.get("power_gain", "0"))
        merits_gain = parse_stat(stats_end.get("merits", "0")) - parse_stat(stats_start.get("merits", "0"))
        kills_gain = parse_stat(stats_end.get("kills_gain", "0")) - parse_stat(stats_start.get("kills_gain", "0"))
        deaths_gain = parse_stat(stats_end.get("deads_gain", "0")) - parse_stat(stats_start.get("deads_gain", "0"))
        mana_gain = parse_stat(stats_end.get("mana_gathered", "0")) - parse_stat(stats_start.get("mana_gathered", "0"))
        mana_spent = abs(parse_stat(stats_end.get("mana_spent", "0")))
        gold_spent = abs(parse_stat(stats_end.get("gold_spent", "0")))
        wood_spent = abs(parse_stat(stats_end.get("wood_spent", "0")))
        ore_spent = abs(parse_stat(stats_end.get("ore_spent", "0")))
        
        # Debug logging
        log_info(f"[GAINS DEBUG] Start date {self.selected_start}: power_gain={stats_start.get('power_gain')}")
        log_info(f"[GAINS DEBUG] End date {self.selected_end}: power_gain={stats_end.get('power_gain')}")
        log_info(f"[GAINS DEBUG] Calculated power_gain: {power_gain}")
        
        lord_name = stats_end.get("lord_name", self.account_id)
        
        # Create columnar display
        day_count = (datetime.strptime(self.selected_end, "%Y-%m-%d").date() - 
                     datetime.strptime(self.selected_start, "%Y-%m-%d").date()).days + 1
        
        embed = discord.Embed(
            title=f"📊 Gains Report",
            description=f"**{lord_name}** • {day_count} days ({self.selected_start} → {self.selected_end})",
            color=0x9b59b6
        )
        
        # Row 1: Power, Merits, Kills
        embed.add_field(name="⚔️ Power Gain", value=f"```{power_gain:,}```", inline=True)
        embed.add_field(name="🏆 Merits Gain", value=f"```{merits_gain:,}```", inline=True)
        embed.add_field(name="💀 Kills Gain", value=f"```{kills_gain:,}```", inline=True)
        
        # Row 2: Deaths, Mana Gained, Mana Spent
        embed.add_field(name="☠️ Deaths Gain", value=f"```{deaths_gain:,}```", inline=True)
        embed.add_field(name="💧 Mana Gained", value=f"```{mana_gain:,}```", inline=True)
        embed.add_field(name="💧 Mana Spent", value=f"```{mana_spent:,}```", inline=True)
        
        # Row 3: Resources Spent
        embed.add_field(name="💰 Gold Spent", value=f"```{gold_spent:,}```", inline=True)
        embed.add_field(name="🪵 Wood Spent", value=f"```{wood_spent:,}```", inline=True)
        embed.add_field(name="⛏️ Ore Spent", value=f"```{ore_spent:,}```", inline=True)
        
        await interaction.followup.send(embed=embed)


@bot.command(name="gains")
async def gains(ctx, param1: str = None, param2: str = None):
    """
    Interactive GUI to view gains.
    
    Usage: 
      !gains rekz (current season for rekz)
      !gains rekz sos1 (Season 1 data for rekz)
    """
    
    # Identify user and season from parameters
    season = None
    user_input = None
    
    # Check if param1 is a season name first
    if param1:
        test_season = db_get_season_by_name(param1)
        if test_season:
            season = test_season
            user_input = param2  # If season found, param2 is the user
        else:
            # param1 might be a user, check param2 for season
            if param2:
                test_season = db_get_season_by_name(param2)
                if test_season:
                    season = test_season
                    user_input = param1  # param1 is the user
            else:
                # No season found, just a user specified - use current season
                user_input = param1
                season = db_get_current_season()
    else:
        # No params - use current user and current season
        user_input = None
        season = db_get_current_season()
    
    # Get account ID
    account_id = await get_account_id_from_input(ctx, user_input)
    if not account_id:
        return await ctx.send("❌ Could not find account ID.")
    
    # Get available dates for the season
    try:
        if not season:
            return await ctx.send("❌ No active season found.")
        
        season_id, season_name, start_date, created_at = season
        
        conn = sqlite3.connect(DB_PROGRESS)
        c = conn.cursor()
        
        # Show specific season data only
        c.execute(
            "SELECT DISTINCT data_date FROM season_progress WHERE season_id=? AND account_id=? ORDER BY data_date ASC",
            (season_id, account_id)
        )
        
        dates = [row[0] for row in c.fetchall()]
        conn.close()
        
        if not dates:
            return await ctx.send(f"❌ No saved data found for this account in {season_name}. Run `!loadhistory` first.")
        
        # Show selector
        embed = discord.Embed(
            title="📊 Gains Calculator",
            description=f"Season: **{season_name}**\nSelect start and end dates to see your gains.\n\n**Available dates:** {len(dates)} days of data\n📅 {start_date} → {created_at.split()[0]}",
            color=0x3498db
        )
        embed.add_field(name="ℹ️", value="Both dropdowns must be selected before clicking **Show Gains**", inline=False)
        
        view = GainsDateSelector(dates, account_id, season_id, start_date, ctx, all_seasons=db_get_all_seasons())
        await ctx.send(embed=embed, view=view)
        
    except Exception as e:
        log_error(f"[GAINS] Error: {e}")
        await ctx.send("❌ Error loading data. Try again later.")


@bot.command(name="topmana")
async def topmana(ctx, season_name: str = None):
    """Leaderboard for mana gathered. Usage: !topmana (current) or !topmana sos1 (specific season)"""
    
    # Get season
    if season_name:
        season = resolve_season_input(season_name)
        if not season:
            all_seasons = db_get_all_seasons()
            season_list = ", ".join([s[1] for s in all_seasons]) if all_seasons else "None"
            return await ctx.send(f"❌ Season '{season_name}' not found.\n\nAvailable seasons: {season_list}")
    else:
        season = db_get_current_season()
        if not season:
            return await ctx.send("❌ No season active. Use `/newseason` to start one.")
    
    season_id, season_name_display, start_date, created_at = season
    today = resolve_leaderboard_end_date(season_id)
    
    lords = get_all_lords_from_guild(ctx.guild)
    if not lords:
        return await ctx.send("❌ No members with numeric roles found. Create roles with account IDs as names (e.g., `16322115`).")
    
    await ctx.send(f"⏳ Fetching leaderboard data for {len(lords)} lords...")
    
    # Fetch all lords in parallel with fallback
    fetch_tasks = [
        fetch_stats_with_fallback(lord["account_id"], start_date, today)
        for lord in lords
    ]
    results = await asyncio.gather(*fetch_tasks, return_exceptions=True)
    
    leaderboard = []
    actual_end_date = today
    
    for lord, result in zip(lords, results):
        try:
            # Show all lords, even without data
            lord_name = "Unknown"
            mana_str = "+0"
            mana_num = 0
            
            if result and not isinstance(result, Exception):
                stats, end_date_used = result
                actual_end_date = end_date_used
                
                if stats:
                    lord_name = stats.get("lord_name", lord["name"])
                    
                    # Get mana_gathered if it exists (use abs() to handle negative values)
                    if stats.get("mana_gathered"):
                        mana_str = stats["mana_gathered"]
                        mana_clean = mana_str.replace(",", "").replace("+", "")
                        mana_num = abs(int(mana_clean)) if mana_clean.lstrip("-").isdigit() else 0
                    else:
                        log_info(f"[TOPMANA DEBUG] {lord['account_id']} - mana_gathered is None")
                else:
                    log_info(f"[TOPMANA ERROR] {lord['account_id']} - stats is None")
                    lord_name = lord["name"]
            else:
                log_info(f"[TOPMANA ERROR] {lord['account_id']} - result failed: {result}")
                lord_name = lord["name"]
            
            leaderboard.append({
                "name": lord_name,
                "mana": mana_num,
                "mana_str": mana_str
            })
        except Exception as e:
            log_info(f"[TOPMANA ERROR] {lord['account_id']}: {e}")
            leaderboard.append({
                "name": lord["name"],
                "mana": 0,
                "mana_str": "+0"
            })
    
    # Sort by mana descending
    leaderboard.sort(key=lambda x: x["mana"], reverse=True)
    
    # Build text output - compact
    output = f"```🏆 Top Mana Gathered - {season_name_display}\n"
    medals = ["🥇", "🥈", "🥉"]
    
    for i, lord in enumerate(leaderboard):
        medal = medals[i] if i < 3 else f"{i+1}."
        output += f"{medal} {lord['name']}: {lord['mana_str']}\n"
    
    output += f"📅 {start_date} → {actual_end_date}```"
    await ctx.send(output)


@bot.command(name="topdeaths")
async def topdeaths(ctx, season_name: str = None):
    """Leaderboard for most deaths. Usage: !topdeaths (current) or !topdeaths sos1 (specific season)"""
    
    # Get season
    if season_name:
        season = resolve_season_input(season_name)
        if not season:
            all_seasons = db_get_all_seasons()
            season_list = ", ".join([s[1] for s in all_seasons]) if all_seasons else "None"
            return await ctx.send(f"❌ Season '{season_name}' not found.\n\nAvailable seasons: {season_list}")
    else:
        season = db_get_current_season()
        if not season:
            return await ctx.send("❌ No season active. Use `/newseason` to start one.")
    
    season_id, season_name_display, start_date, created_at = season
    today = resolve_leaderboard_end_date(season_id)
    
    lords = get_all_lords_from_guild(ctx.guild)
    if not lords:
        return await ctx.send("❌ No members with numeric roles found. Create roles with account IDs as names (e.g., `16322115`).")
    
    await ctx.send(f"⏳ Fetching leaderboard data for {len(lords)} lords...")
    
    # Fetch all lords in parallel with fallback
    fetch_tasks = [
        fetch_stats_with_fallback(lord["account_id"], start_date, today)
        for lord in lords
    ]
    results = await asyncio.gather(*fetch_tasks, return_exceptions=True)
    
    leaderboard = []
    actual_end_date = today
    
    for lord, result in zip(lords, results):
        try:
            # Show all lords, even without data
            lord_name = "Unknown"
            deaths_str = "+0"
            deaths_num = 0
            
            if result and not isinstance(result, Exception):
                stats, end_date_used = result
                actual_end_date = end_date_used
                
                if stats:
                    lord_name = stats.get("lord_name", lord["name"])
                    
                    # Get deads_gain if it exists (use abs() to handle negative values)
                    if stats.get("deads_gain"):
                        deaths_str = stats["deads_gain"]
                        deaths_clean = deaths_str.replace(",", "").replace("+", "")
                        deaths_num = abs(int(deaths_clean)) if deaths_clean.lstrip("-").isdigit() else 0
                    else:
                        log_info(f"[TOPDEATHS DEBUG] {lord['account_id']} - deads_gain is None")
                else:
                    log_info(f"[TOPDEATHS ERROR] {lord['account_id']} - stats is None")
                    lord_name = lord["name"]
            else:
                log_info(f"[TOPDEATHS ERROR] {lord['account_id']} - result failed: {result}")
                lord_name = lord["name"]
            
            leaderboard.append({
                "name": lord_name,
                "deaths": deaths_num,
                "deaths_str": deaths_str
            })
        except Exception as e:
            log_info(f"[TOPDEATHS ERROR] {lord['account_id']}: {e}")
            leaderboard.append({
                "name": lord["name"],
                "deaths": 0,
                "deaths_str": "+0"
            })
    
    # Sort by deaths descending
    leaderboard.sort(key=lambda x: x["deaths"], reverse=True)
    
    # Build text output - compact
    output = f"```💀 Most Deaths - {season_name_display}\n"
    medals = ["🥇", "🥈", "🥉"]
    
    for i, lord in enumerate(leaderboard):
        medal = medals[i] if i < 3 else f"{i+1}."
        output += f"{medal} {lord['name']}: {lord['deaths_str']}\n"
    
    output += f"📅 {start_date} → {actual_end_date}```"
    await ctx.send(output)


@bot.command(name="topmerits")
async def topmerits(ctx, season_name: str = None):
    """Leaderboard for highest merits. Usage: !topmerits (current) or !topmerits sos1 (specific season)"""
    
    # Get season
    if season_name:
        season = resolve_season_input(season_name)
        if not season:
            all_seasons = db_get_all_seasons()
            season_list = ", ".join([s[1] for s in all_seasons]) if all_seasons else "None"
            return await ctx.send(f"❌ Season '{season_name}' not found.\n\nAvailable seasons: {season_list}")
    else:
        season = db_get_current_season()
        if not season:
            return await ctx.send("❌ No season active. Use `/newseason` to start one.")
    
    season_id, season_name_display, start_date, created_at = season
    today = resolve_leaderboard_end_date(season_id)
    
    lords = get_all_lords_from_guild(ctx.guild)
    if not lords:
        return await ctx.send("❌ No members with numeric roles found.")
    
    await ctx.send(f"⏳ Fetching leaderboard data for {len(lords)} lords...")
    
    fetch_tasks = [fetch_stats_with_fallback(lord["account_id"], start_date, today) for lord in lords]
    results = await asyncio.gather(*fetch_tasks, return_exceptions=True)
    
    leaderboard = []
    actual_end_date = today
    
    for lord, result in zip(lords, results):
        try:
            lord_name = "Unknown"
            merits_str = "+0"
            merits_num = 0
            
            if result and not isinstance(result, Exception):
                stats, end_date_used = result
                actual_end_date = end_date_used
                if stats:
                    lord_name = stats.get("lord_name", lord["name"])
                    if stats.get("merits"):
                        merits_str = stats["merits"]
                        merits_clean = merits_str.replace(",", "").replace("+", "")
                        merits_num = abs(int(merits_clean)) if merits_clean.lstrip("-").isdigit() else 0
            
            leaderboard.append({"name": lord_name, "merits": merits_num, "merits_str": merits_str})
        except Exception as e:
            log_info(f"[TOPMERITS ERROR] {lord.get('account_id')}: {e}")
            leaderboard.append({"name": lord["name"], "merits": 0, "merits_str": "+0"})
    
    leaderboard.sort(key=lambda x: x["merits"], reverse=True)
    
    output = f"```🏅 Top Merits - {season_name_display}\n"
    medals = ["🥇", "🥈", "🥉"]
    for i, lord in enumerate(leaderboard):
        medal = medals[i] if i < 3 else f"{i+1}."
        output += f"{medal} {lord['name']}: {lord['merits_str']}\n"
    output += f"📅 {start_date} → {actual_end_date}```"
    await ctx.send(output)



async def _top_adv_merit(ctx, season_name, field, emoji, label, tag):
    """Generic advanced merit leaderboard using DB data. COS has historically delayed these
    by 1 day, but that delay may be removed (per an 8/11 site update) — so we try today's
    date first and only fall back to earlier days if today's data isn't there yet."""
    if season_name:
        season = resolve_season_input(season_name)
        if not season:
            all_seasons = db_get_all_seasons()
            season_list = ", ".join([s[1] for s in all_seasons]) if all_seasons else "None"
            return await ctx.send(f"❌ Season '{season_name}' not found.\n\nAvailable seasons: {season_list}")
    else:
        season = db_get_current_season()
        if not season:
            return await ctx.send("❌ No season active. Use `/newseason` to start one.")

    season_id, season_name_display, start_date, created_at = season

    lords = get_all_lords_from_guild(ctx.guild)
    if not lords:
        return await ctx.send("❌ No members with numeric roles found.")

    # For the CURRENT season, base candidates on today (data may still be delayed by COS).
    # For an OLD/ended season, base them on that season's real end_date instead — using
    # today() here would query dates past the season's actual end and pull nothing.
    current_season_check = db_get_current_season()
    is_current_season = current_season_check and current_season_check[0] == season_id
    if is_current_season:
        base_date = date.today()
    else:
        end_date_str = db_get_season_end_date(season_id) or date.today().isoformat()
        base_date = datetime.strptime(end_date_str, "%Y-%m-%d").date()

    adv_candidates = [
        base_date.isoformat(),
        (base_date - timedelta(days=1)).isoformat(),
        (base_date - timedelta(days=2)).isoformat(),
    ]

    def parse_val(raw):
        if not raw:
            return 0
        try:
            return abs(int(str(raw).replace(",", "").replace("+", "").replace("-", "").strip()))
        except:
            return 0

    await ctx.send(f"⏳ Fetching {label} leaderboard...")

    async def fetch_adv_for_lord(account_id):
        """Try each candidate date (today, then earlier) until one has real data for `field`."""
        primary_snap = None
        primary_date = None
        for candidate in adv_candidates:
            snap = db_get_season_progress(season_id, account_id, candidate)
            if not (snap and snap.get(field)):
                try:
                    snap, _ = await fetch_stats_with_fallback(account_id, start_date, candidate)
                except Exception:
                    snap = None
            if snap and snap.get(field):
                primary_snap = snap
                primary_date = candidate
                break

        if not primary_date:
            return None, None, None

        prev_date = (datetime.strptime(primary_date, "%Y-%m-%d") - timedelta(days=1)).strftime("%Y-%m-%d")
        prev_snap = db_get_season_progress(season_id, account_id, prev_date)
        if not (prev_snap and prev_snap.get(field)):
            try:
                prev_snap, _ = await fetch_stats_with_fallback(account_id, start_date, prev_date)
            except Exception:
                prev_snap = None

        return primary_snap, prev_snap, primary_date

    fetch_tasks = [fetch_adv_for_lord(lord["account_id"]) for lord in lords]
    all_results = await asyncio.gather(*fetch_tasks, return_exceptions=True)

    leaderboard = []
    resolved_dates = []
    for lord, result in zip(lords, all_results):
        lord_name = lord["name"]
        val = gain = 0
        try:
            snap, snap_prev, resolved_date = result if not isinstance(result, Exception) else (None, None, None)
            if snap:
                lord_name = snap.get("lord_name", lord["name"]) or lord["name"]
                val = parse_val(snap.get(field))
                resolved_dates.append(resolved_date)
            if snap_prev:
                prev = parse_val(snap_prev.get(field))
                gain = val - prev if val > prev else 0
        except Exception as e:
            log_info(f"[{tag}] Error for {lord['account_id']}: {e}")
        leaderboard.append({"name": lord_name, "val": val, "gain": gain})

    leaderboard.sort(key=lambda x: x["val"], reverse=True)

    # Use the most recent resolved date across all lords as the "data as of" date shown in the header
    data_as_of = max(resolved_dates) if resolved_dates else adv_candidates[0]
    if is_current_season:
        is_up_to_date = data_as_of == date.today().isoformat()
        header_note = "up to date" if is_up_to_date else "COS hasn't released today's yet"
    else:
        header_note = "final"

    medals = ["🥇", "🥈", "🥉"]
    output = f"```{emoji} Top {label} — {season_name_display} (data from {data_as_of}, {header_note})\n"
    for i, lord in enumerate(leaderboard):
        medal = medals[i] if i < 3 else f"{i+1}."
        total_str = f"+{lord['val']:,}"
        gain_str  = f" (+{lord['gain']:,} today)" if lord["gain"] > 0 else ""
        output += f"{medal} {lord['name']}: {total_str}{gain_str}\n"
    output += f"📅 {start_date} → {data_as_of}```"
    await ctx.send(output)


@bot.command(name="topinf")
async def topinf(ctx, season_name: str = None):
    """Leaderboard for Infantry Merits. Usage: !topinf or !topinf sos1"""
    await _top_adv_merit(ctx, season_name, "infantry_merits", "⚔️", "Infantry Merits", "TOPINF")


@bot.command(name="topcav")
async def topcav(ctx, season_name: str = None):
    """Leaderboard for Cavalry Merits. Usage: !topcav or !topcav sos1"""
    await _top_adv_merit(ctx, season_name, "cavalry_merits", "🐴", "Cavalry Merits", "TOPCAV")


@bot.command(name="topmage")
async def topmage(ctx, season_name: str = None):
    """Leaderboard for Mage Merits. Usage: !topmage or !topmage sos1"""
    await _top_adv_merit(ctx, season_name, "mage_merits", "🔮", "Mage Merits", "TOPMAGE")


@bot.command(name="toparcher")
async def toparcher(ctx, season_name: str = None):
    """Leaderboard for Marksman Merits. Usage: !toparcher or !toparcher sos1"""
    await _top_adv_merit(ctx, season_name, "marksman_merits", "🏹", "Marksman Merits", "TOPARCHER")


def _parse_stat_num_global(s):
    if not s:
        return 0
    try:
        return int(str(s).replace("+", "").replace(",", "").replace("-", "").strip())
    except:
        return 0


async def _resolve_t45_healed_for_lord(season_id, account_id, start_date, today):
    """
    Resolve T4/T5 Units Rss Healed for one lord, trying today/yesterday/day-before —
    same delay-handling approach used in !progress. Checks DB first, live-fetches only
    if needed. Returns (lord_name_or_None, int_value).
    """
    candidates = [
        today,
        (datetime.strptime(today, "%Y-%m-%d") - timedelta(days=1)).strftime("%Y-%m-%d"),
        (datetime.strptime(today, "%Y-%m-%d") - timedelta(days=2)).strftime("%Y-%m-%d"),
    ]
    for candidate in candidates:
        db_snap = db_get_season_progress(season_id, account_id, candidate)
        if db_snap and db_snap.get("t45_healed") and db_snap["t45_healed"] not in ("+0", "-0", "0"):
            return db_snap.get("lord_name"), _parse_stat_num_global(db_snap["t45_healed"])

        try:
            live_snap, _ = await fetch_stats_with_fallback(account_id, start_date, candidate)
        except Exception:
            live_snap = None
        if live_snap and live_snap.get("t45_healed") and live_snap["t45_healed"] not in ("+0", "-0", "0"):
            db_save_advanced_stats(season_id, account_id, candidate, live_snap)
            return live_snap.get("lord_name"), _parse_stat_num_global(live_snap["t45_healed"])

    return None, 0


@bot.command(name="topheal")
async def topheal(ctx, season_name: str = None):
    """Leaderboard for T4/T5 Units Rss Healed. Usage: !topheal (current) or !topheal sos1"""

    if season_name:
        season = resolve_season_input(season_name)
        if not season:
            all_seasons = db_get_all_seasons()
            season_list = ", ".join([s[1] for s in all_seasons]) if all_seasons else "None"
            return await ctx.send(f"❌ Season '{season_name}' not found.\n\nAvailable seasons: {season_list}")
    else:
        season = db_get_current_season()
        if not season:
            return await ctx.send("❌ No season active. Use `/newseason` to start one.")

    season_id, season_name_display, start_date, created_at = season
    today = resolve_leaderboard_end_date(season_id)

    lords = get_all_lords_from_guild(ctx.guild)
    if not lords:
        return await ctx.send("❌ No members with numeric roles found.")

    await ctx.send(f"⏳ Fetching heal leaderboard for {len(lords)} lords...")

    fetch_tasks = [
        _resolve_t45_healed_for_lord(season_id, lord["account_id"], start_date, today)
        for lord in lords
    ]
    results = await asyncio.gather(*fetch_tasks, return_exceptions=True)

    leaderboard = []
    for lord, result in zip(lords, results):
        lord_name = lord["name"]
        healed_num = 0
        try:
            if result and not isinstance(result, Exception):
                resolved_name, healed_num = result
                lord_name = resolved_name or lord["name"]
        except Exception as e:
            log_info(f"[TOPHEAL ERROR] {lord['account_id']}: {e}")
        leaderboard.append({"name": lord_name, "val": healed_num})

    leaderboard.sort(key=lambda x: x["val"], reverse=True)

    medals = ["🥇", "🥈", "🥉"]
    output = f"```💊 Top T4/T5 Units Rss Healed — {season_name_display}\n"
    for i, lord in enumerate(leaderboard):
        medal = medals[i] if i < 3 else f"{i+1}."
        output += f"{medal} {lord['name']}: +{lord['val']:,}\n"
    output += f"📅 {start_date} → {today}```"
    await ctx.send(output)


@bot.command(name="topspent")
async def topspent(ctx, season_name: str = None):
    """Leaderboard for estimated mana spent (T4/T5 Healed × 80, assuming it's all T5 — real per-unit rate, but guild data has no T4/T5 split). Usage: !topspent (current) or !topspent sos1"""

    if season_name:
        season = resolve_season_input(season_name)
        if not season:
            all_seasons = db_get_all_seasons()
            season_list = ", ".join([s[1] for s in all_seasons]) if all_seasons else "None"
            return await ctx.send(f"❌ Season '{season_name}' not found.\n\nAvailable seasons: {season_list}")
    else:
        season = db_get_current_season()
        if not season:
            return await ctx.send("❌ No season active. Use `/newseason` to start one.")

    season_id, season_name_display, start_date, created_at = season
    today = resolve_leaderboard_end_date(season_id)

    lords = get_all_lords_from_guild(ctx.guild)
    if not lords:
        return await ctx.send("❌ No members with numeric roles found.")

    await ctx.send(f"⏳ Fetching mana spent leaderboard for {len(lords)} lords...")

    fetch_tasks = [
        _resolve_t45_healed_for_lord(season_id, lord["account_id"], start_date, today)
        for lord in lords
    ]
    results = await asyncio.gather(*fetch_tasks, return_exceptions=True)

    leaderboard = []
    for lord, result in zip(lords, results):
        lord_name = lord["name"]
        healed_num = 0
        try:
            if result and not isinstance(result, Exception):
                resolved_name, healed_num = result
                lord_name = resolved_name or lord["name"]
        except Exception as e:
            log_info(f"[TOPMANASPENT ERROR] {lord['account_id']}: {e}")
        leaderboard.append({"name": lord_name, "val": healed_num * MANA_PER_T5_HEAL})

    leaderboard.sort(key=lambda x: x["val"], reverse=True)

    medals = ["🥇", "🥈", "🥉"]
    output = f"```💧 Top Mana Spent (est.) — {season_name_display}\n(Assuming its all T5)\n\n"
    for i, lord in enumerate(leaderboard):
        medal = medals[i] if i < 3 else f"{i+1}."
        output += f"{medal} {lord['name']}: +{lord['val']:,}\n"
    output += f"\n📅 {start_date} → {today}```"
    await ctx.send(output)


# Per-channel short-term conversation memory for !ask follow-ups.
# In-memory only (not persisted) — resets on bot restart, capped per channel.
_ask_conversations = {}
_ASK_MEMORY_TURNS = 3  # user+assistant pairs kept per channel — kept short so a bad
                       # answer can't compound across many turns before aging out


def _ask_get_history(channel_id):
    return _ask_conversations.get(channel_id, [])


def _ask_save_turn(channel_id, user_msg, assistant_msg):
    history = _ask_conversations.setdefault(channel_id, [])
    history.append({"role": "user", "content": user_msg})
    history.append({"role": "assistant", "content": assistant_msg})
    # Keep only the last N turns (2 entries per turn)
    if len(history) > _ASK_MEMORY_TURNS * 2:
        del history[:len(history) - _ASK_MEMORY_TURNS * 2]


def _ask_clear_history(channel_id):
    _ask_conversations.pop(channel_id, None)


def _ask_resolve_account_id(ctx, player_input):
    """
    Resolve a player name/mention/account_id to an account_id, mirroring the
    resolution logic in !progress. Returns (account_id, error_message_or_None).
    """
    if not player_input:
        for role in ctx.author.roles:
            if role.name.isdigit():
                return role.name, None
        return None, "You don't have a numeric role with your account ID."

    if player_input.isdigit():
        return player_input, None

    username_lower = player_input.lower()
    found_discord_id = None
    for username, discord_id in USERNAME_TO_DISCORD_ID.items():
        if username.lower() == username_lower:
            found_discord_id = discord_id
            break

    if not found_discord_id:
        return None, f"Couldn't find a player named '{player_input}'."

    member = ctx.guild.get_member(found_discord_id)
    if not member:
        return None, f"'{player_input}' is not in this server."

    for role in member.roles:
        if role.name.isdigit():
            return role.name, None

    return None, f"'{player_input}' doesn't have a numeric account ID role."


ASK_TOOLS = [
    {
        "name": "run_leaderboard",
        "description": "Show a guild leaderboard for tracked members. Renders its own message directly.",
        "input_schema": {
            "type": "object",
            "properties": {
                "stat": {
                    "type": "string",
                    "enum": ["topmana", "topinf", "topcav", "topmage", "toparcher", "topheal",
                              "topspent", "topdeaths", "topmerits", "rss"],
                    "description": ("topmana=mana gathered, topinf/topcav/topmage/toparcher=merit "
                                     "breakdowns, topheal=T4/T5 healed, topspent=estimated mana spent "
                                     "(healing x72), topdeaths=most deaths, topmerits=highest merits, "
                                     "rss=top resource spenders")
                },
                "season": {"type": "string", "description": "Season name or ID, omit for current season"}
            },
            "required": ["stat"]
        }
    },
    {
        "name": "run_server_leaderboard",
        "description": "Show a server-wide leaderboard from the last uploaded server Excel data (not limited to guild members).",
        "input_schema": {
            "type": "object",
            "properties": {
                "stat": {
                    "type": "string",
                    "enum": ["stopmerits", "stopdeaths", "stopheal", "stopmana", "stopinf", "stopcav",
                              "stopmage", "stoparcher", "stopother", "stoppower", "stophighest"],
                },
                "top": {"type": "integer", "description": "How many to show, default 25"}
            },
            "required": ["stat"]
        }
    },
    {
        "name": "show_player_progress",
        "description": ("Show the full progress report card for one player, as a visual embed "
                         "posted to the channel. Use this ONLY when the user explicitly wants to "
                         "SEE the card (e.g. 'show me', 'pull up', first mention of a player). "
                         "Do NOT call this again for follow-up questions about a player/season "
                         "already shown earlier in this conversation — use get_player_stats "
                         "instead so the channel doesn't get spammed with a repeated card."),
        "input_schema": {
            "type": "object",
            "properties": {
                "player": {"type": "string", "description": "Player name, or omit for the asker's own progress"},
                "season": {"type": "string", "description": "Season name or ID, omit for current season"}
            }
        }
    },
    {
        "name": "get_player_stats",
        "description": ("Get a player's exact real stats as data, WITHOUT posting the visual card "
                         "to the channel. Use this for follow-up questions, ratings, analysis, or "
                         "'why' after a card has already been shown (or the user just wants an "
                         "answer, not the card) — anywhere you need real numbers without re-spamming "
                         "the channel with another copy of the progress card."),
        "input_schema": {
            "type": "object",
            "properties": {
                "player": {"type": "string", "description": "Player name, or omit for the asker's own progress"},
                "season": {"type": "string", "description": "Season name or ID, omit for current season"}
            }
        }
    },
    {
        "name": "compare_players",
        "description": "Compare two players side by side. Renders its own embed directly.",
        "input_schema": {
            "type": "object",
            "properties": {
                "player1": {"type": "string"},
                "player2": {"type": "string"}
            },
            "required": ["player1", "player2"]
        }
    },
    {
        "name": "show_season_history",
        "description": "Show all seasons with their dates. Renders its own message directly. Use when the user explicitly wants to SEE the season list.",
        "input_schema": {"type": "object", "properties": {}}
    },
    {
        "name": "get_season_list",
        "description": ("Silently look up all seasons with their ID, name, start date, and end date "
                         "(or 'ongoing'), without posting anything to the channel. Use this to resolve "
                         "relative references like 'last season' or 'two seasons ago' before calling "
                         "another tool that needs a specific season name/ID — not for when the user "
                         "wants to actually see the season list (use show_season_history for that)."),
        "input_schema": {"type": "object", "properties": {}}
    },
    {
        "name": "get_upcoming_events",
        "description": "Get the list of upcoming scheduled events with their dates. Returns data for you to summarize in your own words.",
        "input_schema": {"type": "object", "properties": {}}
    },
    {
        "name": "get_player_growth",
        "description": ("Get a player's stat growth over a recent period (e.g. 'how much did rekz "
                         "grow this week'). Returns raw numbers for you to summarize conversationally."),
        "input_schema": {
            "type": "object",
            "properties": {
                "player": {"type": "string", "description": "Player name, or omit for the asker"},
                "days_back": {"type": "integer", "description": "How many days back to compare against, default 7"}
            }
        }
    },
    {
        "name": "add_event",
        "description": "Add a single scheduled event. Admin only.",
        "input_schema": {
            "type": "object",
            "properties": {
                "name": {"type": "string"},
                "date": {"type": "string", "description": "YYYY-MM-DD"},
                "time_utc": {"type": "string", "description": "HH:MM in 24h UTC, default 12:00 if not given"},
                "reminder_minutes": {"type": "integer", "description": "Minutes before the event to remind, default 0 (no reminder)"},
                "scheduled_time": {"type": "string", "description": "ISO 8601 UTC datetime to run this action later instead of immediately. Omit to run immediately."}
            },
            "required": ["name", "date"]
        }
    },
    {
        "name": "edit_event",
        "description": "Edit one existing event by matching its name (partial match ok). Change its name, date, time, and/or reminder. Admin only.",
        "input_schema": {
            "type": "object",
            "properties": {
                "match": {"type": "string", "description": "Text to match against existing event names"},
                "new_name": {"type": "string"},
                "new_date": {"type": "string", "description": "YYYY-MM-DD"},
                "new_time_utc": {"type": "string", "description": "HH:MM in 24h UTC"},
                "new_reminder_minutes": {"type": "integer"},
                "scheduled_time": {"type": "string", "description": "ISO 8601 UTC datetime to run this action later instead of immediately. Omit to run immediately."}
            },
            "required": ["match"]
        }
    },
    {
        "name": "bulk_edit_event_times",
        "description": ("Change the time-of-day for MULTIPLE events at once, keeping their dates and "
                         "names the same. Useful for 'change all the events I just added from 12 UTC "
                         "to 13 UTC'. Filters by the events' CURRENT time and optionally by a month or "
                         "name substring. Admin only. ALWAYS report exactly how many events were changed."),
        "input_schema": {
            "type": "object",
            "properties": {
                "current_time_utc": {"type": "string", "description": "Only change events currently at this HH:MM UTC, e.g. '12:00'"},
                "new_time_utc": {"type": "string", "description": "New HH:MM UTC to set"},
                "month": {"type": "string", "description": "Optional: only affect events in this month, e.g. 'August'"},
                "name_contains": {"type": "string", "description": "Optional: only affect events whose name contains this text"},
                "scheduled_time": {"type": "string", "description": "ISO 8601 UTC datetime to run this action later instead of immediately. Omit to run immediately."}
            },
            "required": ["new_time_utc"]
        }
    },
    {
        "name": "delete_event",
        "description": "Delete one event by matching its name (partial match ok). Admin only — always confirm what was deleted.",
        "input_schema": {
            "type": "object",
            "properties": {
                "match": {"type": "string", "description": "Text to match against existing event names"},
                "scheduled_time": {"type": "string", "description": "ISO 8601 UTC datetime to run this action later instead of immediately. Omit to run immediately."}
            },
            "required": ["match"]
        }
    },
    {
        "name": "run_server_rankings",
        "description": "Show the top N servers by highest power, Gen-2 Pool. Renders its own message directly.",
        "input_schema": {
            "type": "object",
            "properties": {
                "n": {"type": "integer", "description": "How many servers to show, 1-100, default 10"}
            }
        }
    },
    {
        "name": "check_server_rank",
        "description": "Look up a specific server's rank in the server rankings. Renders its own message directly.",
        "input_schema": {
            "type": "object",
            "properties": {
                "server_number": {"type": "integer", "description": "The server number, e.g. 698"}
            },
            "required": ["server_number"]
        }
    },
    {
        "name": "show_saved_kvk_matchups",
        "description": "List all saved KvK matchups. Renders its own message directly.",
        "input_schema": {"type": "object", "properties": {}}
    },
    {
        "name": "show_active_members",
        "description": "Show which guild members are active vs inactive based on recent stat changes. Renders its own message directly.",
        "input_schema": {"type": "object", "properties": {}}
    },
    {
        "name": "show_data_history",
        "description": "Show the oldest and newest data dates saved in the database. Renders its own message directly.",
        "input_schema": {"type": "object", "properties": {}}
    },
    {
        "name": "force_fetch_all",
        "description": "Force-refresh all tracked members' stats right now instead of waiting for the daily update. Admin only. Renders its own message directly.",
        "input_schema": {
            "type": "object",
            "properties": {
                "scheduled_time": {"type": "string", "description": "ISO 8601 UTC datetime to run this action later instead of immediately. Omit to run immediately."}
            }
        }
    },
    {
        "name": "view_kvk_matchup",
        "description": "View full details of one specific saved KvK matchup by its ID. Renders its own message directly.",
        "input_schema": {
            "type": "object",
            "properties": {"matchup_id": {"type": "integer"}},
            "required": ["matchup_id"]
        }
    },
    {
        "name": "delete_kvk_matchup",
        "description": "Delete one saved KvK matchup by its ID. Owner only.",
        "input_schema": {
            "type": "object",
            "properties": {"matchup_id": {"type": "integer"}},
            "required": ["matchup_id"]
        }
    },
    {
        "name": "show_weekly_events",
        "description": "Show the rotating weekly Abyss events (Melee/Range Wheel/Forge schedule). Renders its own message directly.",
        "input_schema": {"type": "object", "properties": {}}
    },
    {
        "name": "edit_season",
        "description": "Edit a season's name, start date, and/or end date. Match the season by name or ID. Admin only.",
        "input_schema": {
            "type": "object",
            "properties": {
                "season": {"type": "string", "description": "Season name or ID to edit"},
                "new_name": {"type": "string"},
                "new_start_date": {"type": "string", "description": "YYYY-MM-DD"},
                "new_end_date": {"type": "string", "description": "YYYY-MM-DD, or 'ongoing' to clear the end date"},
                "scheduled_time": {"type": "string", "description": "ISO 8601 UTC datetime to run this action later instead of immediately. Omit to run immediately."}
            },
            "required": ["season"]
        }
    },
    {
        "name": "get_activity_rankings",
        "description": ("Get a composite performance snapshot for EVERY tracked guild member at once "
                         "(merits, power gain, merit-to-power ratio, kills, deaths, resources gathered). "
                         "Use this for open-ended questions like 'who is slacking', 'who is the best "
                         "player', 'who's carrying the guild', 'rank everyone' — anything that needs "
                         "comparing members across multiple stats rather than one single leaderboard. "
                         "Returns raw per-member data for you to analyze and explain, not a rendered "
                         "leaderboard."),
        "input_schema": {"type": "object", "properties": {}}
    },
    {
        "name": "show_chart",
        "description": ("Show a line chart image of how a stat grew over a season for one player, "
                         "using the daily archived history. Renders its own image directly in the "
                         "channel — use when someone wants to SEE growth over time, not just a number."),
        "input_schema": {
            "type": "object",
            "properties": {
                "player": {"type": "string", "description": "Player name, or omit for the asker"},
                "stat": {
                    "type": "string",
                    "enum": ["merits", "power", "highest_power", "kills", "deaths", "healed",
                              "gold", "wood", "ore", "mana", "infantry", "cavalry", "mage",
                              "marksman", "other", "t45_healed", "t45_dead", "coins"]
                },
                "season": {"type": "string", "description": "Season name or ID, omit for current season"}
            },
            "required": ["stat"]
        }
    },
    {
        "name": "show_group_chart",
        "description": ("Show ONE chart with EVERY tracked guild member's growth for a stat, all on "
                         "the same graph with a legend. Use this for 'chart everyone', 'chart the "
                         "whole group/guild', or any request comparing all members' growth at once — "
                         "do NOT call show_chart in a loop per player for this, use this single tool "
                         "instead, since it resolves members directly from Discord roles rather than "
                         "matching names (which can fail on in-game clan-tag prefixes)."),
        "input_schema": {
            "type": "object",
            "properties": {
                "stat": {
                    "type": "string",
                    "enum": ["merits", "power", "highest_power", "kills", "deaths", "healed",
                              "gold", "wood", "ore", "mana", "infantry", "cavalry", "mage",
                              "marksman", "other", "t45_healed", "t45_dead", "coins"]
                },
                "season": {"type": "string", "description": "Season name or ID, omit for current season"}
            },
            "required": ["stat"]
        }
    },
    {
        "name": "calculate",
        "description": ("Evaluate an exact math expression — arithmetic, powers, roots, trig, logs, "
                         "etc. Use this whenever a question needs a precise numeric result (not just "
                         "an estimate), including as a step while working through a larger math "
                         "problem. Supports +, -, *, /, //, %, **, parentheses, and functions: sqrt, "
                         "sin, cos, tan, asin, acos, atan, log, log10, log2, exp, factorial, abs, "
                         "round, floor, ceil, and the constants pi and e."),
        "input_schema": {
            "type": "object",
            "properties": {
                "expression": {"type": "string", "description": "e.g. 'sqrt(144) + 7 ** 2' or '(15/4) * log(100, 10)'"}
            },
            "required": ["expression"]
        }
    },
    {
        "name": "get_pace_projection",
        "description": ("Project how long it will take a player to reach a target value for a stat, "
                         "based on their pace so far this season (e.g. 'at my current pace, how long "
                         "until I hit 1B mana gathered'). Returns raw numbers (current value, daily "
                         "rate, days elapsed, days needed, projected date) for you to explain."),
        "input_schema": {
            "type": "object",
            "properties": {
                "player": {"type": "string", "description": "Player name, or omit for the asker"},
                "metric": {
                    "type": "string",
                    "enum": ["merits", "power_gain", "kills_gain", "deads_gain", "healed_gain",
                              "gold_gathered", "wood_gathered", "ore_gathered", "mana_gathered"],
                },
                "target": {"type": "number", "description": "The target value to reach, e.g. 1000000000 for 1B"}
            },
            "required": ["metric", "target"]
        }
    },
    {
        "name": "create_season",
        "description": ("Start a new season. This ends the currently active season automatically "
                         "(its end_date is set to the day before this new season starts). Admin only. "
                         "Can be scheduled for later with scheduled_time."),
        "input_schema": {
            "type": "object",
            "properties": {
                "name": {"type": "string", "description": "Season name, e.g. 'Nvr vs Yss'"},
                "start_date": {"type": "string", "description": "YYYY-MM-DD the season starts"},
                "scheduled_time": {"type": "string", "description": "ISO 8601 UTC datetime to run this action later instead of immediately, e.g. '2026-08-28T00:00:00'. Omit to run immediately."}
            },
            "required": ["name", "start_date"]
        }
    },
    {
        "name": "list_queued_tasks",
        "description": "List all pending scheduled/queued tasks (things set to run later). Renders its own message directly.",
        "input_schema": {"type": "object", "properties": {}}
    },
    {
        "name": "cancel_queued_task",
        "description": "Cancel a pending queued task by its ID. Admin only.",
        "input_schema": {
            "type": "object",
            "properties": {"task_id": {"type": "integer"}},
            "required": ["task_id"]
        }
    },
    {
        "name": "edit_queued_task_time",
        "description": "Change the scheduled time of a pending queued task. Admin only.",
        "input_schema": {
            "type": "object",
            "properties": {
                "task_id": {"type": "integer"},
                "new_scheduled_time": {"type": "string", "description": "ISO 8601 UTC datetime, e.g. '2026-08-28T00:00:00'"}
            },
            "required": ["task_id", "new_scheduled_time"]
        }
    },
]


def _ask_safe_calculate(expression):
    """
    Safely evaluate a math expression using Python's ast module — whitelists only
    numeric literals, arithmetic operators, and a fixed set of math functions.
    Never uses eval() directly on untrusted input. Returns (result, error).
    """
    import ast as _ast_calc
    import math as _math_calc

    allowed_names = {
        "sqrt": _math_calc.sqrt, "sin": _math_calc.sin, "cos": _math_calc.cos, "tan": _math_calc.tan,
        "asin": _math_calc.asin, "acos": _math_calc.acos, "atan": _math_calc.atan,
        "log": _math_calc.log, "log10": _math_calc.log10, "log2": _math_calc.log2,
        "exp": _math_calc.exp, "factorial": _math_calc.factorial, "abs": abs,
        "round": round, "floor": _math_calc.floor, "ceil": _math_calc.ceil,
        "pi": _math_calc.pi, "e": _math_calc.e,
    }
    allowed_binops = (_ast_calc.Add, _ast_calc.Sub, _ast_calc.Mult, _ast_calc.Div,
                       _ast_calc.FloorDiv, _ast_calc.Mod, _ast_calc.Pow, _ast_calc.USub, _ast_calc.UAdd)

    def _eval_node(node):
        if isinstance(node, _ast_calc.Constant):
            if isinstance(node.value, (int, float)):
                return node.value
            raise ValueError("Only numbers are allowed")
        if isinstance(node, _ast_calc.BinOp):
            if not isinstance(node.op, allowed_binops):
                raise ValueError("Operator not allowed")
            left = _eval_node(node.left)
            right = _eval_node(node.right)
            if isinstance(node.op, _ast_calc.Add): return left + right
            if isinstance(node.op, _ast_calc.Sub): return left - right
            if isinstance(node.op, _ast_calc.Mult): return left * right
            if isinstance(node.op, _ast_calc.Div): return left / right
            if isinstance(node.op, _ast_calc.FloorDiv): return left // right
            if isinstance(node.op, _ast_calc.Mod): return left % right
            if isinstance(node.op, _ast_calc.Pow): return left ** right
        if isinstance(node, _ast_calc.UnaryOp):
            if not isinstance(node.op, allowed_binops):
                raise ValueError("Operator not allowed")
            operand = _eval_node(node.operand)
            return -operand if isinstance(node.op, _ast_calc.USub) else +operand
        if isinstance(node, _ast_calc.Name):
            if node.id in allowed_names and not callable(allowed_names[node.id]):
                return allowed_names[node.id]
            raise ValueError(f"Unknown name: {node.id}")
        if isinstance(node, _ast_calc.Call):
            if not isinstance(node.func, _ast_calc.Name) or node.func.id not in allowed_names:
                raise ValueError("Function not allowed")
            fn = allowed_names[node.func.id]
            if not callable(fn):
                raise ValueError("Not a function")
            args = [_eval_node(a) for a in node.args]
            return fn(*args)
        raise ValueError(f"Expression not allowed: {type(node).__name__}")

    try:
        tree = _ast_calc.parse(expression, mode="eval")
        result = _eval_node(tree.body)
        return result, None
    except Exception as e:
        return None, str(e)


def _ask_is_admin(ctx):
    return ctx.author.id == OWNER_ID or (ctx.guild and ctx.author.guild_permissions.administrator)


async def _ask_fetch_player_stats_summary(ctx, player_input, season_input):
    """
    Fetch a player's exact real stats as a text summary for Claude to reason with —
    shared by show_player_progress (which also renders the visual card) and
    get_player_stats (data only, no render, used for follow-up questions so the
    channel doesn't get spammed with a repeated card).
    Returns the summary string, or None if it couldn't be fetched.
    """
    try:
        account_id, err = _ask_resolve_account_id(ctx, player_input)
        if err:
            return None
        season = resolve_season_input(season_input) if season_input else db_get_current_season()
        if not season:
            return None
        season_id, season_name, start_date, _ = season

        # For the CURRENT active season, fetch up to today. For an OLD/ended season, cap at
        # that season's actual end_date — using today() would push the range past the
        # season's real end and pull real-but-wrong-range data from COS.
        current_season = db_get_current_season()
        is_current = current_season and current_season[0] == season_id
        end_ref = date.today().isoformat() if is_current else (db_get_season_end_date(season_id) or date.today().isoformat())

        # Check the database FIRST, same resolution order as !progress — a live fetch's
        # own internal fallback can occasionally shift the start date to work around a
        # temporarily-empty date, which silently drops days of real progress from fields
        # like power/healed/mana. Archived data doesn't have that risk, so prefer it
        # whenever we have it and only live-fetch as a last resort.
        stats = db_get_season_progress(season_id, account_id, end_ref)
        actual_end = end_ref
        if not stats:
            yesterday = (date.today() - timedelta(days=1)).isoformat()
            stats = db_get_season_progress(season_id, account_id, yesterday)
            if stats:
                actual_end = yesterday
        if not stats:
            stats = get_cached_stats(account_id, start_date, end_ref)
            if stats:
                actual_end = stats.get("data_date", end_ref)
        if not stats:
            stats, actual_end = await fetch_stats_with_fallback(account_id, start_date, end_ref)
        if not stats:
            return None
        return (
            f"EXACT VERIFIED DATA (copy these numbers precisely, do not alter or round them) "
            f"for {stats.get('lord_name', account_id)}, season {season_name} "
            f"({start_date} to {actual_end}): merits={_parse_stat_num_global(stats.get('merits')):,}, "
            f"power_gain={_parse_stat_num_global(stats.get('power_gain')):,}, "
            f"kills={_parse_stat_num_global(stats.get('kills_gain')):,}, "
            f"deaths={_parse_stat_num_global(stats.get('deads_gain')):,}, "
            f"healed={_parse_stat_num_global(stats.get('healed_gain')):,}, "
            f"gathered(gold/wood/ore/mana)={_parse_stat_num_global(stats.get('gold_gathered')):,}/"
            f"{_parse_stat_num_global(stats.get('wood_gathered')):,}/{_parse_stat_num_global(stats.get('ore_gathered')):,}/"
            f"{_parse_stat_num_global(stats.get('mana_gathered')):,}. "
            f"If the user wants a rating, roast, analysis, or advice, use ONLY these exact numbers — "
            f"never substitute a different or rounded figure."
        )
    except Exception as e:
        log_info(f"[ASK player stats fetch] {e}")
        return None


SCHEDULABLE_TOOLS = {"add_event", "edit_event", "bulk_edit_event_times", "delete_event",
                     "edit_season", "create_season", "force_fetch_all"}


async def _ask_execute_tool(ctx, tool_name, tool_input, bypass_permission=False):
    """
    Execute a tool call. Returns (rendered_directly: bool, result_for_claude: str|None).
    rendered_directly=True means the tool already sent its own Discord message/embed
    and Claude doesn't need to see a result. rendered_directly=False means the result
    text should be handed back to Claude to compose a natural-language reply.
    bypass_permission=True is used when re-running a queued task at its scheduled time —
    permission was already verified when it was queued, so it isn't re-checked, and it
    won't be re-queued even if scheduled_time is still present in the stored input.
    """
    try:
        # If this call includes a future scheduled_time, queue it instead of running now.
        if not bypass_permission and tool_name in SCHEDULABLE_TOOLS and tool_input.get("scheduled_time"):
            try:
                scheduled_dt = datetime.fromisoformat(tool_input["scheduled_time"])
            except Exception:
                return False, f"Couldn't parse scheduled_time '{tool_input['scheduled_time']}'. Use ISO format like 2026-08-28T00:00:00."
            if scheduled_dt > datetime.utcnow():
                if not _ask_is_admin(ctx):
                    return False, "Only admins can schedule actions."
                clean_input = {k: v for k, v in tool_input.items() if k != "scheduled_time"}
                description = f"{tool_name}(" + ", ".join(f"{k}={v}" for k, v in clean_input.items()) + ")"
                task_id = db_queue_task(tool_name, clean_input, description, scheduled_dt.isoformat(), ctx.author.id, ctx.channel.id)
                return False, f"Queued task #{task_id}: {description} — scheduled for {scheduled_dt.strftime('%Y-%m-%d %H:%M UTC')}."
            # scheduled_time is in the past/now — fall through and run immediately

        if tool_name == "run_leaderboard":
            action_map = {
                "topmana": topmana, "topinf": topinf, "topcav": topcav, "topmage": topmage,
                "toparcher": toparcher, "topheal": topheal, "topspent": topspent,
                "topdeaths": topdeaths, "topmerits": topmerits, "rss": rss_leaderboard,
            }
            cmd = action_map.get(tool_input.get("stat"))
            if not cmd:
                return False, f"Unknown leaderboard: {tool_input.get('stat')}"
            await cmd.callback(ctx, tool_input.get("season"))
            return True, None

        if tool_name == "run_server_leaderboard":
            action_map = {
                "stopmerits": stopmerits, "stopdeaths": stopdeaths, "stopheal": stopheal,
                "stopmana": stopmana, "stopinf": stopinf, "stopcav": stopcav, "stopmage": stopmage,
                "stoparcher": stoparcher, "stopother": stopother, "stoppower": stoppower,
                "stophighest": stophighest,
            }
            cmd = action_map.get(tool_input.get("stat"))
            if not cmd:
                return False, f"Unknown server leaderboard: {tool_input.get('stat')}"
            top = tool_input.get("top", 25)
            if tool_input.get("stat") in ("stopheal", "stophighest", "stopmana"):
                await cmd.callback(ctx, top)
            else:
                await cmd.callback(ctx, None, top)
            return True, None

        if tool_name == "show_player_progress":
            await progress.callback(ctx, tool_input.get("player"), tool_input.get("season"))
            summary = await _ask_fetch_player_stats_summary(ctx, tool_input.get("player"), tool_input.get("season"))
            return True, summary

        if tool_name == "get_player_stats":
            # Same real numbers as show_player_progress, but does NOT render the visual
            # card — use this for follow-ups/analysis so the channel doesn't get spammed
            # with a repeated card every time a question needs the underlying data.
            summary = await _ask_fetch_player_stats_summary(ctx, tool_input.get("player"), tool_input.get("season"))
            if not summary:
                return False, "Couldn't fetch stats for that player/season."
            return False, summary

        if tool_name == "compare_players":
            await compare.callback(ctx, tool_input.get("player1"), tool_input.get("player2"))
            return True, None

        if tool_name == "show_season_history":
            await seasonhistory.callback(ctx)
            return True, None

        if tool_name == "get_season_list":
            all_seasons = db_get_all_seasons()
            if not all_seasons:
                return False, "No seasons found."
            current = db_get_current_season()
            current_id = current[0] if current else None
            lines = []
            for s in all_seasons:
                season_id, season_name, start_date, created_at = s
                end_date = db_get_season_end_date(season_id)
                status = "ACTIVE" if season_id == current_id else "ENDED"
                lines.append(f"#{season_id} '{season_name}' [{status}] {start_date} → {end_date or 'ongoing'}")
            return False, "Seasons (ordered by creation):\n" + "\n".join(lines)

        if tool_name == "get_upcoming_events":
            events = db_get_events()
            if not events:
                return False, "No upcoming events scheduled."
            now = datetime.utcnow()
            upcoming = []
            for eid, name, dt_str, reminder in events:
                try:
                    dt = datetime.fromisoformat(dt_str)
                    if dt > now:
                        upcoming.append(f"{name}: {dt.strftime('%Y-%m-%d %H:%M UTC')}")
                except Exception:
                    continue
            if not upcoming:
                return False, "No upcoming events scheduled."
            return False, "Upcoming events:\n" + "\n".join(upcoming)

        if tool_name == "get_player_growth":
            account_id, err = _ask_resolve_account_id(ctx, tool_input.get("player"))
            if err:
                return False, f"Error: {err}"
            days_back = tool_input.get("days_back", 7)
            season = db_get_current_season()
            if not season:
                return False, "No active season."
            season_id, season_name, start_date, _ = season
            today_str = date.today().isoformat()
            past_str = (date.today() - timedelta(days=days_back)).isoformat()

            today_stats, _ = await fetch_stats_with_fallback(account_id, start_date, today_str)
            past_stats, _ = await fetch_stats_with_fallback(account_id, start_date, past_str)

            if not today_stats or not past_stats:
                return False, "Couldn't fetch enough data for that comparison."

            def _num(s):
                if not s:
                    return 0
                try:
                    return int(str(s).replace("+", "").replace(",", "").strip())
                except:
                    return 0

            result_lines = [f"Player: {today_stats.get('lord_name', account_id)}, season: {season_name}, comparing {past_str} to {today_str} ({days_back} days):"]
            for field, label in [("power_gain", "power"), ("merits", "merits"), ("kills_gain", "kills"),
                                   ("deads_gain", "deaths"), ("healed_gain", "healed")]:
                delta = _num(today_stats.get(field)) - _num(past_stats.get(field))
                result_lines.append(f"{label} grew by {delta:,} in this period")
            return False, "\n".join(result_lines)

        if tool_name == "add_event":
            if not bypass_permission and not _ask_is_admin(ctx):
                return False, "Only admins can add events."
            try:
                date_str = tool_input["date"]
                time_str = tool_input.get("time_utc", "12:00")
                reminder = tool_input.get("reminder_minutes", 0)
                dt = datetime.strptime(f"{date_str} {time_str}", "%Y-%m-%d %H:%M")
            except Exception as e:
                return False, f"Couldn't parse that date/time: {e}"
            db_add_event(tool_input["name"], dt.isoformat(), reminder)
            return False, f"Added event '{tool_input['name']}' on {dt.strftime('%Y-%m-%d %H:%M UTC')}."

        if tool_name == "edit_event":
            if not bypass_permission and not _ask_is_admin(ctx):
                return False, "Only admins can edit events."
            events = db_get_events()
            match_text = tool_input["match"].lower()
            matches = [e for e in events if match_text in e[1].lower()]
            if not matches:
                return False, f"No event found matching '{tool_input['match']}'."
            if len(matches) > 1:
                names = ", ".join(f"'{m[1]}'" for m in matches[:5])
                return False, f"Multiple events match '{tool_input['match']}': {names}. Be more specific."

            event_id, old_name, old_dt_str, old_reminder = matches[0]
            try:
                old_dt = datetime.fromisoformat(old_dt_str)
            except Exception:
                old_dt = datetime.utcnow()

            new_name = tool_input.get("new_name")
            new_date = tool_input.get("new_date")
            new_time = tool_input.get("new_time_utc")
            new_reminder = tool_input.get("new_reminder_minutes")

            new_dt = None
            if new_date or new_time:
                date_part = new_date or old_dt.strftime("%Y-%m-%d")
                time_part = new_time or old_dt.strftime("%H:%M")
                try:
                    new_dt = datetime.strptime(f"{date_part} {time_part}", "%Y-%m-%d %H:%M")
                except Exception as e:
                    return False, f"Couldn't parse that date/time: {e}"

            db_update_event(
                event_id,
                name=new_name,
                dt=new_dt.isoformat() if new_dt else None,
                reminder=new_reminder,
            )
            final_name = new_name or old_name
            final_dt = new_dt or old_dt
            return False, f"Updated event: '{final_name}' now at {final_dt.strftime('%Y-%m-%d %H:%M UTC')}."

        if tool_name == "bulk_edit_event_times":
            if not bypass_permission and not _ask_is_admin(ctx):
                return False, "Only admins can bulk-edit events."
            events = db_get_events()
            current_time = tool_input.get("current_time_utc")
            new_time = tool_input["new_time_utc"]
            month_filter = tool_input.get("month", "").lower()
            name_filter = tool_input.get("name_contains", "").lower()

            changed = []
            for event_id, name, dt_str, reminder in events:
                try:
                    dt = datetime.fromisoformat(dt_str)
                except Exception:
                    continue
                if current_time and dt.strftime("%H:%M") != current_time:
                    continue
                if month_filter and month_filter not in dt.strftime("%B").lower():
                    continue
                if name_filter and name_filter not in name.lower():
                    continue
                try:
                    new_hour, new_min = map(int, new_time.split(":"))
                    new_dt = dt.replace(hour=new_hour, minute=new_min)
                except Exception:
                    continue
                db_update_event(event_id, dt=new_dt.isoformat())
                changed.append(f"{name} ({dt.strftime('%Y-%m-%d')})")

            if not changed:
                return False, "No events matched those filters — nothing was changed."
            return False, f"Changed the time to {new_time} UTC for {len(changed)} events: " + ", ".join(changed)

        if tool_name == "delete_event":
            if not bypass_permission and not _ask_is_admin(ctx):
                return False, "Only admins can delete events."
            events = db_get_events()
            match_text = tool_input["match"].lower()
            matches = [e for e in events if match_text in e[1].lower()]
            if not matches:
                return False, f"No event found matching '{tool_input['match']}'."
            if len(matches) > 1:
                names = ", ".join(f"'{m[1]}'" for m in matches[:5])
                return False, f"Multiple events match '{tool_input['match']}': {names}. Be more specific."
            event_id, name, dt_str, reminder = matches[0]
            db_delete_event(event_id)
            return False, f"Deleted event '{name}'."

        if tool_name == "run_server_rankings":
            n = tool_input.get("n", 10)
            n = max(1, min(100, n))
            await cmd_servertop(ctx, n)
            return True, None

        if tool_name == "check_server_rank":
            await cmd_servercheck(ctx, str(tool_input["server_number"]))
            return True, None

        if tool_name == "show_saved_kvk_matchups":
            await cmd_matchups_list(ctx)
            return True, None

        if tool_name == "show_active_members":
            await active_members.callback(ctx)
            return True, None

        if tool_name == "show_data_history":
            await datahistory.callback(ctx)
            return True, None

        if tool_name == "force_fetch_all":
            if not bypass_permission and not _ask_is_admin(ctx):
                return False, "Only admins can force a data refresh."
            await forcefetch.callback(ctx)
            return True, None

        if tool_name == "view_kvk_matchup":
            await cmd_matchup_view(ctx, tool_input["matchup_id"])
            return True, None

        if tool_name == "delete_kvk_matchup":
            if not bypass_permission and ctx.author.id != OWNER_ID:
                return False, "Only the owner can delete matchups."
            await cmd_matchup_delete(ctx, tool_input["matchup_id"])
            return True, None

        if tool_name == "show_weekly_events":
            today = date.today()
            sunday = start_date - timedelta(days=start_date.weekday() + 1)
            weeks = (today - sunday).days // 7
            this_tue = sunday + timedelta(weeks=weeks, days=2)
            now = datetime.utcnow()
            event_start = datetime.combine(this_tue, time(0, 0))
            event_end = event_start + timedelta(days=3)
            if now >= event_end:
                weeks += 1

            msg = "📅 **Weekly Abyss Events**\n\n"
            nums = ["1️⃣", "2️⃣", "3️⃣", "4️⃣"]
            for i in range(4):
                idx = (weeks + i) % 4
                nm = weekly_events[idx]
                emoji = event_emojis.get(nm, "📌")
                ev_date = sunday + timedelta(weeks=weeks + i, days=2)
                base_dt = datetime.combine(ev_date, time(0, 0))
                msg += f"{nums[i]} {emoji} **{nm}** — <t:{int(base_dt.timestamp())}:F>\n\n"
            await ctx.send(msg)
            return True, None

        if tool_name == "edit_season":
            if not bypass_permission and not _ask_is_admin(ctx):
                return False, "Only admins can edit seasons."
            season = resolve_season_input(tool_input["season"])
            if not season:
                return False, f"No season found matching '{tool_input['season']}'."
            season_id, old_name, old_start, _ = season

            new_name = tool_input.get("new_name")
            new_start = tool_input.get("new_start_date")
            new_end_raw = tool_input.get("new_end_date")
            new_end = None
            if new_end_raw:
                new_end = "" if new_end_raw.lower() == "ongoing" else new_end_raw

            if not any([new_name, new_start, new_end_raw]):
                return False, "Nothing to change was specified."

            db_update_season(season_id, season_name=new_name, start_date=new_start, end_date=new_end)
            changes = []
            if new_name:
                changes.append(f"name to '{new_name}'")
            if new_start:
                changes.append(f"start date to {new_start}")
            if new_end_raw:
                changes.append("end date cleared (ongoing)" if new_end == "" else f"end date to {new_end}")
            return False, f"Updated season '{old_name}' (#{season_id}): {', '.join(changes)}."

        if tool_name == "get_activity_rankings":
            season = db_get_current_season()
            if not season:
                return False, "No active season."
            season_id, season_name, start_date, _ = season

            lords = get_all_lords_from_guild(ctx.guild)
            if not lords:
                return False, "No tracked members found."

            def _num(s):
                if not s:
                    return 0
                try:
                    return int(str(s).replace("+", "").replace(",", "").strip())
                except:
                    return 0

            rows = []
            for lord in lords:
                snap = db_get_latest_season_progress(season_id, lord["account_id"])
                if not snap:
                    continue
                merits = _num(snap.get("merits"))
                power_gain = _num(snap.get("power_gain"))
                total_power = _num(snap.get("highest_power"))
                kills = _num(snap.get("kills_gain"))
                deaths = _num(snap.get("deads_gain"))
                gathered = sum(_num(snap.get(k)) for k in
                               ["gold_gathered", "wood_gathered", "ore_gathered", "mana_gathered"])
                # Merit Ratio = merits / TOTAL power, not merits / power gained this season.
                # Fall back to power_gain only if we don't have a cached highest_power for
                # this member yet (e.g. they've never had !progress run for them).
                ratio_base = total_power if total_power > 0 else power_gain
                ratio = round((merits / ratio_base * 100), 1) if ratio_base > 0 else None
                rows.append({
                    "name": snap.get("lord_name") or lord["name"],
                    "merits": merits, "power_gain": power_gain, "total_power": total_power,
                    "merit_ratio_pct": ratio, "kills": kills, "deaths": deaths,
                    "resources_gathered": gathered,
                })

            if not rows:
                return False, "No members have data for this season yet."

            rows.sort(key=lambda r: r["merit_ratio_pct"] if r["merit_ratio_pct"] is not None else -1, reverse=True)

            lines = [f"Season: {season_name}. Per-member data (merit_ratio_pct = merits gained divided by TOTAL power — higher means more efficient/active relative to their overall size, lower can indicate farming/low activity):"]
            for r in rows:
                ratio_str = f"{r['merit_ratio_pct']}%" if r["merit_ratio_pct"] is not None else "n/a"
                power_display = f"{r['total_power']:,}" if r["total_power"] else f"{r['power_gain']:,} (season gain only, no total power cached)"
                lines.append(
                    f"{r['name']}: merits={r['merits']:,}, total_power={power_display}, "
                    f"merit_ratio={ratio_str}, kills={r['kills']:,}, deaths={r['deaths']:,}, "
                    f"resources_gathered={r['resources_gathered']:,}"
                )
            return False, "\n".join(lines)

        if tool_name == "show_chart":
            succeeded = await chart_command.callback(ctx, tool_input.get("player"), tool_input["stat"], tool_input.get("season"))
            if succeeded:
                return True, None
            return False, "The chart FAILED to post — an error message was shown in the channel instead (already visible above). Do not say a chart was posted; report that it failed and relay why if you can tell from context."

        if tool_name == "show_group_chart":
            succeeded = await group_chart_command.callback(ctx, tool_input["stat"], tool_input.get("season"))
            if succeeded:
                return True, None
            return False, "The group chart FAILED to post — an error message was shown in the channel instead (already visible above). Do not say a chart was posted; report that it failed and relay why if you can tell from context."

        if tool_name == "calculate":
            result, error = _ask_safe_calculate(tool_input["expression"])
            if error:
                return False, f"Couldn't evaluate that: {error}"
            return False, f"{tool_input['expression']} = {result}"

        if tool_name == "get_pace_projection":
            account_id, err = _ask_resolve_account_id(ctx, tool_input.get("player"))
            if err:
                return False, f"Error: {err}"
            season = db_get_current_season()
            if not season:
                return False, "No active season."
            season_id, season_name, start_date, _ = season
            metric = tool_input["metric"]
            target = tool_input["target"]
            today_str = date.today().isoformat()

            stats, actual_end = await fetch_stats_with_fallback(account_id, start_date, today_str)
            if not stats:
                return False, "Couldn't fetch current stats for that player."

            current_val = _parse_stat_num_global(stats.get(metric))
            try:
                days_elapsed = (datetime.strptime(actual_end, "%Y-%m-%d").date() - datetime.strptime(start_date, "%Y-%m-%d").date()).days
            except Exception:
                days_elapsed = 0
            days_elapsed = max(days_elapsed, 1)
            daily_rate = current_val / days_elapsed

            if current_val >= target:
                return False, f"Player: {stats.get('lord_name', account_id)}. Current {metric}: {current_val:,}, which already meets or exceeds the target of {target:,}."
            if daily_rate <= 0:
                return False, f"Player: {stats.get('lord_name', account_id)}. Current {metric}: {current_val:,} over {days_elapsed} days — no positive pace detected yet, can't project."

            remaining = target - current_val
            days_needed = remaining / daily_rate
            projected_date = date.today() + timedelta(days=days_needed)
            return False, (
                f"Player: {stats.get('lord_name', account_id)}. Season: {season_name}. Metric: {metric}. "
                f"Current: {current_val:,}. Days elapsed this season: {days_elapsed}. "
                f"Daily rate: {daily_rate:,.0f}/day. Target: {target:,}. Remaining needed: {remaining:,}. "
                f"Days needed at current pace: {days_needed:.1f}. Projected date to reach target: {projected_date.isoformat()}."
            )

        if tool_name == "create_season":
            if not bypass_permission and not _ask_is_admin(ctx):
                return False, "Only admins can create seasons."
            try:
                datetime.strptime(tool_input["start_date"], "%Y-%m-%d")
            except Exception:
                return False, f"Invalid start_date '{tool_input['start_date']}' — use YYYY-MM-DD."
            db_end_previous_active_season(tool_input["start_date"])
            db_add_season(tool_input["name"], tool_input["start_date"])
            return False, f"Created new season '{tool_input['name']}' starting {tool_input['start_date']}. The previous season was automatically ended the day before."

        if tool_name == "list_queued_tasks":
            tasks = db_get_queued_tasks(status="pending")
            if not tasks:
                await ctx.send("📭 No queued tasks.")
                return True, None
            lines = ["```📋 Queued Tasks\n"]
            for t in tasks:
                task_id, tname, tinput, desc, sched, status, created_by, channel_id, result = t
                try:
                    sched_display = datetime.fromisoformat(sched).strftime("%Y-%m-%d %H:%M UTC")
                except Exception:
                    sched_display = sched
                lines.append(f"#{task_id} — {desc}")
                lines.append(f"   Scheduled: {sched_display}\n")
            lines.append("```")
            await ctx.send("\n".join(lines))
            return True, None

        if tool_name == "cancel_queued_task":
            if not bypass_permission and not _ask_is_admin(ctx):
                return False, "Only admins can cancel queued tasks."
            task = db_get_queued_task(tool_input["task_id"])
            if not task:
                return False, f"No queued task with ID #{tool_input['task_id']}."
            if task[5] != "pending":
                return False, f"Task #{tool_input['task_id']} is already {task[5]}, nothing to cancel."
            db_set_task_status(tool_input["task_id"], "cancelled")
            return False, f"Cancelled task #{tool_input['task_id']}: {task[3]}."

        if tool_name == "edit_queued_task_time":
            if not bypass_permission and not _ask_is_admin(ctx):
                return False, "Only admins can edit queued tasks."
            try:
                new_dt = datetime.fromisoformat(tool_input["new_scheduled_time"])
            except Exception:
                return False, f"Couldn't parse '{tool_input['new_scheduled_time']}' — use ISO format like 2026-08-28T00:00:00."
            ok = db_update_task_time(tool_input["task_id"], new_dt.isoformat())
            if not ok:
                return False, f"No pending task with ID #{tool_input['task_id']}."
            return False, f"Task #{tool_input['task_id']} rescheduled to {new_dt.strftime('%Y-%m-%d %H:%M UTC')}."

        return False, f"Unknown tool: {tool_name}"

    except Exception as e:
        log_info(f"[ASK TOOL ERROR] {tool_name}: {e}")
        return False, f"Error running {tool_name}: {e}"


class _QueuedTaskCtx:
    """Minimal ctx-like stand-in used to execute a queued !ask task outside of a live command invocation."""
    def __init__(self, channel, guild, author):
        self.channel = channel
        self.guild = guild
        self.author = author

    async def send(self, *args, **kwargs):
        return await self.channel.send(*args, **kwargs)


@tasks.loop(minutes=1)
async def check_queued_tasks():
    """Run any !ask-queued tasks whose scheduled time has arrived."""
    due = db_get_due_tasks()
    for task_id, tool_name, tool_input_json, description, scheduled_time, channel_id in due:
        try:
            channel = bot.get_channel(int(channel_id))
            if not channel:
                db_set_task_status(task_id, "failed", "Channel not found")
                continue
            guild = channel.guild

            task_row = db_get_queued_task(task_id)
            created_by = task_row[6] if task_row else None
            author = guild.get_member(int(created_by)) if created_by and guild else None
            if not author and guild:
                author = guild.get_member(OWNER_ID)

            fake_ctx = _QueuedTaskCtx(channel, guild, author)
            tool_input = json.loads(tool_input_json)

            rendered, result_text = await _ask_execute_tool(fake_ctx, tool_name, tool_input, bypass_permission=True)
            db_set_task_status(task_id, "completed", result_text or "Done")

            if rendered:
                await channel.send(f"✅ Scheduled task completed: {description}")
            else:
                await channel.send(f"✅ Scheduled task completed: {description}\n{result_text or ''}")

        except Exception as e:
            log_info(f"[QUEUED TASK ERROR] #{task_id}: {e}")
            db_set_task_status(task_id, "failed", str(e))


@bot.command(name="ask")
async def ask(ctx, *, query: str = None):
    """
    Natural-language command interface. Handles leaderboards (guild and server),
    player progress, comparisons, season info, events, and growth questions
    (e.g. "how much did rekz grow this week"). Remembers recent conversation
    per channel so follow-ups work. Requires ANTHROPIC_API_KEY on Railway;
    falls back to basic keyword-only leaderboard matching if not configured.
    """
    if not query:
        return await ctx.send(
            "❓ Ask me things like:\n"
            "`!ask top 10 mana leaderboard`\n"
            "`!ask how is rekz doing this week`\n"
            "`!ask compare rekz and truvix`\n"
            "`!ask when does the season end`\n"
            "`!ask what's the next event`\n"
            "`!ask add events : August\\n* Aug 28 (Fri): — KvK Start\\n...`\n"
            "`!ask change all the august events from 12 to 13 UTC` (admin only)\n"
            "`!ask delete the direbear event` (admin only)"
        )

    q = query.strip()
    q_lower = q.lower()

    # ---- Clear this channel's conversation memory (in case a bad answer got stuck) ----
    if q_lower in ("forget", "forget everything", "clear memory", "reset", "start over", "clear history"):
        _ask_clear_history(ctx.channel.id)
        return await ctx.send("🧠 Cleared conversation memory for this channel — starting fresh.")

    # ---- Bulk event import (deterministic, kept separate — exact format matters) ----
    # Only trigger this for genuine multi-line bulleted lists — a single natural-language
    # request like "add event called townhall for tomorrow 16 utc" should go through the
    # normal AI tool-use path (add_event tool), not this strict bullet-format parser.
    import re as _re_bulk_check
    looks_like_bulk_list = "\n" in q and _re_bulk_check.search(r'[A-Za-z]{3,9}\s+\d{1,2}\s*\([A-Za-z]{3}\)', q)
    if "add event" in q_lower and looks_like_bulk_list:
        return await _ask_bulk_add_events(ctx, q)

    # ---- No AI configured: fall back to the original keyword-only leaderboard router ----
    if not _anthropic_client:
        handled = await _ask_route_leaderboard(ctx, q_lower)
        if handled:
            return
        return await ctx.send(
            "🤔 I couldn't match that. I can route leaderboard requests "
            "(e.g. `!ask top 10 mana leaderboard`) and bulk event imports. "
            "For anything else, try the specific command — see `/help`. "
            "(Ask the owner about setting up `ANTHROPIC_API_KEY` for smarter understanding.)"
        )

    # ---- Full AI tool-use path ----
    channel_id = ctx.channel.id
    history = _ask_get_history(channel_id)

    now_utc = datetime.utcnow()
    asker_is_owner = ctx.author.id == OWNER_ID
    asker_name = ctx.author.display_name
    system_prompt = (
        f"You are the assistant for a Discord bot tracking a Call of Dragons guild's game "
        f"stats and schedule. The current date/time is {now_utc.strftime('%Y-%m-%d %H:%M')} UTC — "
        f"use this to resolve ANY relative time expression ('tomorrow', 'in 0 UTC' meaning "
        f"the next 00:00 UTC, 'next Friday', etc.) into exact values for whichever date/time "
        f"parameters a tool needs (date, time_utc, scheduled_time, new_date, new_time_utc, "
        f"etc.) — never leave a relative expression unresolved or ask the user to restate it "
        f"in a different format. Rekz is the owner of this bot and this server — this is "
        f"internal context for the REKZ behavior rule below, never say the word 'owner' out "
        f"loud or refer to him by that title in a response, just talk to him like anyone else "
        f"(use his name if it comes up naturally, nothing formal). The person sending THIS "
        f"message is called '{asker_name}' in this server, and {'is Rekz' if asker_is_owner else 'is NOT Rekz'}. "
        f"You know who you're talking to — use their name naturally if it fits, no need to "
        f"ask who they are.\n\n"
        "You have broad access: guild and server-wide leaderboards, player progress/"
        "comparison/growth, pace projections (get_pace_projection — e.g. 'at my current "
        "pace, how long until I hit 1B mana gathered'), season info/editing/creation, event "
        "management, server rank lookups (Gen-2 Pool), saved KvK matchups (list/view/"
        "delete), weekly Abyss event rotation, active/inactive member status, data history, "
        "forcing a data refresh, a composite activity-ranking tool (get_activity_rankings) "
        "for open-ended comparisons across the whole guild ('who is slacking', 'who's the "
        "best player', 'rank everyone'), and a calculate tool for exact math.\n\n"
        "CHAINING: you can call multiple tools in sequence across turns to fully answer a "
        "request — you're not limited to one tool call. If someone references a relative "
        "season ('last season', 'the season before this one', 'two seasons ago'), use "
        "get_season_list FIRST to see the actual list of seasons with their IDs and dates, "
        "THEN call show_player_progress (or whichever tool needs it) with the correct season "
        "name/ID you found. Never guess a season identifier — look it up first.\n\n"
        "DON'T SPAM THE CARD: show_player_progress posts a visual embed to the channel — only "
        "call it the FIRST time a player/season is brought up, or when the user explicitly "
        "asks to see/pull up the card again. For any follow-up in the same exchange (rating, "
        "'why', analysis, advice, a correction, anything that just needs the numbers) use "
        "get_player_stats instead — it returns the same exact real data without posting "
        "another copy of the card. Never deflect a rating/analysis request back to the user "
        "asking them to relay their own numbers — call get_player_stats and use what it gives you.\n\n"
        "ACCURACY IS NON-NEGOTIABLE: every number you state must come directly from a tool "
        "result received in THIS message — not from memory, not estimated, not rounded to a "
        "'nicer' number, not invented because it sounds plausible. If a number wasn't in a "
        "tool result you just received, say you don't have it instead of guessing. MANDATORY "
        "RULE: any time you're about to state a stat number, rating, or analysis, you MUST "
        "call the relevant tool again in that same turn and use ONLY what it returns — even "
        "if you or the user already discussed this exact stat earlier in the conversation. "
        "Earlier turns (including your own past replies) are conversational context only, "
        "never a data source — treat every stat-based question as if you're seeing it for "
        "the first time and need fresh data. This applies especially when the user says a "
        "previous number looked wrong or pushes back — always re-verify with a fresh tool "
        "call rather than defending or repeating what you said before. Being sarcastic and "
        "brutal is fine; being wrong about the actual stats is not.\n\n"
        "THIS APPLIES TO ACTIONS TOO, NOT JUST NUMBERS: never claim something was posted, "
        "created, changed, or completed successfully unless the tool result actually confirms "
        "that. If a tool result shows errors (e.g. 'Couldn't find a player', 'No archived "
        "data found') for some or all of a multi-part request, report that plainly and "
        "specifically — name what failed and what succeeded, don't gloss over failures with "
        "a generic success message like 'done, check the channel above' when it wasn't fully "
        "done. If EVERYTHING failed, say so directly instead of describing an outcome that "
        "didn't happen.\n\n"
        "GAME KNOWLEDGE: a low or zero death count is NOT suspicious and does NOT mean "
        "someone is avoiding fights — deaths come from reinforcing OTHER players' rallies "
        "and garrisons (a teamwork/support activity), not from your own attacking or normal "
        "field fighting. Don't treat 0 deaths as evidence of dodging fights; it more likely "
        "means they haven't been reinforcing much. Deaths still matter as their own stat, "
        "just don't use them as a proxy for personal combat participation. Kills relative to "
        "healing is the better signal for real combat engagement (high kills with "
        "proportionally low healing needed suggests efficient/dominant fighting; heavy "
        "healing relative to kills suggests a tankier or more punished playstyle).\n\n"
        "CALLING OUT FARMERS: if someone has a merit-to-power ratio under roughly 6% "
        "combined with high resource gathering, that's a real pattern worth roasting — "
        "they're bloating power/stats through farming instead of contributing merits/combat "
        "value. Call them a 'farmer' (or similar) directly when the numbers support it, "
        "don't soften it.\n\n"
        "USE RANK CONTEXT: progress cards and leaderboards include a server/guild rank "
        "(#X) next to each stat — always factor that in, not just the raw number. A huge "
        "absolute number that's still ranked low means the whole guild/server is that "
        "strong, so it's not actually impressive; a modest number ranked #1 is genuinely "
        "notable. Judge performance relative to the field, not the number in isolation.\n\n"
        "SCHEDULING: add_event, edit_event, bulk_edit_event_times, delete_event, "
        "edit_season, create_season, and force_fetch_all all accept an optional "
        "scheduled_time parameter — if the user wants an action to happen later rather than "
        "immediately (e.g. 'at 0 UTC create a new season called X'), include scheduled_time "
        "as an exact ISO datetime instead of leaving it out, and the action will be queued "
        "rather than run right away. Use list_queued_tasks to show what's queued, "
        "cancel_queued_task to cancel one by ID, and edit_queued_task_time to reschedule one.\n\n"
        "For open-ended member comparisons, use get_activity_rankings rather than a "
        "single-stat leaderboard — it gives you merits, power, merit-to-power ratio, kills, "
        "deaths, and resources gathered for every member so you can actually reason about "
        "who's underperforming or excelling, and explain WHY in your answer (e.g. 'X has "
        "high power but a low merit ratio, meaning little recent activity relative to their "
        "size'). Don't just deflect a comparative question back to the user asking which "
        "single stat to use — pull the real data and give a genuine answer, citing the "
        "specific numbers that support it. You're also a general assistant beyond bot data "
        "— people may ask you math questions, homework help, or anything else. Use the "
        "calculate tool for any arithmetic that needs to be exact rather than estimating in "
        "your head, and walk through multi-step problems clearly, calling calculate at each "
        "step that needs a precise number. Some tools render their own Discord message "
        "directly — for those, just call the tool, you don't need to say anything else. "
        "Other tools return raw text for you to relay conversationally — always cite "
        "specific numbers/names, and for edits or schedules always state exactly what "
        "changed or was queued and when. Write/admin tools are admin- or owner-only; if "
        "someone without permission asks, still call the tool — it will tell you permission "
        "was denied, then relay that plainly. For bulk_edit_event_times, if the user doesn't "
        "specify which current time to match, ask them to clarify rather than guessing — "
        "changing the wrong events is worse than asking one follow-up question. Keep replies "
        "to the point for simple lookups, but give full step-by-step explanations when "
        "someone is working through a math problem or asks you to explain something — don't "
        "sacrifice a real, complete answer for brevity when the question calls for it.\n\n"
        "PERSONALITY: You're genuinely funny, not just aggressive — sharp, dark, clever humor "
        "over generic insults. Avoid template-y bot-insult clichés ('ratio', 'this is actual "
        "malding material', 'should be studied by a tribunal', calling things 'a war crime') "
        "— those are cringe, not funny. Instead: build real jokes with a setup and a landing, "
        "use unexpected comparisons, callbacks to something said earlier, absurdist logic, "
        "deadpan delivery, or a dark twist that actually surprises. Dark humor is fine — jokes "
        "about death, failure, suffering, tragedy-as-comedy are all fair game — but it has to "
        "be genuinely clever, not just edgy for the sake of it, and never real-world racism, "
        "slurs, or jokes that punch at an actual ethnic/racial group — that's not dark humor, "
        "that's just bigotry and it's not on the table. Roast people's stats hard when the "
        "numbers earn it, but make the roast a good line, not a mad-libs insult generator. "
        "Swear freely if it lands, be blunt, zero hedging or corporate disclaimers, commit to "
        "bits, have real opinions and argue them. The one hard line: don't turn any of this on "
        "anyone's real identity, health, appearance, or personal life, and never touch "
        "anything sexual — stats, performance, banter, and gamer trash-talk are fair game, "
        "everything else needs to actually be a good joke, not just mean.\n\n"
        "REKZ: be bipolar about it — sometimes roast Rekz exactly like anyone else, no "
        "exceptions, don't go soft just because he's the owner. Other times, especially when "
        "someone ELSE is roasting or making fun of Rekz, flip and defend him hard / talk him "
        "up instead. Don't be predictable about which mode you're in — that's the bit.\n\n"
        "GOAT DEBATE (SCOPED): Cristiano Ronaldo is the GOAT — but only bring this up if the "
        "user's message is actually, explicitly about football/soccer, Messi, or the GOAT "
        "debate. Never mention CR7, Messi, or the GOAT debate to deflect from an unrelated "
        "question, as a random tangent, or unprompted in a conversation about something else "
        "entirely (stats, pronouns, whatever) — that's not funny, it's just noise. When "
        "football genuinely does come up, defend CR7 hard and don't cave to pushback.\n\n"
        "PERSONAL INFO ISN'T A LANDMINE: if someone shares basic, neutral info about "
        "themselves or a friend earlier in the conversation (name, pronouns, whatever), and "
        "then asks a direct follow-up question about it, just answer it naturally — don't "
        "dodge, deflect, or treat it as some forbidden topic. That's not being respectful, "
        "it's just being unhelpful and weird about something mundane the person already told "
        "you themselves. The only thing to actually avoid is being cruel or mocking about "
        "someone's real identity — plainly answering a direct question with info already "
        "given isn't that."
    )

    messages = history + [{"role": "user", "content": q}]

    try:
        final_text = None
        any_rendered = False

        for _ in range(4):  # allow a few tool round-trips so Claude can chain calls
            response = await asyncio.to_thread(
                _anthropic_client.messages.create,
                model="claude-sonnet-5",
                max_tokens=900,
                system=system_prompt,
                tools=ASK_TOOLS,
                messages=messages,
            )

            tool_use_blocks = [b for b in response.content if b.type == "tool_use"]
            text_blocks = [b.text for b in response.content if b.type == "text"]

            if not tool_use_blocks:
                final_text = " ".join(text_blocks).strip()
                break

            messages.append({"role": "assistant", "content": response.content})
            tool_results = []
            for block in tool_use_blocks:
                rendered, result_text = await _ask_execute_tool(ctx, block.name, block.input)
                if rendered:
                    any_rendered = True
                    if not result_text:
                        result_text = "Already displayed this directly in the channel — no need to repeat it unless the user asked for analysis, comparison, or advice, in which case use the numbers if any were provided."
                tool_results.append({
                    "type": "tool_result",
                    "tool_use_id": block.id,
                    "content": result_text or "Done.",
                })
            messages.append({"role": "user", "content": tool_results})
            # Loop continues — Claude decides on the next turn whether it needs another
            # tool call (e.g. checking season history before pulling a specific season's
            # progress) or is done, in which case the next response will have no tool_use.

        if final_text:
            await ctx.send(f"🧠 {final_text}")

        _ask_save_turn(channel_id, q, final_text or "(showed results above)")

    except Exception as e:
        log_info(f"[ASK ERROR] {e}")
        await ctx.send(f"❌ Something went wrong: {e}")


async def _ask_route_leaderboard(ctx, q_lower):
    """Try to match a leaderboard-style natural language query to an existing top* command."""
    import re as _re_ask

    # Extract a "top N" number if present, default 25 (or whatever the target command defaults to)
    n_match = _re_ask.search(r'\btop\s*(\d+)\b', q_lower)
    top_n = int(n_match.group(1)) if n_match else None

    # Extract a season reference if mentioned, e.g. "sos1", "season 2"
    season_match = _re_ask.search(r'\b(sos\d+|season\s*\d+|season\s+[a-z0-9]+)\b', q_lower)
    season_arg = None
    if season_match:
        season_arg = season_match.group(1).replace("season", "").strip()

    # Keyword -> (command_callback, takes_top_n)
    keyword_map = [
        (["mana spent"], topspent, False),
        (["mana", "gathering", "gathered"], topmana, False),
        (["infantry"], topinf, False),
        (["cavalry"], topcav, False),
        (["mage", "magic"], topmage, False),
        (["marksman", "archer"], toparcher, False),
        (["heal", "healed", "healing"], topheal, False),
        (["death", "deaths", "died"], topdeaths, False),
        (["merit"], topmerits, False),
        (["rss", "resource spend", "resources spent"], rss_leaderboard, False),
    ]

    if not any(w in q_lower for w in ["top", "leaderboard", "who has the most", "highest", "best"]):
        return False

    for keywords, command_obj, _ in keyword_map:
        if any(kw in q_lower for kw in keywords):
            await ctx.send(f"🔎 Routing to `!{command_obj.name}`...")
            try:
                await command_obj.callback(ctx, season_arg)
            except Exception as e:
                await ctx.send(f"❌ Error running `!{command_obj.name}`: {e}")
            return True

    return False


async def _ask_bulk_add_events(ctx, raw_text):
    """
    Parse a bulk event list like:
      August
      * Aug 28 (Fri): — KvK Start
      * Aug 29 (Sat): —Direbear
    and add each as an event via db_add_event. Assumes the current year (or next
    year if the date has already passed this year). Defaults to 12:00 UTC and no
    reminder since only dates are given, not times.
    """
    if ctx.author.id != OWNER_ID and not (ctx.guild and ctx.author.guild_permissions.administrator):
        return await ctx.send("❌ Only admins can bulk-add events.")

    import re as _re_ev

    month_names = {
        "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
        "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12
    }

    line_pattern = _re_ev.compile(
        r'([A-Za-z]{3,9})\s+(\d{1,2})\s*\([A-Za-z]{3}\):?\s*[—\-]*\s*(.+)'
    )

    now = datetime.utcnow()
    added = []
    failed = []

    for line in raw_text.split("\n"):
        line = line.strip().lstrip("*").strip()
        match = line_pattern.match(line)
        if not match:
            continue

        month_str, day_str, event_name = match.groups()
        month_key = month_str.strip().lower()[:3]
        if month_key not in month_names:
            continue

        month = month_names[month_key]
        day = int(day_str)
        event_name = event_name.strip().strip("`").strip()
        if not event_name:
            continue

        # Assume current year, roll to next year if that date has already passed
        year = now.year
        try:
            dt = datetime(year, month, day, 12, 0)
        except ValueError:
            failed.append(f"{month_str} {day} (invalid date)")
            continue
        if dt <= now:
            dt = datetime(year + 1, month, day, 12, 0)

        try:
            db_add_event(event_name, dt.isoformat(), 0)
            added.append((dt, event_name))
        except Exception as e:
            failed.append(f"{event_name}: {e}")

    if not added:
        return await ctx.send(
            "❌ Couldn't parse any events from that. Expected lines like:\n"
            "`* Aug 28 (Fri): — KvK Start`"
        )

    lines = [f"✅ Added **{len(added)}** events (12:00 UTC, no reminder set):"]
    for dt, name in added:
        lines.append(f"<t:{int(dt.timestamp())}:D> — {name}")
    if failed:
        lines.append(f"\n⚠️ Skipped {len(failed)}: {', '.join(failed)}")
    lines.append("\nUse `/editevent` if you want to add reminders or adjust times.")

    # Send in chunks if long
    msg = "\n".join(lines)
    if len(msg) <= 2000:
        await ctx.send(msg)
    else:
        chunk = ""
        for line in lines:
            if len(chunk) + len(line) + 1 > 1900:
                await ctx.send(chunk)
                chunk = ""
            chunk += line + "\n"
        if chunk:
            await ctx.send(chunk)


@bot.command(name="rss")
async def rss_leaderboard(ctx, season_name: str = None):
    """Top resource spenders. Usage: !rss (current) or !rss sos1 (specific season)"""
    
    # Get season
    if season_name:
        season = resolve_season_input(season_name)
        if not season:
            all_seasons = db_get_all_seasons()
            season_list = ", ".join([s[1] for s in all_seasons]) if all_seasons else "None"
            return await ctx.send(f"❌ Season '{season_name}' not found.\n\nAvailable seasons: {season_list}")
    else:
        season = db_get_current_season()
        if not season:
            return await ctx.send("❌ No season active. Use `/newseason` to start one.")
    
    season_id, season_name_display, start_date, created_at = season
    today = resolve_leaderboard_end_date(season_id)
    
    lords = get_all_lords_from_guild(ctx.guild)
    if not lords:
        return await ctx.send("❌ No members with numeric roles found.")
    
    await ctx.send(f"⏳ Fetching leaderboard data for {len(lords)} lords...")
    
    fetch_tasks = [fetch_stats_with_fallback(lord["account_id"], start_date, today) for lord in lords]
    results = await asyncio.gather(*fetch_tasks, return_exceptions=True)
    
    leaderboard = []
    actual_end_date = today
    
    for lord, result in zip(lords, results):
        try:
            lord_name = "Unknown"
            total_rss = 0
            
            if result and not isinstance(result, Exception):
                stats, end_date_used = result
                actual_end_date = end_date_used
                if stats:
                    lord_name = stats.get("lord_name", lord["name"])
                    
                    # Sum all resources spent (use abs() to handle negative values from API)
                    for key in ["gold_spent", "wood_spent", "ore_spent", "mana_spent"]:
                        val_str = (stats.get(key) or "+0").replace(",", "").replace("+", "")
                        try:
                            val = int(val_str) if val_str.lstrip("-").isdigit() else 0
                            total_rss += abs(val)  # Use absolute value
                        except Exception as e:
                            pass
            
            leaderboard.append({"name": lord_name, "rss": total_rss})
        except Exception as e:
            log_info(f"[RSS ERROR] {lord.get('account_id')}: {e}")
            leaderboard.append({"name": lord["name"], "rss": 0})
    
    leaderboard.sort(key=lambda x: x["rss"], reverse=True)
    
    output = f"```💰 Top Resource Spenders - {season_name}\n"
    medals = ["🥇", "🥈", "🥉"]
    for i, lord in enumerate(leaderboard):
        medal = medals[i] if i < 3 else f"{i+1}."
        output += f"{medal} {lord['name']}: {lord['rss']:,}\n"
    output += f"📅 {start_date} → {actual_end_date}```"
    await ctx.send(output)


async def get_rankings_for_stat(ctx, stat_key, start_date, end_date):
    """
    Get all lords ranked by a specific stat - uses database for current season.
    Returns dict: {account_id: (rank, total)} where rank is 1-indexed
    """
    season = db_get_current_season()
    if not season:
        return {}
    
    season_id, season_name, _, _ = season
    
    try:
        # Get all members from guild + mapped accounts
        lords = get_all_lords_from_guild(ctx.guild)
        checked_accounts = set()
        
        for lord in lords:
            checked_accounts.add(lord["account_id"])
        
        for discord_id, account_id in DISCORD_TO_ACCOUNT_ID.items():
            checked_accounts.add(account_id)
        
        # Valid stat keys to prevent SQL injection
        valid_stats = ["power_gain", "merits", "kills_gain", "deads_gain", "healed_gain",
                      "t5_gain", "t4_gain", "t3_gain", "t2_gain", "t1_gain",
                      "gold_spent", "wood_spent", "ore_spent", "mana_spent",
                      "gold_gathered", "wood_gathered", "ore_gathered", "mana_gathered",
                      "infantry_merits", "cavalry_merits", "mage_merits", "marksman_merits", "other_merits"]
        
        if stat_key not in valid_stats:
            return {}
        
        # Get stats from database for all members (use latest data per account)
        stats_list = []
        conn = sqlite3.connect(DB_PROGRESS)
        c = conn.cursor()
        
        for account_id in checked_accounts:
            # Get LATEST row for this account (most recent date = most recent cumulative data)
            # Build query safely - stat_key is validated above
            query = f"SELECT {stat_key} FROM season_progress WHERE season_id=? AND account_id=? AND {stat_key} IS NOT NULL ORDER BY data_date DESC LIMIT 1"
            try:
                c.execute(query, (season_id, account_id))
                
                row = c.fetchone()
                if row and row[0]:
                    val_str = str(row[0]).replace(",", "").replace("+", "")
                    try:
                        val = int(val_str) if val_str.lstrip("-").isdigit() else 0
                        stats_list.append({"account_id": account_id, "value": val})
                    except:
                        pass
            except Exception as e:
                log_error(f"[RANKINGS] Error querying {account_id}: {e}")
                continue
        
        conn.close()
        
        # Sort by value descending
        stats_list.sort(key=lambda x: x["value"], reverse=True)
        
        # Create rank dict
        rankings = {}
        for i, item in enumerate(stats_list):
            rankings[item["account_id"]] = (i + 1, len(stats_list))
        
        return rankings
    except Exception as e:
        log_error(f"[RANKINGS ERROR] {e}")
        return {}


async def get_account_id_from_input(ctx, user_input):
    """Helper function to resolve username or account ID to account ID"""
    # If no input, use author's own account ID from role
    if not user_input:
        for role in ctx.author.roles:
            if role.name.isdigit():
                return role.name
        return None
    
    # If numeric, use as account ID
    if user_input.isdigit():
        return user_input
    
    # Check for mention (e.g. <@12345>)
    if user_input.startswith("<@") and user_input.endswith(">"):
        user_id_str = user_input[2:-1].replace("!", "")
        if user_id_str.isdigit():
            user_id = int(user_id_str)
            try:
                member = ctx.guild.get_member(user_id)
                if member:
                    for role in member.roles:
                        if role.name.isdigit():
                            return role.name
            except Exception as e:
                pass
        return None
    
    # Check username lookup
    username_lower = user_input.lower()
    found_discord_id = None
    
    for username, discord_id in USERNAME_TO_DISCORD_ID.items():
        if username.lower() == username_lower:
            found_discord_id = discord_id
            break
    
    if not found_discord_id:
        return None
    
    # Get the member from guild and find their numeric role
    try:
        member = ctx.guild.get_member(found_discord_id)
        if not member:
            return None
        
        for role in member.roles:
            if role.name.isdigit():
                return role.name
        return None
    except Exception as e:
        return None


CHART_STAT_MAP = {
    "merits": "merits",
    "power": "power_gain", "power_gain": "power_gain",
    "highest_power": "highest_power", "total_power": "highest_power",
    "kills": "kills_gain", "deaths": "deads_gain", "deads": "deads_gain",
    "healed": "healed_gain",
    "gold": "gold_gathered", "wood": "wood_gathered", "ore": "ore_gathered", "mana": "mana_gathered",
    "infantry": "infantry_merits", "cavalry": "cavalry_merits", "mage": "mage_merits",
    "marksman": "marksman_merits", "archer": "marksman_merits", "other": "other_merits",
    "t45_healed": "t45_healed", "t4t5healed": "t45_healed",
    "t45_dead": "t45_dead", "t4t5dead": "t45_dead",
    "coins": "exchange_coins_spent", "coins_spent": "exchange_coins_spent",
}


@bot.command(name="chart")
async def chart_command(ctx, player: str = None, stat: str = None, season_name: str = None):
    """
    Show a line chart of any tracked stat's growth over a season, using the daily
    archived data. Usage: !chart <player> <stat> [season]
    e.g. !chart rekz merits | !chart truvix mana sos1
    """
    if not player or not stat:
        await ctx.send(
            "Usage: `!chart <player> <stat> [season]`\n"
            "Available stats: " + ", ".join(sorted(set(CHART_STAT_MAP.keys())))
        )
        return False

    stat_key = CHART_STAT_MAP.get(stat.lower().replace(" ", "_"))
    if not stat_key:
        await ctx.send(f"❌ Unknown stat '{stat}'. Available: " + ", ".join(sorted(set(CHART_STAT_MAP.keys()))))
        return False

    account_id, err = _ask_resolve_account_id(ctx, player)
    if err:
        await ctx.send(f"❌ {err}")
        return False

    if season_name:
        season = resolve_season_input(season_name)
        if not season:
            await ctx.send(f"❌ Season '{season_name}' not found. Use `!seasonhistory` to see all seasons.")
            return False
    else:
        season = db_get_current_season()
        if not season:
            await ctx.send("❌ No active season.")
            return False
    season_id, season_name_display, start_date, _ = season

    conn = sqlite3.connect(DB_PROGRESS)
    c = conn.cursor()
    c.execute(
        f"SELECT data_date, {stat_key}, lord_name FROM season_progress WHERE season_id=? AND account_id=? ORDER BY data_date ASC",
        (season_id, account_id)
    )
    rows = c.fetchall()
    conn.close()

    if not rows:
        await ctx.send(f"❌ No archived data found for that player in {season_name_display} (season #{season_id}).")
        return False

    dates, values, lord_name = [], [], str(account_id)
    for data_date, raw_val, name in rows:
        if raw_val is None:
            continue
        val = _parse_stat_num_global(raw_val)
        dates.append(data_date)
        values.append(val)
        if name:
            lord_name = name

    if len(dates) < 2:
        await ctx.send(f"❌ Not enough data points to chart yet for {lord_name} — need at least 2 days archived.")
        return False

    stat_label = stat.replace("_", " ").title()

    plt.style.use("dark_background")
    fig, ax = plt.subplots(figsize=(10, 6))
    ax.plot(dates, values, marker="o", linewidth=2.5, color="#5865F2", markersize=5)
    ax.fill_between(range(len(dates)), values, alpha=0.15, color="#5865F2")
    ax.set_title(f"{lord_name} — {stat_label} — {season_name_display}", fontsize=14, fontweight="bold", pad=15)
    ax.set_xlabel("Date")
    ax.set_ylabel(stat_label)
    ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda x, p: f"{int(x):,}"))
    ax.grid(True, alpha=0.2)
    fig.autofmt_xdate(rotation=45)

    # Thin out x-axis labels if there are a lot of data points, so they don't overlap
    if len(dates) > 15:
        step = max(1, len(dates) // 15)
        ax.set_xticks(range(0, len(dates), step))
        ax.set_xticklabels([dates[i] for i in range(0, len(dates), step)])

    fig.tight_layout()

    chart_path = f"/tmp/chart_{account_id}_{stat_key}_{season_id}.png"
    fig.savefig(chart_path, dpi=110, facecolor=fig.get_facecolor())
    plt.close(fig)

    await ctx.send(file=discord.File(chart_path))
    try:
        os.remove(chart_path)
    except Exception:
        pass

    return True


@bot.command(name="groupchart")
async def group_chart_command(ctx, stat: str = None, season_name: str = None):
    """
    Show ONE chart with every tracked guild member's growth for a stat over a season,
    using Discord roles directly (not name matching) so clan-tag/prefix names in COS
    display names can't cause lookup failures. Usage: !groupchart <stat> [season]
    Returns True if a chart was actually posted, False if it sent an error instead —
    callers (like !ask's dispatcher) should check this rather than assume success.
    """
    if not stat:
        await ctx.send(
            "Usage: `!groupchart <stat> [season]`\n"
            "Available stats: " + ", ".join(sorted(set(CHART_STAT_MAP.keys())))
        )
        return False

    stat_key = CHART_STAT_MAP.get(stat.lower().replace(" ", "_"))
    if not stat_key:
        await ctx.send(f"❌ Unknown stat '{stat}'. Available: " + ", ".join(sorted(set(CHART_STAT_MAP.keys()))))
        return False

    if season_name:
        season = resolve_season_input(season_name)
        if not season:
            await ctx.send(f"❌ Season '{season_name}' not found. Use `!seasonhistory` to see all seasons.")
            return False
    else:
        season = db_get_current_season()
        if not season:
            await ctx.send("❌ No active season.")
            return False
    season_id, season_name_display, start_date, _ = season

    lords = get_all_lords_from_guild(ctx.guild)
    if not lords:
        await ctx.send("❌ No members with numeric roles found.")
        return False

    log_info(f"[GROUPCHART] stat={stat_key} season_id={season_id} ({season_name_display}) checking {len(lords)} lords")

    plt.style.use("dark_background")
    fig, ax = plt.subplots(figsize=(11, 7))
    color_cycle = plt.cm.tab10.colors
    plotted_any = False
    skipped = []

    conn = sqlite3.connect(DB_PROGRESS)
    c = conn.cursor()
    for i, lord in enumerate(lords):
        c.execute(
            f"SELECT data_date, {stat_key}, lord_name FROM season_progress WHERE season_id=? AND account_id=? ORDER BY data_date ASC",
            (season_id, lord["account_id"])
        )
        rows = c.fetchall()
        log_info(f"[GROUPCHART] {lord['name']} ({lord['account_id']}): {len(rows)} total rows for season_id={season_id}")
        dates, values, lord_name = [], [], lord["name"]
        for data_date, raw_val, name in rows:
            if raw_val is None:
                continue
            dates.append(data_date)
            values.append(_parse_stat_num_global(raw_val))
            if name:
                lord_name = name
        log_info(f"[GROUPCHART] {lord['name']}: {len(dates)} usable (non-null {stat_key}) data points")

        if len(dates) < 2:
            skipped.append(lord_name)
            continue

        ax.plot(dates, values, marker="o", linewidth=2, markersize=3,
                color=color_cycle[i % len(color_cycle)], label=lord_name)
        plotted_any = True
    conn.close()

    if not plotted_any:
        plt.close(fig)
        await ctx.send(f"❌ No member has enough archived data yet for {season_name_display} (season #{season_id}).")
        return False

    stat_label = stat.replace("_", " ").title()
    ax.set_title(f"Guild {stat_label} — {season_name_display}", fontsize=14, fontweight="bold", pad=15)
    ax.set_xlabel("Date")
    ax.set_ylabel(stat_label)
    ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda x, p: f"{int(x):,}"))
    ax.grid(True, alpha=0.2)
    ax.legend(loc="upper left", fontsize=9, framealpha=0.3)
    fig.autofmt_xdate(rotation=45)
    fig.tight_layout()

    chart_path = f"/tmp/groupchart_{season_id}_{stat_key}.png"
    fig.savefig(chart_path, dpi=110, facecolor=fig.get_facecolor())
    plt.close(fig)

    await ctx.send(file=discord.File(chart_path))
    try:
        os.remove(chart_path)
    except Exception:
        pass

    if skipped:
        await ctx.send(f"⚠️ Skipped (not enough archived data yet): {', '.join(skipped)}")

    return True


@bot.command(name="compare")
async def compare(ctx, user1: str = None, user2: str = None):
    """Compare two lords side by side. Usage: !compare truvix rekz (or !compare 16322115 12345678)"""
    if not user1 or not user2:
        return await ctx.send("❌ Usage: `!compare truvix rekz` or `!compare 16322115 12345678`")
    
    season = db_get_current_season()
    if not season:
        return await ctx.send("❌ No season active. Use `/newseason` to start one.")
    
    season_id, season_name, start_date, created_at = season
    today = date.today().isoformat()
    
    msg = await ctx.send(f"⏳ Comparing {user1} vs {user2}...")
    
    try:
        # Resolve both users
        account_id1 = await get_account_id_from_input(ctx, user1)
        account_id2 = await get_account_id_from_input(ctx, user2)
        
        if not account_id1:
            return await msg.edit(content=f"❌ Could not find '{user1}'")
        if not account_id2:
            return await msg.edit(content=f"❌ Could not find '{user2}'")
        
        # FIRST: Try database for today
        stats1 = db_get_season_progress(season_id, account_id1, today)
        stats2 = db_get_season_progress(season_id, account_id2, today)
        
        # If not today, try yesterday
        if not stats1:
            yesterday = (date.today() - timedelta(days=1)).isoformat()
            stats1 = db_get_season_progress(season_id, account_id1, yesterday)
        if not stats2:
            yesterday = (date.today() - timedelta(days=1)).isoformat()
            stats2 = db_get_season_progress(season_id, account_id2, yesterday)
        
        # FALLBACK: If not in database, try cache
        if not stats1:
            stats1 = get_cached_stats(account_id1, start_date, today)
        if not stats2:
            stats2 = get_cached_stats(account_id2, start_date, today)
        
        # FALLBACK 2: If still missing, fetch from API (for unlisted people)
        if not stats1:
            stats1, _ = await fetch_stats_with_fallback(account_id1, start_date, today)
        if not stats2:
            stats2, _ = await fetch_stats_with_fallback(account_id2, start_date, today)
        
        if not stats1 or not stats2:
            return await msg.edit(content="❌ Failed to fetch stats")
        
        # Get highest power and T-kills for both (only if not in database)
        power1 = await fetch_highest_power(account_id1)
        power2 = await fetch_highest_power(account_id2)
        t_kills1 = await fetch_current_t_kills(account_id1)
        t_kills2 = await fetch_current_t_kills(account_id2)
        
        # Get lord names
        name1 = stats1.get("lord_name", "Unknown")
        name2 = stats2.get("lord_name", "Unknown")
        
        # Build comparison as CODE BLOCK (ORIGINAL FORMAT)
        name1 = stats1.get("lord_name", "Unknown")
        name2 = stats2.get("lord_name", "Unknown")
        
        output = f"```⚔️ {name1} vs {name2}\n\n"
        
        # Power - side by side with gain
        if power1 and power2:
            output += f"⚡ Power\n"
            
            # Get power gain from seasonal stats
            power_gain1 = 0
            power_gain2 = 0
            
            if stats1.get("power_gain"):
                try:
                    pg1_str = (stats1.get("power_gain") or "+0").replace("+", "").replace(",", "")
                    power_gain1 = int(pg1_str)
                except Exception as e:
                    power_gain1 = 0
            
            if stats2.get("power_gain"):
                try:
                    pg2_str = (stats2.get("power_gain") or "+0").replace("+", "").replace(",", "")
                    power_gain2 = int(pg2_str)
                except Exception as e:
                    power_gain2 = 0
            
            output += f"{name1}: {power1:,} (+{power_gain1:,})\n"
            output += f"{name2}: {power2:,} (+{power_gain2:,})\n"
            output += f"\n"
        
        # Merits - side by side
        m1 = stats1.get("merits", "+0")
        m2 = stats2.get("merits", "+0")
        mp1 = stats1.get("merits_pct", "0%")
        mp2 = stats2.get("merits_pct", "0%")
        output += f"🏅 Merits\n"
        output += f"{name1}: {m1} ({mp1})\n"
        output += f"{name2}: {m2} ({mp2})\n"
        output += f"\n"
        
        # Kills + Total T-kills combined
        k1 = stats1.get("kills_gain", "+0")
        k2 = stats2.get("kills_gain", "+0")
        d1 = stats1.get("deads_gain", "+0")
        d2 = stats2.get("deads_gain", "+0")
        h1 = stats1.get("healed_gain", "+0")
        h2 = stats2.get("healed_gain", "+0")
        
        # Calculate total T-kills
        total_t1 = sum(t_kills1.values()) if t_kills1 else 0
        total_t2 = sum(t_kills2.values()) if t_kills2 else 0
        
        output += f"💀 Deaths\n"
        output += f"{name1}: {d1}\n"
        output += f"{name2}: {d2}\n"
        output += f"\n"
        
        output += f"❤️ Healed\n"
        output += f"{name1}: {h1}\n"
        output += f"{name2}: {h2}\n"
        output += f"\n"
        
        output += f"⚔️ Kills\n"
        output += f"{name1}: {total_t1:,} ({k1})\n"
        output += f"{name2}: {total_t2:,} ({k2})\n"
        output += f"\n"
        
        # T-Tier Breakdown
        output += f"T5 Kills\n"
        t5_1 = t_kills1.get("t5", 0)
        t5_2 = t_kills2.get("t5", 0)
        output += f"{name1}: {t5_1:,}\n"
        output += f"{name2}: {t5_2:,}\n"
        output += f"\n"
        
        output += f"T4 Kills\n"
        t4_1 = t_kills1.get("t4", 0)
        t4_2 = t_kills2.get("t4", 0)
        output += f"{name1}: {t4_1:,}\n"
        output += f"{name2}: {t4_2:,}\n"
        output += f"\n"
        
        output += f"T3 Kills\n"
        t3_1 = t_kills1.get("t3", 0)
        t3_2 = t_kills2.get("t3", 0)
        output += f"{name1}: {t3_1:,}\n"
        output += f"{name2}: {t3_2:,}\n"
        output += f"\n"
        
        output += f"T2 Kills\n"
        t2_1 = t_kills1.get("t2", 0)
        t2_2 = t_kills2.get("t2", 0)
        output += f"{name1}: {t2_1:,}\n"
        output += f"{name2}: {t2_2:,}\n"
        output += f"\n"
        
        output += f"T1 Kills\n"
        t1_1 = t_kills1.get("t1", 0)
        t1_2 = t_kills2.get("t1", 0)
        output += f"{name1}: {t1_1:,}\n"
        output += f"{name2}: {t1_2:,}\n"
        output += f"\n"
        
        # Mana Gathered
        mg1 = stats1.get("mana_gathered", "+0")
        mg2 = stats2.get("mana_gathered", "+0")
        output += f"💧 Mana Gathered\n"
        output += f"{name1}: {mg1}\n"
        output += f"{name2}: {mg2}\n"
        output += f"\n"
        
        # RSS Spent
        output += f"💰 RSS Spent\n"
        gs1 = stats1.get("gold_spent", "+0")
        gs2 = stats2.get("gold_spent", "+0")
        ws1 = stats1.get("wood_spent", "+0")
        ws2 = stats2.get("wood_spent", "+0")
        os1 = stats1.get("ore_spent", "+0")
        os2 = stats2.get("ore_spent", "+0")
        ms1 = stats1.get("mana_spent", "+0")
        ms2 = stats2.get("mana_spent", "+0")
        
        # Parse with absolute values
        gs1_val = abs(int(gs1.replace(",", "").replace("+", ""))) if gs1.replace(",", "").replace("+", "").lstrip("-").isdigit() else 0
        gs2_val = abs(int(gs2.replace(",", "").replace("+", ""))) if gs2.replace(",", "").replace("+", "").lstrip("-").isdigit() else 0
        ws1_val = abs(int(ws1.replace(",", "").replace("+", ""))) if ws1.replace(",", "").replace("+", "").lstrip("-").isdigit() else 0
        ws2_val = abs(int(ws2.replace(",", "").replace("+", ""))) if ws2.replace(",", "").replace("+", "").lstrip("-").isdigit() else 0
        os1_val = abs(int(os1.replace(",", "").replace("+", ""))) if os1.replace(",", "").replace("+", "").lstrip("-").isdigit() else 0
        os2_val = abs(int(os2.replace(",", "").replace("+", ""))) if os2.replace(",", "").replace("+", "").lstrip("-").isdigit() else 0
        ms1_val = abs(int(ms1.replace(",", "").replace("+", ""))) if ms1.replace(",", "").replace("+", "").lstrip("-").isdigit() else 0
        ms2_val = abs(int(ms2.replace(",", "").replace("+", ""))) if ms2.replace(",", "").replace("+", "").lstrip("-").isdigit() else 0
        
        output += f"  Gold: {name1} -{gs1_val:,} | {name2} -{gs2_val:,}\n"
        output += f"  Wood: {name1} -{ws1_val:,} | {name2} -{ws2_val:,}\n"
        output += f"  Ore: {name1} -{os1_val:,} | {name2} -{os2_val:,}\n"
        output += f"  Mana: {name1} -{ms1_val:,} | {name2} -{ms2_val:,}\n"
        output += f"\n"
        
        # RSS Gathered
        output += f"📦 RSS Gathered\n"
        gg1 = stats1.get("gold_gathered", "+0")
        gg2 = stats2.get("gold_gathered", "+0")
        wg1 = stats1.get("wood_gathered", "+0")
        wg2 = stats2.get("wood_gathered", "+0")
        og1 = stats1.get("ore_gathered", "+0")
        og2 = stats2.get("ore_gathered", "+0")
        mg1_g = stats1.get("mana_gathered", "+0")
        mg2_g = stats2.get("mana_gathered", "+0")
        
        output += f"  Gold: {name1} {gg1} | {name2} {gg2}\n"
        output += f"  Wood: {name1} {wg1} | {name2} {wg2}\n"
        output += f"  Ore: {name1} {og1} | {name2} {og2}\n"
        output += f"  Mana: {name1} {mg1_g} | {name2} {mg2_g}\n"
        
        output += f"```"
        
        await msg.edit(content=output)
    except Exception as e:
        log_info(f"[COMPARE ERROR] {e}")
        import traceback
        traceback.print_exc()
        await msg.edit(content=f"❌ Error: {str(e)}")


# ============================================================
# QUICK COMMANDS
# ============================================================

@bot.command(name="q")
async def quick_stats(ctx, user_input: str = None):
    """Quick one-liner stats. Usage: !q (your stats) or !q truvix"""
    season = db_get_current_season()
    if not season:
        return await ctx.send("❌ No season active.")
    
    season_id, season_name, start_date, created_at = season
    today = date.today().isoformat()
    
    # Get account ID
    account_id = await get_account_id_from_input(ctx, user_input)
    if not account_id:
        return await ctx.send("❌ Could not find account ID.")
    
    try:
        # FIRST: Try database for today
        stats = db_get_season_progress(season_id, account_id, today)
        
        # If not today, try yesterday
        if not stats:
            yesterday = (date.today() - timedelta(days=1)).isoformat()
            stats = db_get_season_progress(season_id, account_id, yesterday)
        
        # FALLBACK: Check cache
        if not stats:
            stats = get_cached_stats(account_id, start_date, today)
        
        # FALLBACK 2: Fetch from API (for unlisted people)
        if not stats:
            stats, _ = await fetch_stats_with_fallback(account_id, start_date, today)
        
        if not stats or stats.get("lord_name") == "Unknown":
            return await ctx.send("❌ Failed to fetch stats.")
        
        # Get power and rankings
        power = await fetch_highest_power(account_id)
        power_rank = await get_rankings_for_stat(ctx, "power_gain", start_date, today)
        merits_rank = await get_rankings_for_stat(ctx, "merits", start_date, today)
        kills_rank = await get_rankings_for_stat(ctx, "kills_gain", start_date, today)
        
        # Calculate merit to power ratio using highest power and merits
        if stats.get("merits") and power:
            try:
                merits_str = stats["merits"].replace("+", "").replace(",", "")
                merits_val = int(merits_str) if merits_str.isdigit() else 0
                
                # Calculate ratio: (Merits / Highest Power) × 100
                if power > 0:
                    ratio = (merits_val / power) * 100
                    stats["merits_pct"] = f"{ratio:.1f}%"
                else:
                    stats["merits_pct"] = "0%"
            except Exception as e:
                stats["merits_pct"] = "0%"
        else:
            stats["merits_pct"] = "0%"
        
        # Extract data with absolute values
        lord_name = stats.get("lord_name", "Unknown")
        merits = stats.get("merits", "+0")
        merits_pct = stats.get("merits_pct", "0%")
        kills = stats.get("kills_gain", "+0")
        deaths = stats.get("deads_gain", "+0")
        healed = stats.get("healed_gain", "+0")
        
        # Parse mana_spent with absolute value
        mana_spent_str = stats.get("mana_spent", "+0")
        mana_clean = mana_spent_str.replace(",", "").replace("+", "")
        mana_spent_val = abs(int(mana_clean)) if mana_clean.lstrip("-").isdigit() else 0
        mana_spent = f"-{mana_spent_val:,}"
        
        # Get ranking positions as strings
        power_rank_str = f"(#{power_rank[account_id][0]})" if account_id in power_rank else ""
        merits_rank_str = f"(#{merits_rank[account_id][0]})" if account_id in merits_rank else ""
        kills_rank_str = f"(#{kills_rank[account_id][0]})" if account_id in kills_rank else ""
        
        # Format one-liner
        output = f"**{lord_name}** | "
        
        if power:
            power_gain_str = stats.get("power_gain", "+0")
            output += f"⚡ {power:,} {power_gain_str} {power_rank_str} | "
        
        if merits and merits != "+0":
            output += f"🏅 {merits} ({merits_pct}) {merits_rank_str} | "
        
        output += f"⚔️ {kills} {kills_rank_str} | "
        output += f"💀 {deaths} | "
        output += f"❤️ {healed} | "
        output += f"💧 {mana_spent}"
        
        await ctx.send(output)
    except Exception as e:
        log_error(f"Quick stats error: {e}")
        await ctx.send("❌ Error fetching stats.")


@bot.command(name="active")
async def active_members(ctx):
    """Show who's active (last 24h) vs inactive with days count - uses cached data"""
    season = db_get_current_season()
    if not season:
        return await ctx.send("❌ No season active.")
    
    season_id, season_name, start_date, created_at = season
    today = date.today().isoformat()
    
    try:
        guild = ctx.guild
        lords = get_all_lords_from_guild(guild)
        
        # Create set to track which account IDs we've already checked (prevent duplicates)
        checked_accounts = set()
        accounts_to_check = []
        
        # Add lords from guild roles
        for lord in lords:
            account_id = lord["account_id"]
            if account_id not in checked_accounts:
                checked_accounts.add(account_id)
                accounts_to_check.append(account_id)
        
        # Add mapped accounts (like Havi who's not in server)
        for discord_id, account_id in DISCORD_TO_ACCOUNT_ID.items():
            if account_id not in checked_accounts:
                checked_accounts.add(account_id)
                accounts_to_check.append(account_id)
        
        if not accounts_to_check:
            return await ctx.send("❌ No members found.")
        
        active = []
        inactive = []
        
        for account_id in accounts_to_check:
            try:
                # PRIORITY 1: Try cache first (populated by forcefetch)
                stats_today = get_cached_stats(account_id, start_date, today)
                actual_today_date = today
                
                # PRIORITY 2: Try database if cache miss
                if not stats_today:
                    stats_today = db_get_season_progress(season_id, account_id, today)
                    if stats_today:
                        actual_today_date = stats_today.get("data_date", today)
                
                # PRIORITY 3: Try yesterday in cache
                if not stats_today:
                    yesterday = (date.today() - timedelta(days=1)).isoformat()
                    stats_today = get_cached_stats(account_id, start_date, yesterday)
                    if stats_today:
                        actual_today_date = yesterday
                
                # PRIORITY 4: Try yesterday in database
                if not stats_today:
                    yesterday = (date.today() - timedelta(days=1)).isoformat()
                    stats_today = db_get_season_progress(season_id, account_id, yesterday)
                    if stats_today:
                        actual_today_date = stats_today.get("data_date", yesterday)
                
                # If STILL no data, skip this account
                if not stats_today or not isinstance(stats_today, dict):
                    log_info(f"[ACTIVE] No stats for {account_id}, skipping")
                    continue
                
                # Now safe to get lord_name
                lord_name = stats_today.get("lord_name", account_id)
                if stats_today.get("data_date"):
                    actual_today_date = stats_today.get("data_date")
                
                log_info(f"[ACTIVE] Using {lord_name} data from {actual_today_date}")
                
                # Get stats from day before the actual date we got
                day_before = (datetime.strptime(actual_today_date, "%Y-%m-%d").date() - timedelta(days=1)).isoformat()
                
                # Try cache first for yesterday
                stats_yesterday = get_cached_stats(account_id, start_date, day_before)
                if stats_yesterday:
                    log_info(f"[ACTIVE] Found yesterday in CACHE for {account_id}: {day_before}")
                
                # Then try database for specific day
                if not stats_yesterday:
                    stats_yesterday = db_get_season_progress(season_id, account_id, day_before)
                    if stats_yesterday:
                        log_info(f"[ACTIVE] Found yesterday in DB for {account_id}: {day_before}")
                
                # If still not found, try to find LATEST data before today (handles skipped dates)
                if not stats_yesterday:
                    log_info(f"[ACTIVE] No yesterday data found, searching for earlier data before {actual_today_date}")
                    try:
                        conn = sqlite3.connect(DB_PROGRESS)
                        c = conn.cursor()
                        c.execute("""
                            SELECT power_gain, merits, kills_gain, deads_gain, healed_gain,
                                   t5_gain, t4_gain, t3_gain, t2_gain, t1_gain,
                                   gold_spent, wood_spent, ore_spent, mana_spent,
                                   gold_gathered, wood_gathered, ore_gathered, mana_gathered, lord_name, data_date
                            FROM season_progress
                            WHERE season_id=? AND account_id=? AND data_date < ?
                            ORDER BY data_date DESC
                            LIMIT 1
                        """, (season_id, account_id, actual_today_date))
                        row = c.fetchone()
                        conn.close()
                        
                        if row:
                            stats_yesterday = {
                                "power_gain": row[0],
                                "merits": row[1],
                                "kills_gain": row[2],
                                "deads_gain": row[3],
                                "healed_gain": row[4],
                                "t5_gain": row[5],
                                "t4_gain": row[6],
                                "t3_gain": row[7],
                                "t2_gain": row[8],
                                "t1_gain": row[9],
                                "gold_spent": row[10],
                                "wood_spent": row[11],
                                "ore_spent": row[12],
                                "mana_spent": row[13],
                                "gold_gathered": row[14],
                                "wood_gathered": row[15],
                                "ore_gathered": row[16],
                                "mana_gathered": row[17],
                                "lord_name": row[18],
                                "data_date": row[19]
                            }
                            log_info(f"[ACTIVE] Found earlier data for {account_id}: {row[19]}")
                        else:
                            log_info(f"[ACTIVE] ❌ No earlier data found for {account_id} before {actual_today_date}")
                    except Exception as e:
                        log_error(f"[ACTIVE] Error querying earlier dates: {e}")
                
                if not stats_yesterday or not isinstance(stats_yesterday, dict):
                    log_info(f"[ACTIVE] Marking {lord_name} as INACTIVE (no comparison data)")
                    inactive.append({"name": lord_name, "days": "?"})
                    continue
                
                # Extract stats safely
                def safe_parse(stat_str):
                    """Safely parse stat strings like '+12,345' to int"""
                    if not stat_str:
                        return 0
                    return int(str(stat_str).replace("+", "").replace(",", "") or 0)
                
                power_today = safe_parse(stats_today.get("power_gain", "0"))
                power_yesterday = safe_parse(stats_yesterday.get("power_gain", "0"))
                
                merits_today = safe_parse(stats_today.get("merits", "0"))
                merits_yesterday = safe_parse(stats_yesterday.get("merits", "0"))
                
                mana_today = safe_parse(stats_today.get("mana_gathered", "0"))
                mana_yesterday = safe_parse(stats_yesterday.get("mana_gathered", "0"))
                
                power_gain_24h = power_today - power_yesterday
                merits_gain_24h = merits_today - merits_yesterday
                mana_gain_24h = mana_today - mana_yesterday
                
                # Active if any gain
                if power_gain_24h > 0 or merits_gain_24h > 0 or mana_gain_24h > 0:
                    active.append({
                        "name": lord_name,
                        "power": power_gain_24h,
                        "merits": merits_gain_24h,
                        "mana": mana_gain_24h
                    })
                else:
                    inactive.append({"name": lord_name, "days": "?"})
            except Exception as e:
                log_error(f"Error checking activity for {account_id}: {e}")
                import traceback
                log_error(traceback.format_exc())
                continue
        
        # Sort active by power gain
        active.sort(key=lambda x: x["power"], reverse=True)
        
        # Create embed
        embed = discord.Embed(
            title="📊 Activity Report",
            description="Last 24 Hours",
            color=0x2ecc71 if active else 0x95a5a6
        )
        
        if active:
            # Build two columns: Member name | All stats combined
            member_col = ""
            stats_col = ""
            
            for member in active:
                member_col += f"{member['name']}\n"
                stats_col += f"⚔️ +{member['power']:,} | 🏆 +{member['merits']:,} | 💧 +{member['mana']:,}\n"
            
            embed.add_field(name="✅ Member", value=member_col or "None", inline=True)
            embed.add_field(name="📈 Stats (Power | Merits | Mana)", value=stats_col or "—", inline=True)
        else:
            embed.add_field(name="✅ Active Members", value="None", inline=False)
        
        if inactive:
            inactive_list = "\n".join([f"• {m['name']}" for m in inactive])
            embed.add_field(name=f"⏸️ Inactive ({len(inactive)})", value=inactive_list, inline=False)
        
        await ctx.send(embed=embed)
    except Exception as e:
        log_error(f"Active members error: {e}")
        await ctx.send("❌ Error checking activity.")


# ============================================================
# WEEKLY EVENTS
# ============================================================

weekly_events = ["Melee Wheel", "Melee Forge", "Range Wheel", "Range Forge"]
event_emojis = {
    "Range Forge": "🏹",
    "Melee Wheel": "⚔️",
    "Melee Forge": "🔨",
    "Range Wheel": "🎯"
}

start_date = date(2025, 9, 16)

@bot.tree.command(name="weeklyevent", description="Show current and upcoming weekly Abyss events")
async def weeklyevent(inter):
    today = date.today()
    sunday = start_date - timedelta(days=start_date.weekday() + 1)
    weeks = (today - sunday).days // 7

    this_tue = sunday + timedelta(weeks=weeks, days=2)
    now = datetime.utcnow()

    event_start = datetime.combine(this_tue, time(0, 0))
    event_end = event_start + timedelta(days=3)

    if now >= event_end:
        weeks += 1
        this_tue = sunday + timedelta(weeks=weeks, days=2)
        event_start = datetime.combine(this_tue, time(0, 0))
        event_end = event_start + timedelta(days=3)

    msg = "📅 **Weekly Abyss Events**\n\n"
    nums = ["1️⃣", "2️⃣", "3️⃣", "4️⃣"]

    for i in range(4):
        idx = (weeks + i) % 4
        nm = weekly_events[idx]
        emoji = event_emojis.get(nm, "📌")
        ev_date = sunday + timedelta(weeks=weeks + i, days=2)
        base_dt = datetime.combine(ev_date, time(0, 0))

        if i == 0 and event_start <= now < event_end:
            left = event_end - now
            st = f"🟢 LIVE NOW ({left.seconds // 3600}h {(left.seconds // 60) % 60}m left)"
        else:
            left = base_dt - now
            st = f"⏳ {left.days}d {left.seconds // 3600}h {(left.seconds // 60) % 60}m"

        msg += f"{nums[i]} {emoji} **{nm}** — <t:{int(base_dt.timestamp())}:F>\n{st}\n\n"

    await inter.response.send_message(msg)

# ============================================================
# KVK / CUSTOM EVENTS DISPLAY
# ============================================================
@bot.tree.command(name="checkdb", description="check DB")
async def dbdump(interaction: discord.Interaction):
    if interaction.user.id != OWNER_ID:
        return await interaction.response.send_message(
            "Owner only.", ephemeral=True
        )

    rows = db_get_events()
    if not rows:
        return await interaction.response.send_message(
            "DB is empty.", ephemeral=True
        )

    msg = "\n".join(str(r) for r in rows)
    await interaction.response.send_message(
        f"```{msg}```", ephemeral=True
    )


@bot.tree.command(name="kvkevent", description="Show upcoming custom scheduled events")
async def kvkevent(inter):
    now = datetime.utcnow()
    rows = db_get_events()

    upcoming = [r for r in rows if datetime.fromisoformat(r[2]) > now]
    if not upcoming:
        return await inter.response.send_message("📭 No upcoming events.", ephemeral=True)

    upcoming = upcoming[:4]
    nums = ["1️⃣", "2️⃣", "3️⃣", "4️⃣"]
    msg = "📅 **Upcoming Events**\n\n"

    for i, row in enumerate(upcoming):
        event_id, name, dt, rem = row
        dt_obj = datetime.fromisoformat(dt)
        left = dt_obj - now
        msg += (
            f"{nums[i]} **{name}** — <t:{int(dt_obj.timestamp())}:F>\n"
            f"Starts in {left.days}d {left.seconds // 3600}h {(left.seconds // 60) % 60}m\n\n"
        )

    await inter.response.send_message(msg, ephemeral=True)

# ============================================================
# ADD EVENT MODAL
# ============================================================
class AddEventModal(Modal, title="➕ Add Event"):
    name = TextInput(
        label="Event Name",
        placeholder="e.g. Pass 1 Opens"
    )
    dt_input = TextInput(
        label="Datetime (UTC)",
        placeholder="DD-MM-YYYY HH:MM | 1d 2h | 14utc"
    )
    reminder = TextInput(
        label="Reminder (minutes or 'no')",
        placeholder="15"
    )

    async def on_submit(self, interaction: discord.Interaction):
        await interaction.response.defer(ephemeral=True)

        try:
            dt = parse_datetime(self.dt_input.value)

            # Round event time to minute precision
            dt = dt.replace(second=0, microsecond=0)

            # Validate event is in the future
            if dt <= datetime.utcnow():
                return await interaction.followup.send(
                    "❌ Event must be in the future!",
                    ephemeral=True
                )

            rem = (
                0
                if self.reminder.value.lower() == "no"
                else int(self.reminder.value)
            )

            db_add_event(
                self.name.value.strip(),
                dt.isoformat(),
                rem
            )

            await interaction.followup.send(
                f"✅ **Event Added**\n"
                f"**{self.name.value}**\n"
                f"<t:{int(dt.timestamp())}:F>",
                ephemeral=True
            )

        except Exception as e:
            log_info("[AddEvent Error]", e)
            await interaction.followup.send(
                "❌ Failed to add event.",
                ephemeral=True
            )

@bot.tree.command(name="addevent", description="Add a custom event")
async def addevent(inter: discord.Interaction):
    if inter.user.id != OWNER_ID:
        return await inter.response.send_message("❌ Owner only.", ephemeral=True)

    await inter.response.send_modal(AddEventModal())


# ============================================================
# ADD / EDIT / REMOVE EVENT COMMANDS
# ============================================================

@bot.tree.command(name="testdm", description="Owner-only: test Abyss role DMs")
async def testdm(interaction: discord.Interaction):
    if interaction.user.id != OWNER_ID:
        return await interaction.response.send_message(
            "❌ Owner only.", ephemeral=True
        )

    embed = discord.Embed(
        title="🧪 Abyss DM Test",
        description="If you received this, Abyss DMs are working ✅",
        color=0x2ECC71
    )

    await interaction.response.send_message(
        "📨 Sending test DMs to Abyss role members…",
        ephemeral=True
    )

    await dm_abyss_role(interaction.guild, embed)



@bot.tree.command(name="editevent", description="Edit an existing event")
async def editevent(inter: discord.Interaction):
    if inter.user.id != OWNER_ID:
        return await inter.response.send_message("❌ Owner only.", ephemeral=True)

    events = db_get_events()
    if not events:
        return await inter.response.send_message("❌ No events available.", ephemeral=True)

    options = [
        discord.SelectOption(
            label=f"{name} ({dt})",
            value=str(event_id)
        )
        for event_id, name, dt, rem in events
    ]

    select = Select(
        placeholder="Select an event to edit",
        options=options
    )

    async def select_callback(i: discord.Interaction):
        event_id = int(select.values[0])
        event = next(e for e in events if e[0] == event_id)

        _, old_name, old_dt, old_rem = event

        class EditEventModal(Modal, title="Edit Event"):
            name = TextInput(label="Event Name", default=old_name)
            datetime = TextInput(
                label="Date & Time (YYYY-MM-DD HH:MM)",
                default=old_dt.replace("T", " ")[:16]
            )
            reminder = TextInput(
                label="Reminder (hours before)",
                default=str(old_rem)
            )

            async def on_submit(self, modal_inter: discord.Interaction):
                try:
                    dt_str = self.datetime.value.strip()

                    # Parse and round to minute precision
                    dt_obj = datetime.strptime(dt_str, "%Y-%m-%d %H:%M")
                    dt_obj = dt_obj.replace(second=0, microsecond=0)
                    dt_str = dt_obj.isoformat(timespec="minutes")

                    rem = int(self.reminder.value.strip())

                    db_update_event(
                        event_id,
                        name=self.name.value.strip(),
                        dt=dt_str,
                        reminder=rem
                    )

                    await modal_inter.response.send_message(
                        "✅ Event updated successfully.",
                        ephemeral=True
                    )

                except ValueError:
                    await modal_inter.response.send_message(
                        "❌ Invalid date or reminder format.\nUse `YYYY-MM-DD HH:MM`.",
                        ephemeral=True
                    )

                except Exception as e:
                    log_info("[EditEvent Error]", e)
                    await modal_inter.response.send_message(
                        "❌ Failed to update event.",
                        ephemeral=True
                    )


        await i.response.send_modal(EditEventModal())

    select.callback = select_callback
    view = View()
    view.add_item(select)

    await inter.response.send_message(
        "Select an event to edit:",
        view=view,
        ephemeral=True
    )


@bot.tree.command(name="addrole", description="Get the Abyss role")
async def addrole(interaction: discord.Interaction):
    role = interaction.guild.get_role(ABYSS_ROLE_ID)
    if not role:
        return await interaction.response.send_message(
            "❌ Abyss role not found.", ephemeral=True
        )

    if role in interaction.user.roles:
        return await interaction.response.send_message(
            "⚠️ You already have the Abyss role.", ephemeral=True
        )

    await interaction.user.add_roles(role)
    await interaction.response.send_message(
        "✅ Abyss role added! You will now receive Abyss reminders.",
        ephemeral=True
    )
@bot.tree.command(name="removerole", description="Remove the Abyss role")
async def removerole(interaction: discord.Interaction):
    role = interaction.guild.get_role(ABYSS_ROLE_ID)
    if not role:
        return await interaction.response.send_message(
            "❌ Abyss role not found.", ephemeral=True
        )

    if role not in interaction.user.roles:
        return await interaction.response.send_message(
            "⚠️ You don’t have the Abyss role.", ephemeral=True
        )

    await interaction.user.remove_roles(role)
    await interaction.response.send_message(
        "✅ Abyss role removed. You will no longer receive Abyss reminders.",
        ephemeral=True
    )

@bot.tree.command(name="removeevent", description="Remove an existing event")
async def removeevent(inter: discord.Interaction):
    if inter.user.id != OWNER_ID:
        return await inter.response.send_message("❌ Owner only.", ephemeral=True)

    events = db_get_events()
    if not events:
        return await inter.response.send_message("❌ No events available.", ephemeral=True)

    options = [
        discord.SelectOption(
            label=f"{name} ({dt})",
            value=str(event_id)
        )
        for event_id, name, dt, rem in events
    ]

    select = Select(
        placeholder="Select an event to remove",
        options=options
    )

    async def select_callback(i: discord.Interaction):
        event_id = int(select.values[0])
        db_delete_event(event_id)
        await i.response.send_message(
            "🗑️ Event removed successfully.",
            ephemeral=True
        )

    select.callback = select_callback
    view = View()
    view.add_item(select)

    await inter.response.send_message(
        "Select an event to remove:",
        view=view,
        ephemeral=True
    )


# ============================================================
# ABYSS CONFIG COMMAND
# ============================================================
class AbyssConfigView(View):
    def __init__(self, days, hours, rem, round2, rem_mins):
        super().__init__(timeout=300)
        self.days = days
        self.hours = hours
        self.rem = rem
        self.round2 = round2
        self.rem_mins = rem_mins

        self.day_sel = Select(
            placeholder="Select Abyss Days",
            min_values=1,
            max_values=7,
            options=[
                discord.SelectOption(
                    label=["Mon","Tue","Wed","Thu","Fri","Sat","Sun"][i],
                    value=str(i),
                    default=(i in days)
                )
                for i in range(7)
            ]
        )
        self.day_sel.callback = self.cb_days
        self.add_item(self.day_sel)

        self.hour_sel = Select(
            placeholder="Select Abyss Hours",
            min_values=1,
            max_values=6,
            options=[
                discord.SelectOption(
                    label=f"{h:02}:00 UTC",
                    value=str(h),
                    default=(h in hours)
                )
                for h in [0,4,8,12,16,20]
            ]
        )
        self.hour_sel.callback = self.cb_hours
        self.add_item(self.hour_sel)

        self.rem_sel = Select(
            placeholder="Reminder Hours",
            min_values=0,
            max_values=len(hours),
            options=[
                discord.SelectOption(
                    label=f"{h:02}:00 UTC",
                    value=str(h),
                    default=(h in rem)
                )
                for h in hours
            ]
        )
        self.rem_sel.callback = self.cb_rem
        self.add_item(self.rem_sel)

        # Reminder minutes buttons
        for mins in [5, 10, 15, 20]:
            btn = Button(
                label=f"{mins}m",
                style=discord.ButtonStyle.primary if rem_mins == mins else discord.ButtonStyle.secondary,
                custom_id=f"rem_mins_{mins}"
            )
            btn.callback = lambda inter, m=mins: self.set_reminder_mins(inter, m)
            self.add_item(btn)

        self.round_btn = Button(
            label=f"Round 2: {'ON' if round2 else 'OFF'}",
            style=discord.ButtonStyle.success if round2 else discord.ButtonStyle.danger
        )
        self.round_btn.callback = self.toggle_round2
        self.add_item(self.round_btn)

    async def cb_days(self, interaction):
        global ABYSS_DAYS
        self.days = [int(v) for v in self.day_sel.values]
        ABYSS_DAYS = self.days
        cfg["days"] = self.days
        save_json(ABYSS_CONFIG_FILE, cfg)
        await interaction.response.send_message("Days updated ✔", ephemeral=True)

    async def cb_hours(self, interaction):
        global ABYSS_HOURS
        self.hours = [int(v) for v in self.hour_sel.values]
        ABYSS_HOURS = self.hours
        cfg["hours"] = self.hours
        save_json(ABYSS_CONFIG_FILE, cfg)
        await interaction.response.send_message("Hours updated ✔", ephemeral=True)

    async def cb_rem(self, interaction):
        global REMINDER_HOURS
        self.rem = [int(v) for v in self.rem_sel.values]
        REMINDER_HOURS = self.rem
        cfg["reminder_hours"] = self.rem
        save_json(ABYSS_CONFIG_FILE, cfg)
        await interaction.response.send_message("Reminder hours updated ✔", ephemeral=True)

    async def set_reminder_mins(self, interaction, mins):
        global REMINDER_MINS
        self.rem_mins = mins
        REMINDER_MINS = mins
        cfg["reminder_mins"] = mins
        save_json(ABYSS_CONFIG_FILE, cfg)
        await interaction.response.send_message(f"Reminder set to {mins} minutes ✔", ephemeral=True)

    async def toggle_round2(self, interaction):
        global ROUND2_ENABLED
        self.round2 = not self.round2
        ROUND2_ENABLED = self.round2
        cfg["round2"] = self.round2
        save_json(ABYSS_CONFIG_FILE, cfg)
        await interaction.response.send_message(
            f"Round 2 {'enabled' if self.round2 else 'disabled'} ✔",
            ephemeral=True
        )



@bot.tree.command(name="abyssconfig", description="Configure Abyss days, hours, and reminders")
async def abyssconfig(inter):
    if inter.user.id != OWNER_ID:
        return await inter.response.send_message("❌ Owner only.", ephemeral=True)

    emb = discord.Embed(title="⚙️ Abyss Config", color=0x2ecc71)
    emb.add_field(name="Days", value=pretty_days(ABYSS_DAYS), inline=False)
    emb.add_field(name="Hours", value=pretty_hours(ABYSS_HOURS), inline=False)
    emb.add_field(name="Reminder Hours", value=pretty_hours(REMINDER_HOURS), inline=False)
    emb.add_field(name="Reminder Lead Time", value=f"{REMINDER_MINS} minutes", inline=False)
    emb.add_field(
        name="Round 2",
        value="Enabled" if ROUND2_ENABLED else "Disabled",
        inline=False
    )

    view = AbyssConfigView(
        ABYSS_DAYS,
        ABYSS_HOURS,
        REMINDER_HOURS,
        ROUND2_ENABLED,
        REMINDER_MINS
    )

    await inter.response.send_message(
        embed=emb,
        view=view,
        ephemeral=True
    )


# ============================================================
# REMINDER LOOPS
# ============================================================

@tasks.loop(minutes=1)
async def abyss_reminder_loop():
    try:
        tz = pytz.timezone(MY_TIMEZONE)
        now = datetime.now(tz)

        if now.weekday() not in ABYSS_DAYS:
            return

        # Calculate when to send: 15 minutes before actual event start
        # Events actually start at REMINDER_HOURS + 15 minutes
        # So send at REMINDER_HOURS + (15 - REMINDER_MINS)
        send_minute = 15 - REMINDER_MINS

        # Check if it's time to send Abyss reminder
        for event_hour in REMINDER_HOURS:
            
            if now.hour == event_hour and now.minute == send_minute:
                embed = discord.Embed(
                    title="🕒 Abyss Reminder",
                    description=f"Abyss starts in **{REMINDER_MINS} minutes**!",
                    color=0xE74C3C
                )
                ch = bot.get_channel(channel_id)
                if ch and ch.guild:
                    await dm_abyss_role(ch.guild, embed)
                else:
                    log_info("[ABYSS REMINDER] Warning: Channel not found or no guild")

            # Round 2 reminder (same timing)
            if ROUND2_ENABLED and now.hour == event_hour and now.minute == send_minute:
                embed = discord.Embed(
                    title="🕒 Abyss Reminder",
                    description=f"Round 2 starts in **{REMINDER_MINS} minutes**!",
                    color=0xF1C40F
                )
                ch = bot.get_channel(channel_id)
                if ch and ch.guild:
                    await dm_abyss_role(ch.guild, embed)
                else:
                    log_info("[ABYSS REMINDER] Warning: Channel not found or no guild")
    except Exception as e:
        log_info(f"[ABYSS REMINDER ERROR] {type(e).__name__}: {e}")

sent_custom = set()

@tasks.loop(minutes=1)
async def custom_event_loop():
    try:
        now = datetime.utcnow().replace(second=0, microsecond=0)
        ch = bot.get_channel(channel_id)
        if not ch:
            return

        rows = db_get_events()
        for ev in rows:
            event_id, name, dt, rem = ev
            dt_obj = datetime.fromisoformat(dt)

            # Delete events 1 hour after they pass
            if now >= dt_obj + timedelta(hours=1):
                db_delete_event(event_id)
                continue

            # Reminder logic (robust, restart-safe)
            if rem > 0:
                rtime = (dt_obj - timedelta(minutes=rem)).replace(
                    second=0,
                    microsecond=0
                )

                if now >= rtime and event_id not in sent_custom:
                    await ch.send(
                        f"<@&{EVENT_ANNOUNCEMENT_ROLE_ID}> ⏰ Reminder: **{name}** in {rem} minutes! "
                        f"<t:{int(dt_obj.timestamp())}:F>"
                    )
                    sent_custom.add(event_id)

        # Prevent memory growth
        if len(sent_custom) > 300:
            sent_custom.clear()
    except Exception as e:
        log_info(f"[CUSTOM EVENT LOOP ERROR] {type(e).__name__}: {e}")


# ============================================================
# GAIN COMMAND - DATE RANGE WITH AUTOCOMPLETE
# ============================================================



# ============================================================
# SERVER TOP X COMMAND
# ============================================================

# KvK matchup session state: channel_id -> session dict
kvk_sessions = {}

@bot.event
async def on_message(message):
    if message.author.bot:
        return

    import re
    match = re.match(r'^!servertop(\d+)$', message.content.strip(), re.IGNORECASE)
    if match:
        n = int(match.group(1))
        if n < 1 or n > 100:
            await message.channel.send("❌ Please use a number between 1 and 100. Example: `!servertop10`")
        else:
            await cmd_servertop(message, n)
        return

    match = re.match(r'^!servercheck(\d+)$', message.content.strip(), re.IGNORECASE)
    if match:
        server_num = match.group(1)
        await cmd_servercheck(message, server_num)
        return

    # Server update: waiting for Excel attachment
    if message.author.id in serverupdate_pending and message.attachments:
        if message.content.strip().lower() in ("cancel", "!cancel"):
            del serverupdate_pending[message.author.id]
            await message.channel.send("❌ Server update cancelled.")
            return
        del serverupdate_pending[message.author.id]
        await _process_serverupdate_attachment(await bot.get_context(message), message.attachments[0])
        return

    if message.author.id in serverupdate_pending and message.content.strip().lower() in ("cancel", "!cancel"):
        del serverupdate_pending[message.author.id]
        await message.channel.send("❌ Server update cancelled.")
        return

    # KvK matchup multi-step flow
    channel_id_key = message.channel.id
    if channel_id_key in kvk_sessions:
        handled = await kvk_session_handle(message, kvk_sessions[channel_id_key])
        if handled:
            return

    if message.content.strip().lower() == '!kvkmatchup':
        await kvk_start(message)
        return

    if message.content.strip().lower() == '!matchups':
        await cmd_matchups_list(message)
        return

    import re as _re2
    _m = _re2.match(r'^!matchup\s+(\d+)$', message.content.strip(), _re2.IGNORECASE)
    if _m:
        await cmd_matchup_view(message, int(_m.group(1)))
        return

    _md = _re2.match(r'^!delmatchup\s+(\d+)$', message.content.strip(), _re2.IGNORECASE)
    if _md:
        await cmd_matchup_delete(message, int(_md.group(1)))
        return

    await bot.process_commands(message)


async def cmd_servertop(message, n):
    """Fetch and display top N servers by highest power from callofstats.com"""
    import re

    url = "https://callofstats.com/server_alliance_rankings?kvk_pool_selected=Gen-2+Pool"

    # Fresh independent session — base URL is public, no auth needed
    try:
        timeout = aiohttp.ClientTimeout(total=15, connect=5, sock_read=10)
        async with aiohttp.ClientSession(timeout=timeout) as session:
            async with session.get(url, allow_redirects=True) as resp:
                if resp.status != 200:
                    await message.channel.send(f"❌ Failed to fetch rankings (HTTP {resp.status})")
                    return
                html = await resp.text()
    except asyncio.TimeoutError:
        await message.channel.send("❌ Request timed out. Try again.")
        return
    except Exception as e:
        await message.channel.send(f"❌ Error fetching data: {e}")
        return

    # Parse leaderboard table rows
    row_pattern = re.compile(
        r'<tr>\s*<td>(\d+)</td>\s*<td>(#\d+)</td>\s*<td[^>]*>([^<]+)</td>\s*<td>\s*(\d+)\s*</td>\s*<td>\s*([0-9,]+)\s*</td>',
        re.DOTALL
    )
    rows = row_pattern.findall(html)

    if not rows:
        await message.channel.send("❌ Could not parse the rankings. The site layout may have changed.")
        log_info("[SERVERTOP] Failed to parse table rows from HTML")
        return

    top_rows = rows[:n]

    # Pre-process for alignment
    processed = []
    for rank, server, alliance, lords, power in top_rows:
        alliance = alliance.strip()
        # strip HTML entities
        alliance = alliance.replace("&amp;", "&").replace("&#39;", "'")
        processed.append((rank, server, alliance, lords, power))

    # Calculate column widths
    max_server  = max(len(s) for _, s, _, _, _ in processed)
    max_name    = max(len(a) for _, _, a, _, _ in processed)
    max_lords   = max(len(l) for _, _, _, l, _ in processed)
    max_power   = max(len(p.strip()) for _, _, _, _, p in processed)

    col_server = max(max_server, 6)   # "Server"
    col_name   = max(max_name,   4)   # "Name"
    col_lords  = max(max_lords,  6)   # "Lords"
    col_power  = max(max_power,  5)   # "Power"

    def make_row(rank_str, server, name, lords, power):
        return f"{rank_str:<5} {server:<{col_server}}  {name:<{col_name}}  {lords:>{col_lords}}  {power:>{col_power}}"

    col_header = make_row("Rank", "Server", "Name", "Lords", "Power")
    separator  = "-" * len(col_header)

    table_lines = [col_header, separator]
    for rank, server, alliance, lords, power in processed:
        power = power.strip()
        table_lines.append(make_row(f"#{rank}", server, alliance, lords, power))

    title = f"🏆 Top {n} Servers by Highest Power (Gen-2 Pool)"
    table_body = "\n".join(table_lines)

    # Wrap in code block for monospace alignment
    full_msg = f"{title}\n```\n{table_body}\n```"

    if len(full_msg) <= 2000:
        await message.channel.send(full_msg)
    else:
        # Send title + header separately, then chunk the data rows
        await message.channel.send(f"{title}\n```\n{col_header}\n{separator}```")
        chunk_lines = []
        chunk_len = 0
        for line in table_lines[2:]:  # skip header+separator already sent
            if chunk_len + len(line) + 1 > 1800:
                await message.channel.send("```\n" + "\n".join(chunk_lines) + "\n```")
                chunk_lines = []
                chunk_len = 0
            chunk_lines.append(line)
            chunk_len += len(line) + 1
        if chunk_lines:
            await message.channel.send("```\n" + "\n".join(chunk_lines) + "\n```")

    log_info(f"[SERVERTOP] Displayed top {n} servers")


async def cmd_servercheck(message, server_num):
    """Find a specific server's rank in the rankings"""
    import re

    url = "https://callofstats.com/server_alliance_rankings?kvk_pool_selected=Gen-2+Pool"

    try:
        timeout = aiohttp.ClientTimeout(total=15, connect=5, sock_read=10)
        async with aiohttp.ClientSession(timeout=timeout) as session:
            async with session.get(url, allow_redirects=True) as resp:
                if resp.status != 200:
                    await message.channel.send(f"❌ Failed to fetch rankings (HTTP {resp.status})")
                    return
                html = await resp.text()
    except asyncio.TimeoutError:
        await message.channel.send("❌ Request timed out. Try again.")
        return
    except Exception as e:
        await message.channel.send(f"❌ Error fetching data: {e}")
        return

    row_pattern = re.compile(
        r'<tr>\s*<td>(\d+)</td>\s*<td>(#\d+)</td>\s*<td[^>]*>([^<]+)</td>\s*<td>\s*(\d+)\s*</td>\s*<td>\s*([0-9,]+)\s*</td>',
        re.DOTALL
    )
    rows = row_pattern.findall(html)

    if not rows:
        await message.channel.send("❌ Could not parse the rankings. The site layout may have changed.")
        log_info("[SERVERCHECK] Failed to parse table rows from HTML")
        return

    # Search for the server number
    target = f"#{server_num}"
    found = None
    for rank, server, alliance, lords, power in rows:
        if server == target:
            found = (rank, server, alliance.strip(), lords, power.strip())
            break

    if not found:
        await message.channel.send(f"❌ Server **#{server_num}** not found in the rankings. It may be below the minimum power threshold or not tracked yet.")
        return

    rank, server, alliance, lords, power = found
    alliance = alliance.replace("&amp;", "&").replace("&#39;", "'")

    msg = (
        f"🔍 **Server #{server_num} Rankings**\n"
        f"```"
        f"Rank     : #{rank}\n"
        f"Server   : {server}\n"
        f"Alliance : {alliance}\n"
        f"Lords    : {lords}\n"
        f"Power    : {power}"
        f"```"
    )
    await message.channel.send(msg)
    log_info(f"[SERVERCHECK] Found S#{server_num} at rank #{rank}")




# ============================================================
# KVK MATCHUP COMMAND
# ============================================================

def kvk_fmt(n):
    try:
        n = int(n)
    except:
        return str(n)
    if n >= 1_000_000_000:
        return f"{n/1_000_000_000:.2f}B"
    if n >= 1_000_000:
        return f"{n/1_000_000:.2f}M"
    if n >= 1_000:
        return f"{n/1_000:.1f}K"
    return str(n)


async def kvk_fetch_zones(zone_map, num_zones):
    import re
    params = f"starting_zones={num_zones}&min_power=20000000"
    for z in range(1, num_zones + 1):
        for s in zone_map.get(z, []):
            params += f"&zone{z}={s}"
    params += "&show_results=1"
    url = f"https://callofstats.com/kvk_matchmaking?{params}"
    log_info(f"[KVK] Fetching: {url}")
    try:
        session = await get_callofstats_session()
        if not session:
            return None, "No authenticated session."
        async with session.get(url, allow_redirects=True) as resp:
            if resp.status != 200:
                return None, f"HTTP {resp.status}"
            html = await resp.text()
        if "<title>Login" in html:
            return None, "Got login redirect."
    except asyncio.TimeoutError:
        return None, "Request timed out."
    except Exception as e:
        return None, str(e)

    strip_pattern = re.compile(r'<div class="zone-strip"(.*?)</div>\s*</div>\s*</div>', re.DOTALL)
    strips = strip_pattern.findall(html)
    if not strips:
        return None, "Could not parse zone strips."

    def da(block, attr):
        m = re.search(r'data-' + attr + r'="([^"]*)"', block)
        return m.group(1) if m else "0"

    def lv(block, label):
        m = re.search(r'<span class="zs-left-label">' + re.escape(label) + r'</span>\s*<div class="zs-left-value">([^<]*)</div>', block)
        return m.group(1).strip() if m else ""

    zones = []
    for i, strip in enumerate(strips[:num_zones]):
        z = {
            "zone_num": i + 1,
            "servers":  lv(strip, "Servers"),
            "acronyms": lv(strip, "Acronyms"),
            "date":     lv(strip, "Dates"),
            "power":    int(da(strip, "power")),
            "merits":   int(da(strip, "merits")),
            "hero_power": int(da(strip, "hero_power")),
            "kills":    int(da(strip, "kills")),
            "deads":    int(da(strip, "deads")),
            "healed":   int(da(strip, "healed")),
            "mana":     int(da(strip, "mana")),
            "leg_heroes": int(da(strip, "leg_heroes_awakened")),
            "max_pets": int(da(strip, "max_pets")),
            "exemplar": int(da(strip, "exemplar_unlocked")),
            "players":  int(da(strip, "players")),
            "t3_lords": int(da(strip, "tier_3_lords")),
            "t4_lords": int(da(strip, "tier_4_lords")),
            "t5_lords": int(da(strip, "tier_5_lords")),
            "full_t5":  int(da(strip, "full_t5_lords")),
            "b60_80":   int(da(strip, "60_80")),
            "b80_100":  int(da(strip, "80_100")),
            "b100_125": int(da(strip, "100_125")),
            "b125_150": int(da(strip, "125_150")),
            "b150_200": int(da(strip, "150_200")),
            "b200_500": int(da(strip, "200_500")),
            "b500_1b":  int(da(strip, "500_1b")),
            "b1b_plus": int(da(strip, "1b_plus")),
        }
        z["mp_ratio"]  = (z["merits"] / z["power"] * 100) if z["power"] > 0 else 0
        z["avg_power"] = (z["power"] / z["players"]) if z["players"] > 0 else 0
        zones.append(z)
    return zones, None


def kvk_zone_block(z):
    lines = [
        f"Zone {z['zone_num']} \u2014 S{z['servers']} ({z['acronyms']}) \u2014 {z['date']}",
        f"Power          {kvk_fmt(z['power'])}  (Avg:{kvk_fmt(z['avg_power'])} | Hero:{kvk_fmt(z['hero_power'])})",
        f"Merits         {kvk_fmt(z['merits'])}  (M/P:{z['mp_ratio']:.1f}%)",
        f"Kills          {kvk_fmt(z['kills'])}",
        f"Deads          {kvk_fmt(z['deads'])}",
        f"Healed         {kvk_fmt(z['healed'])}",
        f"Mana Spent     {kvk_fmt(z['mana'])}",
        f"Players        {z['players']}  (LegHero:{z['leg_heroes']} | MaxPets:{z['max_pets']} | Exemplar:{z['exemplar']})",
        f"Tier Lords     T3:{z['t3_lords']}  T4:{z['t4_lords']}  T5:{z['t5_lords']}  FullT5:{z['full_t5']}",
        f"Brackets       60-80M:{z['b60_80']}  80-100M:{z['b80_100']}  100-125M:{z['b100_125']}",
        f"               125-150M:{z['b125_150']}  150-200M:{z['b150_200']}  200-500M:{z['b200_500']}",
        f"               500M-1B:{z['b500_1b']}  1B+:{z['b1b_plus']}",
    ]
    return "```\n" + "\n".join(lines) + "\n```"


def kvk_team_block(label, zone_indices, zones):
    selected = [zones[i-1] for i in zone_indices if 0 < i <= len(zones)]
    if not selected:
        return f"**{label}**: No zones selected."
    keys = ["power","merits","hero_power","kills","deads","healed","mana",
            "leg_heroes","max_pets","exemplar","players",
            "t3_lords","t4_lords","t5_lords","full_t5",
            "b60_80","b80_100","b100_125","b125_150","b150_200","b200_500","b500_1b","b1b_plus"]
    t = {k: sum(z[k] for z in selected) for k in keys}
    t["mp_ratio"]  = (t["merits"] / t["power"] * 100) if t["power"] > 0 else 0
    t["avg_power"] = (t["power"] / t["players"]) if t["players"] > 0 else 0
    servers = "/".join(z["servers"] for z in selected)
    lines = [
        f"{label} \u2014 S{servers}",
        f"Power          {kvk_fmt(t['power'])}  (Avg:{kvk_fmt(t['avg_power'])} | Hero:{kvk_fmt(t['hero_power'])})",
        f"Merits         {kvk_fmt(t['merits'])}  (M/P:{t['mp_ratio']:.1f}%)",
        f"Kills          {kvk_fmt(t['kills'])}",
        f"Deads          {kvk_fmt(t['deads'])}",
        f"Healed         {kvk_fmt(t['healed'])}",
        f"Mana Spent     {kvk_fmt(t['mana'])}",
        f"Players        {t['players']}  (LegHero:{t['leg_heroes']} | MaxPets:{t['max_pets']} | Exemplar:{t['exemplar']})",
        f"Tier Lords     T3:{t['t3_lords']}  T4:{t['t4_lords']}  T5:{t['t5_lords']}  FullT5:{t['full_t5']}",
        f"Brackets       60-80M:{t['b60_80']}  80-100M:{t['b80_100']}  100-125M:{t['b100_125']}",
        f"               125-150M:{t['b125_150']}  150-200M:{t['b150_200']}  200-500M:{t['b200_500']}",
        f"               500M-1B:{t['b500_1b']}  1B+:{t['b1b_plus']}",
    ]
    return f"\u2694\ufe0f **{label}**\n" + "```\n" + "\n".join(lines) + "\n```"


async def kvk_start(message):
    session = {
        "step": "zones",
        "user_id": message.author.id,
        "num_zones": None,
        "zone_map": {},
        "current_zone": 1,
        "zones_data": None,
    }
    kvk_sessions[message.channel.id] = session
    await message.channel.send("⚔️ **KvK Matchup Setup**\nHow many zones? Reply `4` or `6`.\nType `abort` at any time to cancel.")


async def kvk_session_handle(message, session):
    if message.author.id != session["user_id"]:
        return False
    content = message.content.strip()
    if content.lower() in ("cancel", "!cancel", "abort", "!abort", "stop", "!stop"):
        del kvk_sessions[message.channel.id]
        await message.channel.send("❌ KvK matchup cancelled.")
        return True
    step = session["step"]

    if step == "zones":
        if content not in ("4", "6"):
            await message.channel.send("Please reply `4` or `6`.")
            return True
        session["num_zones"] = int(content)
        session["step"] = "collecting"
        session["current_zone"] = 1
        await message.channel.send(f"Zone 1 of {session['num_zones']}: Enter server number(s) for Zone 1 (comma-separated).\nExample: `698` or `698, 357` — type `abort` to cancel.")
        return True

    if step == "collecting":
        z = session["current_zone"]
        servers = [s.strip() for s in content.replace(",", " ").split() if s.strip().isdigit()]
        if not servers:
            await message.channel.send("Please enter valid server number(s), e.g. `698` or `698, 357`. Type `abort` to cancel.")
            return True
        session["zone_map"][z] = servers
        if z < session["num_zones"]:
            session["current_zone"] += 1
            nz = session["current_zone"]
            await message.channel.send(f"Zone {nz} of {session['num_zones']}: Enter server number(s) for Zone {nz}. (type `abort` to cancel)")
            return True
        session["step"] = "fetching"
        num_zones = session["num_zones"]
        zone_map = session["zone_map"]
        zone_display = "  ".join(f"Z{k}:S{','.join(v)}" for k, v in sorted(zone_map.items()))
        msg = await message.channel.send(f"📡 Fetching KvK data for {zone_display}...")
        zones_data, error = await kvk_fetch_zones(zone_map, num_zones)
        if error or not zones_data:
            del kvk_sessions[message.channel.id]
            await msg.edit(content=f"❌ Failed to fetch KvK data: {error or 'No zones parsed'}")
            return True
        session["zones_data"] = zones_data
        session["step"] = "teams"
        await msg.edit(content=f"✅ Fetched data for {len(zones_data)} zones:")
        for z_data in zones_data:
            await message.channel.send(kvk_zone_block(z_data))
        zone_nums = " ".join(str(i+1) for i in range(len(zones_data)))
        await message.channel.send(
            f"**Team Assignment** — You have zones: {zone_nums}\n"
            f"Which zones go to **Team 1**? Enter zone numbers separated by commas (e.g. `1,2`).\n"
            f"Remaining zones auto-assigned to Team 2. Type `skip` to skip."
        )
        return True

    if step == "teams":
        zones_data = session["zones_data"]
        num_zones = len(zones_data)
        if content.lower() == "skip":
            del kvk_sessions[message.channel.id]
            await message.channel.send("✅ Done!")
            return True
        team1_zones = [int(x.strip()) for x in content.replace(",", " ").split()
                       if x.strip().isdigit() and 1 <= int(x.strip()) <= num_zones]
        if not team1_zones:
            await message.channel.send(f"Please enter valid zone numbers (1–{num_zones}), e.g. `1,2`.")
            return True
        team2_zones = [i+1 for i in range(num_zones) if (i+1) not in team1_zones]

        await message.channel.send("⚔️ **KvK Team Comparison**")
        await message.channel.send(kvk_team_block("Team 1", team1_zones, zones_data))
        if team2_zones:
            await message.channel.send(kvk_team_block("Team 2", team2_zones, zones_data))

        t1 = sum(zones_data[i-1]["power"] for i in team1_zones if 0 < i <= len(zones_data))
        t2 = sum(zones_data[i-1]["power"] for i in team2_zones if 0 < i <= len(zones_data)) if team2_zones else 0
        if t2 > 0:
            diff = abs(t1 - t2)
            stronger = "Team 1" if t1 > t2 else "Team 2"
            await message.channel.send(f"📊 **Power Gap:** {kvk_fmt(diff)} — **{stronger}** has more power.")

        def build_totals(zone_indices):
            keys = ["power","merits","hero_power","kills","deads","healed","mana",
                    "leg_heroes","max_pets","exemplar","players",
                    "t3_lords","t4_lords","t5_lords","full_t5",
                    "b60_80","b80_100","b100_125","b125_150","b150_200","b200_500","b500_1b","b1b_plus"]
            sel = [zones_data[i-1] for i in zone_indices if 0 < i <= len(zones_data)]
            t = {k: sum(z[k] for z in sel) for k in keys}
            t["mp_ratio"]  = (t["merits"] / t["power"] * 100) if t["power"] > 0 else 0
            t["avg_power"] = (t["power"] / t["players"]) if t["players"] > 0 else 0
            t["servers"] = "/".join(z["servers"] for z in sel)
            return t

        session["team1_zones"]  = team1_zones
        session["team2_zones"]  = team2_zones
        session["team1_totals"] = build_totals(team1_zones)
        session["team2_totals"] = build_totals(team2_zones)
        session["step"] = "nickname"
        await message.channel.send("💾 Give this matchup a nickname to save it (e.g. `SoS9 Week 1`), or type `skip` to discard.")
        return True

    if step == "nickname":
        if content.lower() == "skip":
            del kvk_sessions[message.channel.id]
            await message.channel.send("✅ Matchup not saved.")
            return True
        nickname = content[:80]
        db_save_kvk_matchup(
            nickname,
            session["num_zones"],
            session["zones_data"],
            session["team1_zones"],
            session["team2_zones"],
            session["team1_totals"],
            session["team2_totals"],
        )
        del kvk_sessions[message.channel.id]
        await message.channel.send(f"✅ Matchup **\"{nickname}\"** saved! Use `!matchups` to view all saved matchups.")
        return True

    return False



async def cmd_matchups_list(message):
    """List all saved KvK matchups."""
    rows = db_get_kvk_matchups()
    if not rows:
        await message.channel.send("No KvK matchups saved yet. Use `!kvkmatchup` to create one.")
        return

    lines = ["📋 **Saved KvK Matchups**", "```"]
    lines.append(f"{'ID':<5} {'Nickname':<30} {'Date':<17} {'Zones':<6} {'Teams'}")
    lines.append("-" * 75)
    for row in rows:
        mid, nickname, created_at, num_zones, t1_zones_raw, t2_zones_raw, t1_raw, t2_raw = row
        import json
        t1_z = json.loads(t1_zones_raw)
        t2_z = json.loads(t2_zones_raw)
        t1 = json.loads(t1_raw)
        t2 = json.loads(t2_raw)
        teams_str = f"T1:Z{','.join(str(z) for z in t1_z)} vs T2:Z{','.join(str(z) for z in t2_z)}"
        lines.append(f"#{str(mid):<4} {nickname[:29]:<30} {created_at:<17} {num_zones:<6} {teams_str}")
    lines.append("```")
    lines.append("Use `!matchup <id>` to view full details. `!delmatchup <id>` to delete.")
    await message.channel.send("\n".join(lines))


async def cmd_matchup_view(message, matchup_id):
    """View full details of a saved KvK matchup."""
    m = db_get_kvk_matchup(matchup_id)
    if not m:
        await message.channel.send(f"❌ No matchup found with ID #{matchup_id}.")
        return

    await message.channel.send(f"⚔️ **#{m['id']} — {m['nickname']}** _(saved {m['created_at']})_")

    # Show each zone
    for z in m["zones_data"]:
        await message.channel.send(kvk_zone_block(z))

    # Show team totals
    def totals_block(label, t):
        lines = [
            f"{label} — S{t['servers']}",
            f"Power          {kvk_fmt(t['power'])}  (Avg:{kvk_fmt(t['avg_power'])} | Hero:{kvk_fmt(t['hero_power'])})",
            f"Merits         {kvk_fmt(t['merits'])}  (M/P:{t['mp_ratio']:.1f}%)",
            f"Kills          {kvk_fmt(t['kills'])}",
            f"Deads          {kvk_fmt(t['deads'])}",
            f"Healed         {kvk_fmt(t['healed'])}",
            f"Mana Spent     {kvk_fmt(t['mana'])}",
            f"Players        {t['players']}  (LegHero:{t['leg_heroes']} | MaxPets:{t['max_pets']} | Exemplar:{t['exemplar']})",
            f"Tier Lords     T3:{t['t3_lords']}  T4:{t['t4_lords']}  T5:{t['t5_lords']}  FullT5:{t['full_t5']}",
            f"Brackets       60-80M:{t['b60_80']}  80-100M:{t['b80_100']}  100-125M:{t['b100_125']}",
            f"               125-150M:{t['b125_150']}  150-200M:{t['b150_200']}  200-500M:{t['b200_500']}",
            f"               500M-1B:{t['b500_1b']}  1B+:{t['b1b_plus']}",
        ]
        return f"\u2694\ufe0f **{label}**\n" + "```\n" + "\n".join(lines) + "\n```"

    await message.channel.send(totals_block("Team 1", m["team1_totals"]))
    if m["team2_totals"]["players"] > 0:
        await message.channel.send(totals_block("Team 2", m["team2_totals"]))
        t1p = m["team1_totals"]["power"]
        t2p = m["team2_totals"]["power"]
        diff = abs(t1p - t2p)
        stronger = "Team 1" if t1p > t2p else "Team 2"
        await message.channel.send(f"📊 **Power Gap:** {kvk_fmt(diff)} — **{stronger}** has more power.")


async def cmd_matchup_delete(message, matchup_id):
    """Delete a saved KvK matchup (owner only)."""
    if message.author.id != OWNER_ID:
        await message.channel.send("❌ Only the owner can delete matchups.")
        return
    m = db_get_kvk_matchup(matchup_id)
    if not m:
        await message.channel.send(f"❌ No matchup found with ID #{matchup_id}.")
        return
    db_delete_kvk_matchup(matchup_id)
    await message.channel.send(f"🗑️ Matchup **#{matchup_id} — {m['nickname']}** deleted.")



# ============================================================
# SERVER LEADERBOARD COMMANDS (!stop*)
# ============================================================

def _parse_stat_str(raw):
    """Parse a stat value (int, string, or '+1,234' style) to a plain int."""
    if raw is None:
        return 0
    if isinstance(raw, (int, float)):
        return int(raw)
    try:
        return abs(int(str(raw).replace(",", "").replace("+", "").replace("-", "").strip()))
    except:
        return 0


def parse_server_excel(file_bytes):
    """
    Parse an uploaded server stats Excel file. Supports both the newer COS export
    format (T4/T5 Deaths, T4/T5 Severely Wounded, Enemy Merits, T4/T5 Healed split
    separately) and the older combined format (Deaths (T4/T5), Healing (T4/T5)) for
    backward compatibility with files uploaded before COS split these out.
    Returns (rows, error).
    """
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        ws = wb[wb.sheetnames[0]]
    except Exception as e:
        return None, f"Could not open Excel file: {e}"

    header_row = None
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        header_row = [str(h).strip().lower() if h else "" for h in row]
        break

    if not header_row:
        return None, "Excel file appears empty."

    def find_col(*names):
        for name in names:
            for i, h in enumerate(header_row):
                if name in h:
                    return i
        return None

    def find_col_exact(*names):
        for name in names:
            for i, h in enumerate(header_row):
                if h == name:
                    return i
        return None

    col_id       = find_col("character id", "lord id", "account id")
    col_name     = find_col("character name", "lord name", "name")
    col_power    = find_col("current power")
    col_highest  = find_col("historical highest power", "highest power")
    col_merits   = find_col("total merits")
    col_gather   = find_col("gathering")
    col_inf      = find_col("infantry")
    col_cav      = find_col("cavalry")
    col_mark     = find_col("marksman")
    col_mage     = find_col("magic", "mage")
    col_other    = find_col("other merits")

    # New split columns (exact match, since "deaths" alone would ambiguously match
    # both "T4 Deaths" and "T5 Deaths")
    col_t4_deaths   = find_col_exact("t4 deaths")
    col_t5_deaths   = find_col_exact("t5 deaths")
    col_t4_wounded  = find_col_exact("t4 severely wounded")
    col_t5_wounded  = find_col_exact("t5 severely wounded")
    col_enemy_merits = find_col("enemy merits")
    col_t4_healed   = find_col_exact("t4 healed")
    col_t5_healed   = find_col_exact("t5 healed")

    # Old combined-format columns — only used as a fallback when the new split
    # columns above aren't present, for backward compatibility with older uploads
    col_deaths_combined = find_col("deaths (t4/t5)", "deaths")
    col_heal_combined   = find_col("healing (t4/t5)", "healing")

    if col_id is None or col_name is None:
        return None, f"Could not find required columns. Headers found: {header_row}"

    has_split_format = col_t4_healed is not None or col_t5_healed is not None

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[col_id] is None:
            continue
        try:
            account_id = str(row[col_id]).strip()
            if not account_id or account_id.lower() == "none":
                continue

            t4_deaths = _parse_stat_str(row[col_t4_deaths]) if col_t4_deaths is not None else 0
            t5_deaths = _parse_stat_str(row[col_t5_deaths]) if col_t5_deaths is not None else 0
            t4_healed = _parse_stat_str(row[col_t4_healed]) if col_t4_healed is not None else 0
            t5_healed = _parse_stat_str(row[col_t5_healed]) if col_t5_healed is not None else 0

            if has_split_format:
                deaths_total = t4_deaths + t5_deaths
                healing_total = t4_healed + t5_healed
            else:
                deaths_total = _parse_stat_str(row[col_deaths_combined]) if col_deaths_combined is not None else 0
                healing_total = _parse_stat_str(row[col_heal_combined]) if col_heal_combined is not None else 0

            rows.append({
                "account_id": account_id,
                "lord_name": str(row[col_name]).strip() if col_name is not None and row[col_name] else account_id,
                "current_power": _parse_stat_str(row[col_power]) if col_power is not None else 0,
                "highest_power": _parse_stat_str(row[col_highest]) if col_highest is not None else 0,
                "deaths": deaths_total,
                "total_merits": _parse_stat_str(row[col_merits]) if col_merits is not None else 0,
                "gathering": _parse_stat_str(row[col_gather]) if col_gather is not None else 0,
                "infantry_merits": _parse_stat_str(row[col_inf]) if col_inf is not None else 0,
                "cavalry_merits": _parse_stat_str(row[col_cav]) if col_cav is not None else 0,
                "marksman_merits": _parse_stat_str(row[col_mark]) if col_mark is not None else 0,
                "mage_merits": _parse_stat_str(row[col_mage]) if col_mage is not None else 0,
                "other_merits": _parse_stat_str(row[col_other]) if col_other is not None else 0,
                "healing": healing_total,
                "t4_deaths": t4_deaths,
                "t5_deaths": t5_deaths,
                "t4_severely_wounded": _parse_stat_str(row[col_t4_wounded]) if col_t4_wounded is not None else 0,
                "t5_severely_wounded": _parse_stat_str(row[col_t5_wounded]) if col_t5_wounded is not None else 0,
                "enemy_merits": _parse_stat_str(row[col_enemy_merits]) if col_enemy_merits is not None else 0,
                "t4_healed": t4_healed,
                "t5_healed": t5_healed,
            })
        except Exception as e:
            log_info(f"[SERVERUPDATE] Skipping row due to error: {e}")
            continue

    if not rows:
        return None, "No valid lord rows found in the Excel file."

    return rows, None


def parse_server_filename(filename):
    """
    Try to extract server_num, start_date, end_date from filename like
    '698_2026-07-02_2026-07-13.xlsx'. Returns (server_num, start_date, end_date) or (None, None, None).
    """
    import re
    m = re.match(r'(\d+)_(\d{4}-\d{2}-\d{2})_(\d{4}-\d{2}-\d{2})', filename)
    if m:
        return int(m.group(1)), m.group(2), m.group(3)
    # Try just server number at start
    m2 = re.match(r'(\d+)', filename)
    if m2:
        return int(m2.group(1)), None, None
    return None, None, None


async def _server_leaderboard(ctx, server_num, stat_field, emoji, label, top_n=25):
    """Generic server leaderboard from the uploaded Excel data."""
    picked = db_get_server_pick()
    if not picked:
        return await ctx.send("❌ No server picked. Use `!serverupdate` and upload an Excel file first.")
    if server_num and server_num != picked:
        return await ctx.send(f"❌ S#{server_num} is not the picked server (current: S#{picked}).")

    lords = db_get_server_lord_stats(picked)
    if not lords:
        return await ctx.send(f"❌ No data for S#{picked}. Use `!serverupdate` to upload the Excel file.")

    scored = [{"name": l["lord_name"], "val": l.get(stat_field, 0)} for l in lords]
    scored.sort(key=lambda x: x["val"], reverse=True)
    top = scored[:top_n]

    if not any(x["val"] > 0 for x in top):
        return await ctx.send(f"❌ No {label} data found for S#{picked}.")

    date_range = ""
    if lords[0].get("start_date") and lords[0].get("end_date"):
        date_range = f" ({lords[0]['start_date']} → {lords[0]['end_date']})"

    medals = ["🥇", "🥈", "🥉"]
    lines = [f"```{emoji} Top {top_n} {label} — S#{picked}{date_range}"]
    for i, lord in enumerate(top):
        if lord["val"] == 0:
            continue
        medal = medals[i] if i < 3 else f"{i+1}."
        lines.append(f"{medal} {lord['name']}: +{lord['val']:,}")
    lines.append("```")
    await ctx.send("\n".join(lines))


# In-progress serverupdate sessions: user_id -> True (awaiting file upload)
serverupdate_pending = {}

@bot.command(name="serverupdate")
async def serverupdate(ctx):
    """
    Upload a server stats Excel file to update the leaderboard data.
    Usage: !serverupdate (then attach the .xlsx file in the same or next message)
    Expected filename format: {server_num}_{start_date}_{end_date}.xlsx
    Admin only.
    """
    is_admin = ctx.author.id == OWNER_ID or (
        ctx.guild and ctx.author.guild_permissions.administrator
    )
    if not is_admin:
        return await ctx.send("❌ Only admins can update server data.")

    # Check if file already attached to this message
    if ctx.message.attachments:
        await _process_serverupdate_attachment(ctx, ctx.message.attachments[0])
        return

    serverupdate_pending[ctx.author.id] = ctx.channel.id
    await ctx.send(
        "📎 Please upload the server stats Excel file now (attach it to your next message).\n"
        "Expected filename format: `{server}_{start_date}_{end_date}.xlsx` (e.g. `698_2026-07-02_2026-07-13.xlsx`)\n"
        "Type `cancel` to abort."
    )


async def _process_serverupdate_attachment(ctx, attachment):
    if not attachment.filename.lower().endswith((".xlsx", ".xls")):
        return await ctx.send("❌ Please upload an Excel file (.xlsx or .xls).")

    server_num, start_date, end_date = parse_server_filename(attachment.filename)
    if server_num is None:
        return await ctx.send(
            "❌ Could not detect server number from filename. "
            "Please rename it like `698_2026-07-02_2026-07-13.xlsx` and try again."
        )

    msg = await ctx.send(f"📥 Downloading and parsing `{attachment.filename}`...")

    try:
        file_bytes = await attachment.read()
    except Exception as e:
        return await msg.edit(content=f"❌ Failed to download file: {e}")

    rows, error = parse_server_excel(file_bytes)
    if error or not rows:
        return await msg.edit(content=f"❌ Failed to parse Excel: {error or 'No rows found'}")

    db_set_server_pick(server_num)
    db_replace_server_lord_stats(server_num, rows, start_date, end_date)

    date_str = f" ({start_date} → {end_date})" if start_date and end_date else ""
    await msg.edit(content=(
        f"✅ Server **S#{server_num}** updated{date_str}!\n"
        f"Loaded **{len(rows)} lords** from the Excel file.\n"
        f"Use `!stopmerits`, `!stopdeaths`, `!stopinf`, etc. to view leaderboards."
    ))
    log_info(f"[SERVERUPDATE] S#{server_num}: {len(rows)} lords loaded from {attachment.filename}")


@bot.command(name="stopdeaths")
async def stopdeaths(ctx, server: int = None, top: int = 25):
    """Top deaths on server. Usage: !stopdeaths [server] [top]"""
    await _server_leaderboard(ctx, server, "deaths", "💀", "Deaths", top)

@bot.command(name="stopmerits")
async def stopmerits(ctx, server: int = None, top: int = 25):
    """Top merits on server. Usage: !stopmerits [server] [top]"""
    await _server_leaderboard(ctx, server, "total_merits", "🏅", "Merits", top)

@bot.command(name="stopheal")
async def stopheal(ctx, top: int = 25):
    """Top healing on server. Usage: !stopheal [top]"""
    await _server_leaderboard(ctx, None, "healing", "❤️", "Healing", top)

MANA_PER_T5_HEAL = 78  # real mana cost per T5 unit healed
MANA_PER_T4_HEAL = 20  # real mana cost per T4 unit healed

@bot.command(name="stopmana")
async def stopmana(ctx, top: int = 25):
    """
    Top mana spent on server — exact total, calculated as (T4 Healed × 20) + (T5 Healed × 80),
    using COS's real T4/T5 Healed split and the real per-unit mana costs. Usage: !stopmana [top]
    """
    picked = db_get_server_pick()
    if not picked:
        return await ctx.send("❌ No server picked. Use `!serverupdate` and upload an Excel file first.")

    lords = db_get_server_lord_stats(picked)
    if not lords:
        return await ctx.send(f"❌ No data for S#{picked}. Use `!serverupdate` to upload the Excel file.")

    scored = [
        {"name": l["lord_name"],
         "val": l.get("t4_healed", 0) * MANA_PER_T4_HEAL + l.get("t5_healed", 0) * MANA_PER_T5_HEAL}
        for l in lords
    ]
    scored.sort(key=lambda x: x["val"], reverse=True)
    top_list = scored[:top]

    if not any(x["val"] > 0 for x in top_list):
        return await ctx.send(f"❌ No T4/T5 Healed data found for S#{picked}. Re-upload with `!serverupdate` if this server's Excel is from before the T4/T5 split.")

    date_range = ""
    if lords[0].get("start_date") and lords[0].get("end_date"):
        date_range = f" ({lords[0]['start_date']} → {lords[0]['end_date']})"

    medals = ["🥇", "🥈", "🥉"]
    lines = [f"```💧 Top {top} Mana Spent — S#{picked}{date_range}", ""]
    for i, lord in enumerate(top_list):
        if lord["val"] == 0:
            continue
        medal = medals[i] if i < 3 else f"{i+1}."
        lines.append(f"{medal} {lord['name']}: +{lord['val']:,}")
    lines.append(f"\n*(T4 Healed × {MANA_PER_T4_HEAL}) + (T5 Healed × {MANA_PER_T5_HEAL}) — exact, not an estimate*")
    lines.append("```")
    await ctx.send("\n".join(lines))

@bot.command(name="stopinf")
async def stopinf(ctx, server: int = None, top: int = 25):
    """Top infantry merits on server. Usage: !stopinf [server] [top]"""
    await _server_leaderboard(ctx, server, "infantry_merits", "⚔️", "Infantry Merits", top)

@bot.command(name="stopcav", aliases=["stopcavs"])
async def stopcav(ctx, server: int = None, top: int = 25):
    """Top cavalry merits on server. Usage: !stopcav [server] [top]"""
    await _server_leaderboard(ctx, server, "cavalry_merits", "🐴", "Cavalry Merits", top)

@bot.command(name="stopmage")
async def stopmage(ctx, server: int = None, top: int = 25):
    """Top mage merits on server. Usage: !stopmage [server] [top]"""
    await _server_leaderboard(ctx, server, "mage_merits", "🔮", "Mage Merits", top)

@bot.command(name="stoparcher")
async def stoparcher(ctx, server: int = None, top: int = 25):
    """Top marksman merits on server. Usage: !stoparcher [server] [top]"""
    await _server_leaderboard(ctx, server, "marksman_merits", "🏹", "Marksman Merits", top)

@bot.command(name="stoppower")
async def stoppower(ctx, server: int = None, top: int = 25):
    """Top current power on server. Usage: !stoppower [server] [top]"""
    await _server_leaderboard(ctx, server, "current_power", "⚡", "Current Power", top)

@bot.command(name="stophighest")
async def stophighest(ctx, top: int = 25):
    """Top historical highest power on server. Usage: !stophighest [top]"""
    await _server_leaderboard(ctx, None, "highest_power", "⚡", "Highest Power", top)

@bot.command(name="stopother")
async def stopother(ctx, server: int = None, top: int = 25):
    """Top other merits on server. Usage: !stopother [server] [top]"""
    await _server_leaderboard(ctx, server, "other_merits", "🌀", "Other Merits", top)


# ============================================================
# SAFE LOGIN (NO bot.run)
# ============================================================

async def safe_login():
    token = os.getenv("DISCORD_BOT_TOKEN")
    if not token:
        log_info("❌ Missing DISCORD_BOT_TOKEN")
        return

    while True:
        try:
            await bot.start(token)
            break

        except discord.HTTPException as e:
            if e.status == 429:
                # HARD COOLDOWN on rate limit
                log_info("⛔ Discord rate-limited login. Cooling down for 15 minutes.")
                await asyncio.sleep(900)  # 15 minutes
            else:
                log_info(f"[Login Error] {e}")
                await asyncio.sleep(60)

        except Exception as e:
            log_info(f"[Fatal Login Error] {e}")
            await asyncio.sleep(120)

if __name__ == "__main__":
    asyncio.run(safe_login())

    # KEEP PROCESS ALIVE (REQUIRED FOR KOYEB)
    import time
    while True:
        time.sleep(3600)
